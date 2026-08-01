from fastapi import FastAPI, APIRouter, HTTPException, Query, File, UploadFile, Form, BackgroundTasks, Depends
from fastapi.responses import StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import sys
import csv
import json
from io import StringIO

# v3.1: Modelos para Radar Farmacêutico
class InteresseCreate(BaseModel):
    medicamento: str
    principio_ativo: Optional[str] = None
    categoria: Optional[str] = "Oncologia"
    prioridade: Optional[str] = "media"
    target_type: Optional[str] = "Importacao"

class RadarFarmaMatch(BaseModel):
    id: str
    medicamento: str
    principio_ativo: str
    status_anvisa: str
    data_interrupcao: Optional[str] = None
    previsao_retorno: Optional[str] = None
    link_fonte: Optional[str] = None

# Limite de upload: 15MB para editais PDF
MAX_UPLOAD_SIZE = 15 * 1024 * 1024  # 15MB

# v63.0: Diretório para armazenar timbrados das empresas
UPLOADS_DIR = Path(__file__).parent / "uploads" / "timbrados"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# Adicionar path para importar m\u00f3dulos locais
sys.path.insert(0, str(Path(__file__).parent))

from models.licitacao import Licitacao, LicitacaoCreate, SearchQuery
from models.lista_medicamentos import (
    ListaMedicamentos, 
    ListaMedicamentosCreate, 
    ListaMedicamentosUpdate
)
from models.notificacao import (
    AlertaConfig, AlertaConfigCreate, AlertaConfigUpdate,
    Notificacao, NotificacaoStats, StatusNotificacao,
    NotificacaoListResponse, AlertaListResponse
)
from services.scraper_service import ScraperService
from services.mock_data_service import MockDataService
from services.health_monitor_service import HealthMonitorService
from services.notificacao_service import NotificacaoService
from services.cache_service import search_cache
from services.smart_cache_service import smart_cache
from services.data_enrichment_service import data_enrichment_service
from scheduler import init_scheduler, shutdown_scheduler
from services.sync_service import SyncService, init_sync_service
from services.email_service import get_email_service
from models.user import User
from utils.security import get_current_user, require_super_admin
from routers.auth_router import router as auth_router
from routers.users_router import router as users_router

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="BEM - Buscador Estadual de Medicamentos")
app.state.db = db

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Inicializar serviços
mock_service = MockDataService()
health_monitor = HealthMonitorService(db)
notificacao_service = NotificacaoService(db)
scraper_service = ScraperService(health_monitor=health_monitor)

# Variáveis globais para serviços (inicializadas no startup)
_sync_service_instance = None
_multi_source_sync_instance = None
_normalizador_instance = None
_motor_gsm_instance = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ==================== ROTAS BEM ====================

@api_router.get("/")
async def root():
    return {
        "message": "BEM - Buscador Estadual de Medicamentos API",
        "version": "2.0.0",
        "endpoints": {
            "search": "/api/search",
            "states": "/api/states",
            "refresh": "/api/refresh",
            "stats": "/api/stats",
            "listas": "/api/listas"
        }
    }

# ==================== ROTAS DE LISTAS CUSTOMIZADAS ====================

@api_router.post("/listas", response_model=dict, status_code=201)
async def criar_lista(lista: ListaMedicamentosCreate):
    """
    Cria uma nova lista customizada de medicamentos
    
    Limite: Máximo 5 listas por usuário
    """
    try:
        # TODO: Quando implementar autenticação, usar user_id real
        user_id = "default_user"
        
        # Verificar limite de 5 listas
        count = await db.listas_medicamentos.count_documents({"user_id": user_id})
        if count >= 5:
            raise HTTPException(
                status_code=400, 
                detail="Limite de 5 listas atingido. Delete uma lista existente para criar nova."
            )
        
        # Verificar se já existe lista com mesmo nome
        existing = await db.listas_medicamentos.find_one({
            "user_id": user_id,
            "nome": lista.nome
        })
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Já existe uma lista com o nome '{lista.nome}'"
            )
        
        # Criar nova lista
        nova_lista = ListaMedicamentos(
            **lista.dict(),
            user_id=user_id,
            id=str(uuid.uuid4()),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        # Converter para dict e inserir no MongoDB
        lista_dict = nova_lista.dict()
        lista_dict['created_at'] = lista_dict['created_at'].isoformat()
        lista_dict['updated_at'] = lista_dict['updated_at'].isoformat()
        
        # Criar cópia para retorno (sem _id do MongoDB)
        lista_response = lista_dict.copy()
        
        await db.listas_medicamentos.insert_one(lista_dict)
        
        logger.info(f"Lista criada: {nova_lista.nome} (ID: {nova_lista.id})")
        
        return {
            "message": "Lista criada com sucesso",
            "lista": lista_response
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao criar lista: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao criar lista: {str(e)}")


@api_router.get("/listas", response_model=dict)
async def listar_listas():
    """
    Lista todas as listas customizadas do usuário
    """
    try:
        # TODO: Quando implementar autenticação, usar user_id real
        user_id = "default_user"
        
        listas = await db.listas_medicamentos.find(
            {"user_id": user_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(5)
        
        logger.info(f"Listando {len(listas)} listas")
        
        return {
            "total": len(listas),
            "listas": listas
        }
        
    except Exception as e:
        logger.error(f"Erro ao listar listas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao listar listas: {str(e)}")


@api_router.get("/listas/{lista_id}", response_model=dict)
async def buscar_lista(lista_id: str):
    """
    Busca uma lista específica por ID
    """
    try:
        lista = await db.listas_medicamentos.find_one(
            {"id": lista_id},
            {"_id": 0}
        )
        
        if not lista:
            raise HTTPException(status_code=404, detail="Lista não encontrada")
        
        return {"lista": lista}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao buscar lista: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao buscar lista: {str(e)}")


@api_router.put("/listas/{lista_id}", response_model=dict)
async def atualizar_lista(lista_id: str, lista_update: ListaMedicamentosUpdate):
    """
    Atualiza uma lista existente
    """
    try:
        # Verificar se lista existe
        lista_existente = await db.listas_medicamentos.find_one({"id": lista_id})
        if not lista_existente:
            raise HTTPException(status_code=404, detail="Lista não encontrada")
        
        # Preparar dados para atualização
        update_data = {
            k: v for k, v in lista_update.dict(exclude_unset=True).items() 
            if v is not None
        }
        
        if not update_data:
            raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
        
        # Se está mudando o nome, verificar duplicação
        if 'nome' in update_data:
            user_id = lista_existente.get('user_id', 'default_user')
            duplicata = await db.listas_medicamentos.find_one({
                "user_id": user_id,
                "nome": update_data['nome'],
                "id": {"$ne": lista_id}
            })
            if duplicata:
                raise HTTPException(
                    status_code=400,
                    detail=f"Já existe outra lista com o nome '{update_data['nome']}'"
                )
        
        # Adicionar timestamp de atualização
        update_data['updated_at'] = datetime.utcnow().isoformat()
        
        # Atualizar no MongoDB
        result = await db.listas_medicamentos.update_one(
            {"id": lista_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            logger.warning(f"Lista {lista_id} não foi modificada")
        
        # Buscar lista atualizada
        lista_atualizada = await db.listas_medicamentos.find_one(
            {"id": lista_id},
            {"_id": 0}
        )
        
        logger.info(f"Lista atualizada: {lista_id}")
        
        return {
            "message": "Lista atualizada com sucesso",
            "lista": lista_atualizada
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao atualizar lista: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar lista: {str(e)}")


@api_router.delete("/listas/{lista_id}", response_model=dict)
async def deletar_lista(lista_id: str):
    """
    Deleta uma lista customizada
    """
    try:
        result = await db.listas_medicamentos.delete_one({"id": lista_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Lista não encontrada")
        
        logger.info(f"Lista deletada: {lista_id}")
        
        return {
            "message": "Lista deletada com sucesso",
            "lista_id": lista_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao deletar lista: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao deletar lista: {str(e)}")

# Campos essenciais para projeção (P1 - Otimização)
PROJECTION_FIELDS = {
    "_id": 0,
    "id": 1,
    "medicamento": 1,
    "estado": 1,
    "status": 1,
    "orgao_licitante": 1,
    "modalidade": 1,
    "numero_processo": 1,
    "data_final": 1,
    "data_abertura": 1,
    "link_origem": 1,
    "link_documento": 1,
    "fonte": 1,
    "esfera": 1,
    "objeto": 1
}

@api_router.post("/search")
async def search_medicamentos(query: SearchQuery):
    """Busca medicamentos nas bases estaduais ou por lista customizada"""
    try:
        # 🗄️ CACHE: Verificar se resultado está em cache
        cache_params = {
            'medicamento': query.medicamento,
            'tags': str(query.tags) if query.tags else None,
            'lista_id': query.lista_id,
            'apenas_reais': query.apenas_reais,
            'apenas_futuras': query.apenas_futuras,
            'status_filtro': query.status_filtro,
            'modalidade_filtro': str(query.modalidade_filtro) if query.modalidade_filtro else None,
            'esfera_filtro': query.esfera_filtro,
            'page': query.page,
            'per_page': query.per_page
        }
        
        cached_result = search_cache.get(**cache_params)
        if cached_result is not None:
            logger.info("✅ Retornando resultado do cache")
            return cached_result
        
        logger.info("❌ Cache MISS - Executando busca completa")
        
        # NOVO: Se tem lista_id, buscar medicamentos da lista
        medicamentos_lista = []
        if query.lista_id:
            logger.info(f"Buscando por lista customizada: {query.lista_id}")
            lista = await db.listas_medicamentos.find_one(
                {"id": query.lista_id},
                {"_id": 0}
            )
            if not lista:
                raise HTTPException(status_code=404, detail="Lista não encontrada")
            
            medicamentos_lista = lista.get('medicamentos', [])
            logger.info(f"Lista '{lista.get('nome')}' contém {len(medicamentos_lista)} medicamentos")
            
            if not medicamentos_lista:
                return {
                    "total": 0,
                    "medicamento": f"Lista: {lista.get('nome')} (vazia)",
                    "resultados": []
                }
        
        # Busca por medicamento específico OU por tags OU por lista
        if query.medicamento:
            logger.info(f"Buscando medicamento: {query.medicamento}")
        if query.tags:
            logger.info(f"Buscando por tags: {query.tags}")
        
        resultados = []
        
        # NOVO: Se tem lista, buscar todos os medicamentos da lista
        if medicamentos_lista:
            for medicamento in medicamentos_lista:
                logger.info(f"Buscando medicamento da lista: {medicamento}")
                # Buscar nos scrapers reais (PNCP, ComprasNet, BEC SP, Estados)
                # OTIMIZADO: Passar apenas_futuras para prospecção (P0)
                dados_reais = await scraper_service.buscar_medicamento(
                    medicamento,
                    apenas_futuras=query.apenas_futuras
                )
                resultados.extend(dados_reais)
        
        # Se tem medicamento, buscar normalmente
        elif query.medicamento:
            # Buscar nos scrapers reais (PNCP, ComprasNet, BEC SP, Estados)
            # OTIMIZADO: Passar apenas_futuras para prospecção (P0)
            dados_reais = await scraper_service.buscar_medicamento(
                query.medicamento,
                apenas_futuras=query.apenas_futuras
            )
            resultados.extend(dados_reais)
        
        # Se busca APENAS por tags (sem medicamento), buscar no banco de dados
        elif query.tags:
            # Buscar no MongoDB licitações que tenham as tags
            query_db = {
                'tags': {'$in': query.tags}
            }
            
            if query.apenas_reais:
                query_db['is_mock'] = False
            
            # P1: Aplicar projeção para otimizar
            licitacoes_db = await db.licitacoes.find(query_db, PROJECTION_FIELDS).to_list(1000)
            
            # Converter timestamps de volta
            for lic in licitacoes_db:
                if isinstance(lic.get('data_referencia'), str):
                    lic['data_referencia'] = datetime.fromisoformat(lic['data_referencia'])
                if isinstance(lic.get('created_at'), str):
                    lic['created_at'] = datetime.fromisoformat(lic['created_at'])
                if isinstance(lic.get('updated_at'), str):
                    lic['updated_at'] = datetime.fromisoformat(lic['updated_at'])
            
            resultados = licitacoes_db
            
            logger.info(f"Encontrados {len(resultados)} resultados no banco para tags: {query.tags}")
        
        # ========== ENRIQUECIMENTO INTELIGENTE DE DADOS ==========
        # Aplica inteligência de negócios: tags de saúde, score de relevância
        logger.info(f"🔬 Enriquecendo {len(resultados)} resultados com tags e scores...")
        resultados = data_enrichment_service.enriquecer_lote(resultados)
        logger.info(f"✅ Enriquecimento concluído")
        
        # ========== APLICAR FILTROS AVANÇADOS ==========
        
        # Filtro 1: Estado
        if query.estado:
            resultados = [r for r in resultados if r.get('estado') == query.estado]
            logger.info(f"Após filtro de estado ({query.estado}): {len(resultados)} resultados")
        
        # Filtro 2: Tags
        if query.tags and query.medicamento:
            resultados = [r for r in resultados if any(tag in r.get('tags', []) for tag in query.tags)]
            logger.info(f"Após filtro de tags ({query.tags}): {len(resultados)} resultados")
        
        # Filtro 3: Status (Ativa/Encerrada) - NOVO
        if query.status_filtro and query.status_filtro != 'Todas':
            resultados = [r for r in resultados if r.get('status') == query.status_filtro]
            logger.info(f"Após filtro de status ({query.status_filtro}): {len(resultados)} resultados")
        
        # Filtro 4: Modalidade - NOVO
        if query.modalidade_filtro:
            resultados = [r for r in resultados if r.get('modalidade') in query.modalidade_filtro]
            logger.info(f"Após filtro de modalidade ({query.modalidade_filtro}): {len(resultados)} resultados")
        
        # Filtro 5: Esfera (Federal/Estadual/Municipal) - NOVO
        if query.esfera_filtro:
            resultados = [r for r in resultados if r.get('esfera') == query.esfera_filtro]
            logger.info(f"Após filtro de esfera ({query.esfera_filtro}): {len(resultados)} resultados")
        
        # Filtro 6: Apenas licitações futuras
        if query.apenas_futuras:
            def is_futura(r):
                # Verificar status
                if r.get('status') in ['FUTURA', 'Ativa']:
                    return True
                
                # Verificar data_final (prioritária)
                data_final = r.get('data_final')
                if data_final:
                    if isinstance(data_final, str):
                        try:
                            data_final = datetime.fromisoformat(data_final.replace('Z', ''))
                        except (ValueError, AttributeError):
                            pass
                    if isinstance(data_final, datetime):
                        return data_final > datetime.now()
                
                # Fallback: verificar data_abertura
                data_ab = r.get('data_abertura')
                if data_ab:
                    if isinstance(data_ab, str):
                        try:
                            data_ab = datetime.fromisoformat(data_ab.replace('Z', ''))
                        except (ValueError, AttributeError):
                            return False
                    if isinstance(data_ab, datetime):
                        return data_ab > datetime.now()
                
                return False
            
            resultados = [r for r in resultados if is_futura(r)]
            logger.info(f"Após filtro de futuras: {len(resultados)} resultados")
        
        # Filtro 7: Período de data limite - NOVO
        if query.data_limite_inicio or query.data_limite_fim:
            def dentro_periodo(r):
                data_final = r.get('data_final')
                if not data_final:
                    return False
                
                if isinstance(data_final, str):
                    try:
                        data_final = datetime.fromisoformat(data_final.replace('Z', ''))
                    except (ValueError, AttributeError):
                        return False
                
                if query.data_limite_inicio and data_final < query.data_limite_inicio:
                    return False
                if query.data_limite_fim and data_final > query.data_limite_fim:
                    return False
                
                return True
            
            resultados = [r for r in resultados if dentro_periodo(r)]
            logger.info(f"Após filtro de período: {len(resultados)} resultados")
        
        # ========== FILTROS DE INTELIGÊNCIA DE NEGÓCIOS (SAÚDE) ==========
        
        # Filtro 8: Apenas Saúde
        if query.apenas_saude:
            resultados = [r for r in resultados if r.get('is_saude', False)]
            logger.info(f"Após filtro apenas_saude: {len(resultados)} resultados")
        
        # Filtro 9: Apenas Urgentes
        if query.apenas_urgentes:
            resultados = [r for r in resultados if r.get('is_urgente', False)]
            logger.info(f"Após filtro apenas_urgentes: {len(resultados)} resultados")
        
        # Filtro 10: Categorias de Saúde específicas
        if query.categorias_saude and len(query.categorias_saude) > 0:
            # Mapa de categorias para tags do backend
            CATEGORIA_TAG_MAP = {
                'hospitalar': '🏥 Hospitalar',
                'medicamentos': '💊 Medicamentos',
                'equipamentos': '🩺 Equipamentos Médicos',
                'laboratorio': '🧪 Laboratório',
                'insumos': '💉 Insumos Médicos',
                'odontologia': '🦷 Odontologia',
                'oftalmologia': '👁️ Oftalmologia',
                'oncologia': '🩻 Oncologia',
                'cardiologia': '🫀 Cardiologia',
                'especialidades': '🧬 Especialidades',
                'servicos': '👨‍⚕️ Serviços de Saúde',
                'saude_geral': '🩹 Saúde Geral',
            }
            
            def tem_categoria(r):
                tags_saude = r.get('tags_saude', [])
                if not tags_saude:
                    return False
                
                for cat_id in query.categorias_saude:
                    tag_esperada = CATEGORIA_TAG_MAP.get(cat_id, '')
                    if tag_esperada and tag_esperada in tags_saude:
                        return True
                return False
            
            resultados = [r for r in resultados if tem_categoria(r)]
            logger.info(f"Após filtro categorias_saude {query.categorias_saude}: {len(resultados)} resultados")
        
        # ========== CÁLCULO DE IMINÊNCIA (DIAS ATÉ O PRAZO) ==========
        # Inspirado no GSM - campo muito útil para priorização
        
        def calcular_iminencia(r):
            """
            Calcula dias até o prazo (data_final ou data_abertura)
            - Valores positivos: dias restantes
            - 0: hoje é o último dia
            - Valores negativos: prazo já passou
            """
            hoje = datetime.now()
            
            # Prioridade: data_final > data_limite > data_abertura
            for campo in ['data_final', 'data_limite', 'data_abertura']:
                data = r.get(campo)
                if data:
                    if isinstance(data, str):
                        try:
                            data = datetime.fromisoformat(data.replace('Z', ''))
                        except (ValueError, AttributeError):
                            continue
                    
                    if isinstance(data, datetime):
                        dias = (data - hoje).days
                        return dias
            
            return None  # Sem data disponível
        
        # Adicionar iminencia a cada resultado
        for r in resultados:
            r['iminencia'] = calcular_iminencia(r)
        
        # ========== ORDENAÇÃO (PRIORIDADE POR URGÊNCIA) ==========
        
        def sort_key(r):
            """
            Ordenação hierárquica (OTIMIZADA - baseada no GSM):
            1. Iminência (mais urgente primeiro - menos dias)
            2. Status Aberto/Agendado > Em Andamento > Encerrado
            3. Fonte (PNCP, ComprasNet primeiro)
            """
            # PRIORIDADE 1: Iminência (urgência)
            iminencia = r.get('iminencia')
            if iminencia is not None and iminencia >= 0:
                # Processos futuros/ativos: ordenar por proximidade
                prioridade_iminencia = iminencia
            elif iminencia is not None and iminencia < 0:
                # Processos passados: baixa prioridade
                prioridade_iminencia = 9999
            else:
                # Sem data: média prioridade
                prioridade_iminencia = 500
            
            # PRIORIDADE 2: Status
            status = r.get('status', '').lower()
            if any(s in status for s in ['aberto', 'agendado', 'publicado', 'proposta']):
                prioridade_status = 0  # Maior prioridade
            elif any(s in status for s in ['andamento', 'ativo', 'ativa', 'futura']):
                prioridade_status = 1
            elif any(s in status for s in ['encerrad', 'concluíd', 'homolog']):
                prioridade_status = 3
            else:
                prioridade_status = 2
            
            # PRIORIDADE 3: Fonte
            fonte = r.get('fonte', '')
            if fonte == 'PNCP':
                prioridade_fonte = 0
            elif fonte == 'ComprasNet':
                prioridade_fonte = 1
            elif 'CSV' in fonte:
                prioridade_fonte = 2
            else:
                prioridade_fonte = 3
            
            # Retornar tupla de ordenação
            return (prioridade_status, prioridade_iminencia, prioridade_fonte)
        
        resultados = sorted(resultados, key=sort_key)
        logger.info("Resultados ordenados por urgência")
        
        # Salvar no MongoDB apenas se for busca por medicamento
        if query.medicamento:
            for resultado in resultados:
                # Verificar se já existe
                existing = await db.licitacoes.find_one({
                    'medicamento': resultado['medicamento'],
                    'estado': resultado['estado'],
                    'numero_processo': resultado['numero_processo']
                })
                
                if not existing:
                    # Converter datetime para ISO string para MongoDB
                    doc = resultado.copy()
                    doc['data_referencia'] = doc['data_referencia'].isoformat()
                    if doc.get('data_abertura'):
                        doc['data_abertura'] = doc['data_abertura'].isoformat()
                    doc['created_at'] = datetime.now().isoformat()
                    doc['updated_at'] = datetime.now().isoformat()
                    doc['id'] = str(uuid.uuid4())
                    await db.licitacoes.insert_one(doc)
        
        total_resultados = len(resultados)
        logger.info(f"Encontrados {total_resultados} resultados totais")
        
        # ========== P1: APLICAR PAGINAÇÃO ==========
        # Calcular índices de paginação
        skip = (query.page - 1) * query.per_page
        limit = query.per_page
        
        # Aplicar paginação nos resultados
        resultados_paginados = resultados[skip:skip + limit]
        
        # Calcular metadados de paginação
        total_pages = (total_resultados + query.per_page - 1) // query.per_page
        has_next = query.page < total_pages
        has_prev = query.page > 1
        
        logger.info(f"📄 Paginação: página {query.page}/{total_pages} ({len(resultados_paginados)} resultados)")
        
        # Preparar mensagem de resposta
        if query.lista_id and medicamentos_lista:
            lista_info = await db.listas_medicamentos.find_one({"id": query.lista_id}, {"_id": 0})
            mensagem = f"Lista: {lista_info.get('nome')} ({len(medicamentos_lista)} medicamentos)"
        elif query.medicamento:
            mensagem = query.medicamento
        elif query.tags:
            mensagem = f"Busca por tags: {', '.join(query.tags)}"
        else:
            mensagem = "Busca geral"
        
        # Preparar resposta com paginação
        response = {
            "total": total_resultados,
            "medicamento": mensagem,
            "resultados": resultados_paginados,
            # Metadados de paginação
            "pagination": {
                "page": query.page,
                "per_page": query.per_page,
                "total_pages": total_pages,
                "total_items": total_resultados,
                "has_next": has_next,
                "has_prev": has_prev
            }
        }
        
        # 💾 CACHE: Armazenar resultado em cache
        search_cache.set(response, **cache_params)
        logger.info("💾 Resultado armazenado em cache")
        
        return response
        
    except Exception as e:
        logger.error(f"Erro na busca: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao buscar medicamento: {str(e)}")

@api_router.get("/states")
async def get_states():
    """Retorna lista de estados com indica\u00e7\u00e3o de quais t\u00eam scraping real"""
    estados_reais = ['CE', 'ES', 'SP']
    estados_mock = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'DF', 'GO', 'MA', 'MT', 'MS', 
        'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 
        'RR', 'SC', 'SE', 'TO'
    ]
    
    estados = []
    for uf in estados_reais:
        estados.append({
            'uf': uf,
            'nome': _get_estado_nome(uf),
            'has_scraping': True,
            'is_mock': False
        })
    
    for uf in estados_mock:
        estados.append({
            'uf': uf,
            'nome': _get_estado_nome(uf),
            'has_scraping': False,
            'is_mock': True
        })
    
    return {"estados": sorted(estados, key=lambda x: x['uf'])}

@api_router.post("/refresh/{estado}")
async def refresh_estado(estado: str, medicamento: str = Query(...)):
    """For\u00e7a refresh de dados de um estado espec\u00edfico"""
    try:
        if estado not in ['CE', 'ES', 'SP']:
            raise HTTPException(status_code=400, detail="Refresh dispon\u00edvel apenas para CE, ES e SP")
        
        logger.info(f"Refresh manual para {estado} - medicamento: {medicamento}")
        
        dados = await scraper_service.refresh_estado(estado, medicamento)
        
        return {
            "estado": estado,
            "medicamento": medicamento,
            "total": len(dados),
            "resultados": dados
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro no refresh: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar dados: {str(e)}")

@api_router.get("/stats")
async def get_stats():
    """Retorna estat\u00edsticas gerais do sistema"""
    try:
        total_licitacoes = await db.licitacoes.count_documents({})
        total_reais = await db.licitacoes.count_documents({'is_mock': False})
        total_mock = await db.licitacoes.count_documents({'is_mock': True})
        
        # Contagem por estado
        pipeline = [
            {'$group': {'_id': '$estado', 'count': {'$sum': 1}}},
            {'$sort': {'count': -1}}
        ]
        por_estado = await db.licitacoes.aggregate(pipeline).to_list(27)
        
        return {
            "total_licitacoes": total_licitacoes,
            "licitacoes_reais": total_reais,
            "licitacoes_mock": total_mock,
            "por_estado": [{'estado': item['_id'], 'total': item['count']} for item in por_estado],
            "estados_com_scraping": ['CE', 'ES', 'SP']
        }
        
    except Exception as e:
        logger.error(f"Erro ao buscar estat\u00edsticas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao buscar estat\u00edsticas: {str(e)}")



# ==================== ROTAS RADAR FARMACÊUTICO (v3.1) ====================

@api_router.get("/radar-farma/stats")
async def get_radar_farma_stats():
    """Retorna estatísticas do Radar Farmacêutico"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        return await svc.estatisticas()
    except Exception as e:
        logger.error(f"Erro em radar-farma/stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/radar-farma/interesse")
async def listar_interesse_radar():
    """Lista medicamentos na lista de interesse estratégica"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        return await svc.listar_interesse()
    except Exception as e:
        logger.error(f"Erro em radar-farma/interesse: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/radar-farma/interesse")
async def adicionar_interesse_radar(item: InteresseCreate):
    """Adiciona medicamento à lista de interesse estratégica"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        return await svc.adicionar_interesse(item.dict())
    except Exception as e:
        logger.error(f"Erro ao adicionar interesse: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/radar-farma/interesse/{item_id}")
async def remover_interesse_radar(item_id: str):
    """Remove medicamento da lista de interesse"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        success = await svc.remover_interesse(item_id)
        if not success:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao remover interesse: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/radar-farma/matches")
async def listar_matches_radar(limite: int = 50):
    """Lista desabastecimentos detectados cruzados com a lista de interesse"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        return await svc.listar_desabastecimento(limite=limite)
    except Exception as e:
        logger.error(f"Erro em radar-farma/matches: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/radar-farma/scan")
async def disparar_scan_radar(background_tasks: BackgroundTasks):
    """Dispara varredura manual do Radar Farmacêutico"""
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        
        # Como o scan do DOU pode ser lento, rodar em background
        background_tasks.add_task(svc.executar_scan)
        
        return {"message": "Varredura iniciada em segundo plano"}
    except Exception as e:
        logger.error(f"Erro ao iniciar scan radar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/status/scrapers")
async def get_scrapers_status():
    """
    Retorna status de saúde de todos os scrapers
    
    Métricas incluem:
    - Status geral (UP/DOWN/DEGRADED)
    - Última execução com sucesso
    - Total de execuções nas últimas 24h
    - Taxa de sucesso
    - Total de resultados obtidos
    """
    try:
        system_health = await health_monitor.get_system_health()
        return system_health.dict()
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar status dos scrapers: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Erro ao buscar status: {str(e)}"
        )


@api_router.get("/cache/stats")
async def get_cache_stats():
    """
    Retorna estatísticas do cache de buscas (10min + 24h smart cache)
    """
    try:
        stats_search = search_cache.get_stats()
        stats_smart = smart_cache.get_stats()
        return {
            "cache_stats": stats_search,
            "smart_cache_stats": stats_smart,
            "message": "Estatísticas do cache de buscas (search 10min + smart 24h)"
        }
    except Exception as e:
        logger.error(f"Erro ao buscar stats do cache: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/cache/clear")
async def clear_cache():
    """
    Limpa todo o cache de buscas (10min + 24h smart cache)
    
    Útil para forçar atualização dos dados
    """
    try:
        search_cache.clear()
        removed = smart_cache.clear()
        return {"message": f"Cache limpo com sucesso (search + smart: {removed} entradas)"}
    except Exception as e:
        logger.error(f"Erro ao limpar cache: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== NOTIFICACOES DE OPORTUNIDADE (P4) ====================

@api_router.get("/notificacoes/oportunidades")
async def listar_alertas_oportunidade(limite: int = Query(15, ge=1, le=50)):
    """
    Lista alertas de oportunidade (score >= 80%) salvos no MongoDB.
    Ordenados do mais recente ao mais antigo.
    """
    try:
        alertas = await db.oportunidades_alertas.find(
            {}, {"_id": 0}
        ).sort("criado_em", -1).to_list(length=limite)
        nao_lidas = sum(1 for a in alertas if not a.get('lida'))
        return {
            "alertas": alertas,
            "total": len(alertas),
            "nao_lidas": nao_lidas,
        }
    except Exception as e:
        logger.error(f"Erro ao listar alertas oportunidade: {e}")
        return {"alertas": [], "total": 0, "nao_lidas": 0}


@api_router.post("/notificacoes/oportunidades/{alerta_id}/lida")
async def marcar_alerta_lida(alerta_id: str):
    """Marca um alerta de oportunidade como lido."""
    try:
        result = await db.oportunidades_alertas.update_one(
            {"id": alerta_id},
            {"$set": {"lida": True}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Alerta nao encontrado")
        return {"message": "Alerta marcado como lido", "id": alerta_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao marcar alerta: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.get("/notificacoes/oportunidades/{alerta_id}")
async def buscar_alerta_oportunidade(alerta_id: str):
    """
    Busca um alerta de oportunidade por ID.
    Retorna dados completos + analise LMR atualizada.
    Usado pelo frontend quando usuario clica no link do email.
    """
    try:
        alerta = await db.oportunidades_alertas.find_one(
            {"id": alerta_id}, {"_id": 0}
        )
        if not alerta:
            raise HTTPException(status_code=404, detail="Alerta nao encontrado")

        # Buscar analise LMR completa para o medicamento
        analise_lmr = None
        med = alerta.get('medicamento', '')
        if med:
            try:
                from services.lmr_service import get_lmr_service
                svc = get_lmr_service(db)
                analise_lmr = await svc.analisar_medicamento(med)
            except Exception as e:
                logger.warning(f"Analise LMR fallback para alerta {alerta_id}: {e}")

        # Marcar como lido automaticamente
        await db.oportunidades_alertas.update_one(
            {"id": alerta_id}, {"$set": {"lida": True}}
        )

        return {
            "alerta": alerta,
            "analise_lmr": analise_lmr,
            "pdf_url": f"/api/dama/prova-documental-lmr/{alerta_id}",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao buscar alerta {alerta_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== ROTAS LOCAL-FIRST (P0) ====================

@api_router.get("/search/local")
async def search_local(
    q: str = Query(None, description="Termo de busca digitado pelo usuário"),
    keywords: str = Query(None, description="Palavras-chave das listas do usuário (separadas por vírgula)"),
    lista_id: str = Query(None, description="ID da lista de medicamentos para busca híbrida"),
    estados: str = Query(None, description="Estados separados por vírgula (ex: SP,RJ,MG)"),
    municipio: str = Query(None, description="🏙️ Filtro por município/cidade (v3.6)"),
    modalidade: str = Query(None, description="Filtro por modalidade (Pregão, Concorrência, etc)"),
    esfera: str = Query(None, description="Filtro por esfera (Federal, Estadual, Municipal)"),
    apenas_saude: bool = Query(False, description="Filtrar apenas editais de saúde"),
    incluir_historico: bool = Query(False, description="Incluir processos antigos (histórico)"),
    periodo_dias: int = Query(90, ge=7, le=365, description="Período em dias para filtro temporal"),
    # 🎯 CLASSIFICAÇÃO DE OPORTUNIDADES V3 (PADRÃO GSM)
    incluir_ativas: bool = Query(True, description="Incluir oportunidades ATIVAS (DEFAULT: True)"),
    incluir_futuras: bool = Query(False, description="Incluir oportunidades FUTURAS (DEFAULT: False)"),
    incluir_encerradas: bool = Query(False, description="Incluir oportunidades ENCERRADAS (DEFAULT: False)"),
    excluir_credenciamentos: bool = Query(False, description="Excluir credenciamentos do resultado (DEFAULT: False = inclui)"),
    # 🔒 P3: CAMADA DE CONFIABILIDADE DE DADOS
    incluir_suspeitos: bool = Query(False, description="Incluir DATA_SUSPEITA no resultado (DEFAULT: False)"),
    incluir_planejamento: bool = Query(False, description="Incluir PLANEJAMENTO_LONGO no resultado (DEFAULT: False)"),
    limite_quality_score: int = Query(70, ge=0, le=100, description="Score mínimo de qualidade (DEFAULT: 70)"),
    smart_search: bool = Query(False, description="🚀 NOVO v65.0: Acionar busca inteligente (plural, gênero, etc.)"),
    limit: int = Query(50, ge=1, le=200, description="Máximo de resultados"),
    page: int = Query(1, ge=1, description="Número da página")
):
    """
    🚀 BUSCA LOCAL-FIRST COM CLASSIFICAÇÃO V3 (PADRÃO GSM)
    
    🎯 CLASSIFICAÇÃO V3 - ATIVA = OPORTUNIDADE ACIONÁVEL AGORA:
    
    - 🟢 ATIVA: Oportunidade que permite participação/adesão AGORA
      - Inclui competitivas (abertura <= 90 dias)
      - Inclui credenciamentos vigentes (principal fonte no mercado de saúde!)
    - 🟡 FUTURA: Publicada mas ainda não permite participação
    - 🔴 ENCERRADA: Prazo passou ou cancelada
    
    🔵 CREDENCIAMENTOS:
    - Vigentes = ATIVA (com badge diferenciado)
    - Incluídos no default (como GSM!)
    - Checkbox opcional para excluir se desejado
    
    🔒 COMPORTAMENTO DEFAULT (PADRÃO GSM):
    - Retorna TODAS oportunidades ATIVAS (competitivas + credenciamentos!)
    - Nunca retorna vazio para termos comuns de saúde
    
    Args:
        q: Termo digitado
        keywords: Palavras-chave separadas por vírgula
        lista_id: ID de lista de medicamentos
        estados: UFs separadas por vírgula
        modalidade: Tipo de modalidade
        esfera: Federal, Estadual ou Municipal
        incluir_historico: Incluir processos antigos
        periodo_dias: Período temporal em dias
        incluir_ativas: Incluir ATIVAS (DEFAULT: True)
        incluir_futuras: Incluir FUTURAS (DEFAULT: False)
        incluir_encerradas: Incluir ENCERRADAS (DEFAULT: False)
        excluir_credenciamentos: Excluir credenciamentos (DEFAULT: False = inclui!)
        apenas_saude: Filtrar apenas saúde
        limit: Máximo de resultados
        page: Página
        
    Returns:
        Lista de editais classificados com status_oportunidade
    """
    import time
    inicio = time.time()
    
    try:
        if _sync_service_instance is None:
            raise HTTPException(
                status_code=503, 
                detail="SyncService não inicializado. Aguarde a sincronização inicial."
            )
        
        # Converter string de estados para lista
        lista_estados = None
        if estados:
            lista_estados = [e.strip().upper() for e in estados.split(',')]
        
        # BUSCA HÍBRIDA: Processar keywords da lista
        lista_keywords = []
        lista_nome = None
        
        # Se passou keywords diretamente
        if keywords:
            lista_keywords = [k.strip() for k in keywords.split(',') if k.strip()]
        
        # Se passou lista_id, buscar medicamentos da lista
        if lista_id:
            lista_doc = await db.listas_medicamentos.find_one(
                {"id": lista_id},
                {"medicamentos": 1, "nome": 1, "_id": 0}
            )
            if lista_doc:
                lista_keywords.extend(lista_doc.get('medicamentos', []))
                lista_nome = lista_doc.get('nome')
        
        # Calcular skip para paginação
        skip = (page - 1) * limit
        
        # BUSCA V2: Alta recall com expansão de termos, FILTRO TEMPORAL e CLASSIFICAÇÃO
        from services.busca_service_v2 import get_busca_service_v2
        busca_service = get_busca_service_v2(db)
        
        resultado = await busca_service.buscar(
            termo_busca=q,
            keywords=lista_keywords if lista_keywords else None,
            estados=lista_estados,
            municipio=municipio,  # 🏙️ NOVO v3.6
            modalidade=modalidade,
            esfera=esfera,
            apenas_saude=apenas_saude,
            limit=limit,
            skip=skip,
            expandir_termos=True,
            smart_search=smart_search,
            incluir_historico=incluir_historico,  # 🔒 FILTRO TEMPORAL
            periodo_dias=periodo_dias,
            # 🎯 CLASSIFICAÇÃO DE OPORTUNIDADES V3 (PADRÃO GSM)
            incluir_ativas=incluir_ativas,
            incluir_futuras=incluir_futuras,
            incluir_encerradas=incluir_encerradas,
            excluir_credenciamentos=excluir_credenciamentos,  # V3: Excluir se solicitado
            # 🔒 P3: CAMADA DE CONFIABILIDADE DE DADOS
            incluir_suspeitos=incluir_suspeitos,
            incluir_planejamento=incluir_planejamento,
            limite_quality_score=limite_quality_score
        )
        
        tempo_total_ms = (time.time() - inicio) * 1000
        
        # Calcular metadados de paginação
        total = resultado.get('total', 0)
        total_pages = (total + limit - 1) // limit if total > 0 else 1
        
        # Construir filtros ativos para resposta
        filtros_ativos = {}
        if q: filtros_ativos['termo'] = q
        if lista_keywords: filtros_ativos['keywords'] = lista_keywords
        if lista_nome: filtros_ativos['lista_nome'] = lista_nome
        if estados: filtros_ativos['estados'] = lista_estados
        if municipio: filtros_ativos['municipio'] = municipio  # 🏙️ v3.6
        if modalidade: filtros_ativos['modalidade'] = modalidade
        if esfera: filtros_ativos['esfera'] = esfera
        if apenas_saude: filtros_ativos['apenas_saude'] = True
        filtros_ativos['incluir_historico'] = incluir_historico
        filtros_ativos['periodo_dias'] = periodo_dias
        # 🎯 Filtros de classificação V3
        filtros_ativos['incluir_ativas'] = incluir_ativas
        filtros_ativos['incluir_futuras'] = incluir_futuras
        filtros_ativos['incluir_encerradas'] = incluir_encerradas
        filtros_ativos['excluir_credenciamentos'] = excluir_credenciamentos
        # 🔒 Filtros P3 - Confiabilidade
        filtros_ativos['incluir_suspeitos'] = incluir_suspeitos
        filtros_ativos['incluir_planejamento'] = incluir_planejamento
        filtros_ativos['limite_quality_score'] = limite_quality_score
        
        # Indicador de busca híbrida
        busca_hibrida = bool(q and lista_keywords)
        
        modo_busca = "HÍBRIDA" if busca_hibrida else ("EXPANDIDA" if lista_keywords else "LOCAL-FIRST")
        logger.info(f"🔍 [{modo_busca}] Busca com {len(filtros_ativos)} filtros → {len(resultado.get('resultados', []))}/{total} resultados em {tempo_total_ms:.1f}ms")
        
        return {
            "termo": q or "",
            "total": total,
            "resultados": resultado.get('resultados', []),
            "origem": resultado.get('origem', 'BuscaService V2'),
            "filtros_ativos": filtros_ativos,
            "busca_hibrida": {
                "ativa": busca_hibrida or bool(lista_keywords),
                "termo_digitado": q,
                "keywords_lista": lista_keywords,
                "lista_nome": lista_nome,
                "termos_combinados": resultado.get('termos_combinados', [])
            },
            "expansao_termos": {
                "termos_originais": resultado.get('termos_originais', []),
                "termos_expandidos": resultado.get('termos_expandidos', []),
                "fontes_consultadas": resultado.get('fontes_consultadas', [])
            },
            # 🎯 NOVA SEÇÃO - CLASSIFICAÇÃO DE OPORTUNIDADES (PADRÃO GSM)
            "classificacao_oportunidade": {
                "contagem_status": resultado.get('contagem_status', {}),
                "contagem_pre_filtro": resultado.get('contagem_pre_filtro', {}),
                "filtros_aplicados": resultado.get('filtros_aplicados', {})
            },
            # 🔒 P3: CONFIABILIDADE DE DADOS
            "confiabilidade_dados": {
                "auditoria": resultado.get('auditoria', {}),
                "qualidade": resultado.get('qualidade', {}),
                "limite_quality_score": limite_quality_score
            },
            "pagination": {
                "page": page,
                "per_page": limit,
                "total_pages": total_pages,
                "total_items": total,
                "has_next": page < total_pages,
                "has_prev": page > 1
            },
            "performance": {
                "tempo_ms": round(tempo_total_ms, 2),
                "tempo_query_ms": round(resultado.get('tempo_ms', 0), 2),
                "tempo_p3_ms": round(resultado.get('tempo_p3_ms', 0), 2),
                "fonte": "MongoDB Local (P0 + P1 + P2 + P3)"
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na busca local: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro na busca local: {str(e)}")



@api_router.get("/search/unified")
async def search_unified(
    q: str = Query(None, description="Termo de busca"),
    municipio: str = Query(None, description="Filtro por município"),
    uf: str = Query(None, description="Filtro por UF"),
    estados: str = Query(None, description="Filtro por estados (alias UF)"),
    apenas_saude: bool = Query(False, description="Filtrar apenas saúde"),
    modalidade: str = Query(None, description="Filtro por modalidade"),
    limit: int = Query(50, ge=1, le=2000, description="Máximo de resultados"),
    page: int = Query(1, ge=1, description="Página")
):
    """
    v78.0: Busca 100% INDEPENDENTE — PNCP + Compras.gov.br.
    Zero dependência de terceiros. PDFs direto do governo.
    """
    import time
    inicio = time.time()
    
    try:
        tem_localizacao = bool((municipio and municipio.strip()) or (uf and uf.strip()) or (estados and estados.strip()))
        if not tem_localizacao and (not q or len(q.strip()) < 2):
            raise HTTPException(status_code=400, detail="Informe um termo de busca (mín. 2 caracteres) ou um filtro de Município/Estado")

        from services.motor_independente import MotorBuscaIndependente
        motor = MotorBuscaIndependente(db=db)

        resultado = await motor.buscar(
            termo=(q or '').strip(),
            pagina=page,
            uf=uf,
            estados=estados,
            municipio=municipio,
            modalidade=modalidade,
            limit=limit
        )
        
        tempo_ms = (time.time() - inicio) * 1000
        
        total = resultado.get('total', 0)
        total_pages = (total + limit - 1) // limit if total > 0 else 1
        
        return {
            "termo": q,
            "total": total,
            "resultados": resultado.get('resultados', []),
            "fontes": resultado.get('fontes', {}),
            "fonte_disponivel": resultado.get('fonte_disponivel', True),
            "aviso": resultado.get('aviso'),
            "pagination": {
                "page": page,
                "per_page": limit,
                "total_items": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            },
            "performance": {
                "tempo_ms": round(tempo_ms, 2),
                "fonte": "PNCP + Compras.gov.br (100% independente)"
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro busca unificada: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== COMPRAS.GOV.BR API (v1.0) ====================

@api_router.get("/compras-gov/search")
async def search_compras_gov(
    q: str = Query(..., description="Termo de busca no objeto"),
    dias: int = Query(30, description="Dias para trás para busca"),
    uf: str = Query(None, description="Filtrar por UF")
):
    """Busca avançada na API de Dados Abertos do Compras.gov.br."""
    try:
        from services.comprasgov_service import get_comprasgov_service
        service = get_comprasgov_service(db=db)
        return await service.buscar_contratacoes_por_objeto(termo=q, dias_atras=dias, uf=uf)
    except Exception as e:
        logger.error(f"Erro busca ComprasGov: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/compras-gov/items/{id_compra}")
async def get_compras_gov_items(id_compra: str):
    """Busca detalhes dos itens de uma contratação PNCP."""
    try:
        from services.comprasgov_service import get_comprasgov_service
        service = get_comprasgov_service(db=db)
        itens = await service.buscar_itens_contratacao(id_compra)
        return {"id_compra": id_compra, "total": len(itens), "itens": itens}
    except Exception as e:
        logger.error(f"Erro itens ComprasGov: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/compras-gov/precos/material")
async def get_compras_gov_precos(q: str = Query(..., description="Descrição do material")):
    """Pesquisa de preço histórico de materiais."""
    try:
        from services.comprasgov_service import get_comprasgov_service
        service = get_comprasgov_service(db=db)
        return await service.pesquisar_preco_material(descricao=q)
    except Exception as e:
        logger.error(f"Erro preço ComprasGov: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/compras-gov/fornecedor")
async def get_compras_gov_fornecedor(cnpj: str = Query(None), cpf: str = Query(None)):
    """Consulta dados de fornecedor no Compras.gov.br."""
    if not cnpj and not cpf:
        raise HTTPException(status_code=400, detail="Informe CNPJ ou CPF")
    try:
        from services.comprasgov_service import get_comprasgov_service
        service = get_comprasgov_service(db=db)
        return await service.consultar_fornecedor(cnpj=cnpj, cpf=cpf)
    except Exception as e:
        logger.error(f"Erro fornecedor ComprasGov: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/compras-gov/sync")
async def sync_compras_gov(dias: int = Query(1), uf: str = Query(None)):
    """Trigger manual para sincronização incremental da API."""
    try:
        from services.comprasgov_service import get_comprasgov_service
        service = get_comprasgov_service(db=db)
        return await service.sync_incremental(dias_atras=dias, uf=uf)
    except Exception as e:
        logger.error(f"Erro sync ComprasGov: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== STATUS DO CLONE ====================

@api_router.get("/clone/status")
async def clone_status():
    """Status dos editais clonados no MongoDB."""
    total = await db['editais_clone'].count_documents({})
    sample = await db['editais_clone'].find_one({}, {'_id': 0, 'objeto': 1, 'portal_captura': 1, 'importado_em': 1})
    portais = await db['editais_clone'].aggregate([
        {'$group': {'_id': '$portal_base', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]).to_list(length=50)
    pdfs_baixados = await db['editais_clone'].count_documents({'pdf_baixado': True})
    return {
        'total_editais': total,
        'pdfs_baixados': pdfs_baixados,
        'portais': {p['_id']: p['count'] for p in portais if p['_id']},
        'sample': sample,
        'independente': True
    }


# ==================== DOWNLOAD DE EDITAIS (PROXY PNCP) ====================

@api_router.get("/editais/download/{cnpj}/{ano}/{seq}")
async def download_edital_pncp(cnpj: str, ano: str, seq: str, doc: int = Query(1)):
    """Baixa o PDF/ZIP do edital direto do PNCP e serve ao usuário."""
    import aiohttp
    
    url = f"https://pncp.gov.br/pncp-api/v1/orgaos/{cnpj}/compras/{ano}/{seq}/arquivos/{doc}"
    
    try:
        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    raise HTTPException(status_code=404, detail="Arquivo não encontrado no PNCP")
                
                content = await resp.read()
                
                # Extrair nome do arquivo do header
                cd = resp.headers.get('Content-Disposition', '')
                filename = f"edital_{cnpj}_{ano}_{seq}.pdf"
                if 'filename=' in cd:
                    import urllib.parse
                    fname = cd.split('filename=')[-1].strip('"').strip("'")
                    filename = urllib.parse.unquote(fname)
                
                content_type = resp.headers.get('Content-Type', 'application/octet-stream')
                
                return Response(
                    content=content,
                    media_type=content_type,
                    headers={
                        'Content-Disposition': f'attachment; filename="{filename}"',
                        'Content-Length': str(len(content))
                    }
                )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro download PNCP {cnpj}/{ano}/{seq}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao baixar do PNCP")


@api_router.get("/editais/arquivos/{cnpj}/{ano}/{seq}")
async def listar_arquivos_edital(cnpj: str, ano: str, seq: str):
    """Lista arquivos/documentos de um edital direto do PNCP."""
    from services.motor_independente import MotorBuscaIndependente
    motor = MotorBuscaIndependente()
    arquivos = await motor.listar_arquivos(cnpj, ano, seq)
    return {"arquivos": arquivos, "total": len(arquivos)}


@api_router.get("/editais/itens/{cnpj}/{ano}/{seq}")
async def buscar_itens_edital(cnpj: str, ano: str, seq: str):
    """Busca itens de um edital direto do PNCP."""
    from services.motor_independente import MotorBuscaIndependente
    motor = MotorBuscaIndependente()
    itens = await motor.buscar_itens(cnpj, ano, seq)
    return {"itens": itens, "total": len(itens)}


# ==================== JANELA ANVISA - RADAR DE DESABASTECIMENTO ====================

@api_router.get("/anvisa/alertas")
async def listar_alertas_anvisa(limit: int = Query(50, ge=1, le=200)):
    """Lista alertas de desabastecimento da ANVISA."""
    from services.desabastecimento_service import DesabastecimentoService
    svc = DesabastecimentoService(db)
    alertas = await svc.listar_alertas(limit=limit)
    stats = await svc.estatisticas()
    return {"alertas": alertas, "estatisticas": stats}


@api_router.post("/anvisa/atualizar")
async def atualizar_anvisa(background_tasks: BackgroundTasks):
    """Executa coleta + análise ANVISA em background."""
    async def _run_scraper():
        try:
            from services.anvisa_scraper import AnvisaScraper
            from services.desabastecimento_service import DesabastecimentoService
            scraper = AnvisaScraper()
            alertas_brutos = await scraper.coletar_tudo()
            descont = await scraper.coletar_descontinuacao()
            alertas_brutos.extend(descont)
            svc = DesabastecimentoService(db)
            await svc.processar_alertas(alertas_brutos)
            logger.info(f"Scraper ANVISA concluído: {len(alertas_brutos)} alertas coletados")
        except Exception as e:
            logger.error(f"Scraper ANVISA erro: {e}")

    import asyncio
    asyncio.create_task(_run_scraper())
    return {"status": "processando", "mensagem": "Coleta ANVISA + DOU iniciada em background. Atualize em ~60s."}


@api_router.get("/anvisa/stats")
async def stats_anvisa():
    """Estatísticas rápidas do radar ANVISA."""
    from services.desabastecimento_service import DesabastecimentoService
    svc = DesabastecimentoService(db)
    return await svc.estatisticas()


@api_router.get("/anvisa/buscar-medicamento")
async def buscar_medicamento_anvisa(q: str = Query(..., min_length=2, description="Nome do medicamento")):
    """
    Busca inteligente de medicamento em múltiplas fontes ANVISA.
    Consulta: DOU, CMED, Notícias ANVISA, AnvisaLegis, Base GSM.
    Smart Cache: TTL 24h por medicamento.
    """
    # Smart Cache (24h)
    cached = smart_cache.get('anvisa_busca', medicamento=q.lower().strip())
    if cached is not None:
        logger.info(f"SmartCache HIT: anvisa_busca '{q}'")
        return cached

    from services.medicamento_search_service import get_medicamento_search_service
    svc = get_medicamento_search_service(db)
    resultado = await svc.buscar(q)

    smart_cache.set(resultado, namespace='anvisa_busca', medicamento=q.lower().strip())
    return resultado


@api_router.post("/anvisa/esclarecimento")
async def gerar_esclarecimento_anvisa(data: dict):
    """
    Gera texto de esclarecimento técnico para órgão público.
    DAMA P0: Inclui validação de vigência normativa.
    """
    from services.esclarecimento_service import gerar_esclarecimento
    from services.vigencia_service import get_vigencia_service

    medicamento = data.get('medicamento', '')
    principio_ativo = data.get('principio_ativo', medicamento)
    situacao = data.get('situacao', '')
    link_prova = data.get('link_prova', '')
    tipo_alerta = data.get('tipo_alerta', '')
    empresa_id = data.get('empresa_id', '')
    edital_info = data.get('edital_info')
    force_generate = data.get('force_generate', False)

    if not medicamento:
        raise HTTPException(status_code=400, detail="Medicamento é obrigatório")

    # DAMA P0: Validar vigência das normas antes de gerar
    vigencia_svc = get_vigencia_service(db)
    stats = await vigencia_svc.get_stats()
    if stats['total'] == 0:
        await vigencia_svc.scrape_resolucoes()

    vigencia_vigentes = await vigencia_svc.get_resolucoes_vigentes()
    
    # Verificar normas-chave
    normas_check = ["Resolução 07/2022", "Resolução 02/2004"]
    vigencia_alertas = []
    for norma in normas_check:
        check = await vigencia_svc.verificar_vigencia(norma)
        if check.get('encontrada') and not check.get('pode_usar'):
            vigencia_alertas.append(check)

    # Se há bloqueio e não forçou geração, retornar aviso
    if vigencia_alertas and not force_generate:
        return {
            'bloqueado': True,
            'vigencia_alertas': vigencia_alertas,
            'mensagem': 'Existem normas CMED caducas/revogadas. Revise antes de gerar.',
        }

    # Buscar dados da empresa
    empresa = {}
    if empresa_id:
        emp_doc = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        if emp_doc:
            empresa = emp_doc

    # Montar contexto de vigência para o LLM
    vigencia_context = []
    for v in vigencia_vigentes[:15]:
        vigencia_context.append(f"- {v['titulo']} | Status: {v['status']}")

    for alerta in vigencia_alertas:
        r = alerta.get('resolucao', {})
        vigencia_context.append(f"- BLOQUEADA: {r.get('titulo', alerta.get('referencia_buscada', ''))} | Status: {r.get('status', 'caduca/revogada')} - NÃO USAR")

    resultado = await gerar_esclarecimento(
        medicamento=medicamento,
        principio_ativo=principio_ativo,
        situacao=situacao,
        link_prova=link_prova,
        tipo_alerta=tipo_alerta,
        empresa=empresa,
        edital_info=edital_info,
        vigencia_context="\n".join(vigencia_context) if vigencia_context else None,
    )

    # Incluir info de vigência na resposta
    resultado['vigencia_alertas'] = vigencia_alertas
    resultado['vigencia_validada'] = True

    return resultado




@api_router.post("/anvisa/cruzar-licitacoes")
async def cruzar_anvisa_licitacoes():
    """
    Cruza alertas ANVISA com licitações do PNCP.
    Para cada medicamento com oportunidade real, busca licitações ativas
    usando medicamento_detectado e principio_ativo como termos de busca.
    Atualiza indice_oportunidade com bonus de licitações encontradas.
    """
    import asyncio
    from services.desabastecimento_service import DesabastecimentoService, calcular_indice
    from services.motor_independente import MotorBuscaIndependente

    svc = DesabastecimentoService(db)
    alertas = await svc.listar_alertas(limit=50)

    # Coletar medicamentos com oportunidade real
    medicamentos_buscar = {}
    for a in alertas:
        med = a.get('medicamento_detectado', '') or a.get('medicamento', '')
        if not med or med in ('N/A', 'Diversos', '-'):
            continue
        oport = a.get('oportunidade', '')
        risco = a.get('risco', '')
        if oport in ('Importação', 'Licitação provável', 'Demanda pública crítica') or risco == 'ALTO':
            if med not in medicamentos_buscar:
                medicamentos_buscar[med] = {
                    'principio_ativo': a.get('principio_ativo', med),
                    'oportunidade': oport,
                    'risco': risco,
                    'situacao': a.get('situacao', ''),
                    'tipo_alerta': a.get('tipo_alerta', ''),
                    'indice_oportunidade': a.get('indice_oportunidade', 0),
                    'gatilhos': a.get('gatilhos', []),
                }

    motor = MotorBuscaIndependente()
    cruzamento = {}

    async def buscar_med(nome, info):
        try:
            # Buscar pelo princípio ativo (mais preciso que nome comercial)
            pa = info.get('principio_ativo', nome)
            termo = pa.split('(')[0].strip()[:50]
            resultado = await motor.buscar(termo=termo, limit=10)
            lics = resultado.get('resultados', [])

            # Se não achou pelo PA, tentar pelo nome detectado
            if len(lics) == 0 and pa != nome:
                termo2 = nome.split('(')[0].strip()[:50]
                resultado2 = await motor.buscar(termo=termo2, limit=10)
                lics = resultado2.get('resultados', [])

            # Recalcular índice com bônus de licitações
            novo_indice = calcular_indice(
                info.get('gatilhos', []),
                info['risco'],
                info['oportunidade'],
                len(lics)
            )

            return nome, {
                'medicamento_detectado': nome,
                'principio_ativo': info['principio_ativo'],
                'oportunidade': info['oportunidade'],
                'risco': info['risco'],
                'situacao': info['situacao'],
                'tipo_alerta': info['tipo_alerta'],
                'indice_oportunidade': novo_indice,
                'licitacoes_encontradas': len(lics),
                'licitacoes': lics[:5]
            }
        except Exception as e:
            logger.error(f"Cruzamento ANVISA-PNCP '{nome}': {e}")
            return nome, {
                'medicamento_detectado': nome,
                'principio_ativo': info.get('principio_ativo', nome),
                'oportunidade': info['oportunidade'],
                'risco': info['risco'],
                'situacao': info['situacao'],
                'tipo_alerta': info.get('tipo_alerta', ''),
                'indice_oportunidade': info.get('indice_oportunidade', 0),
                'licitacoes_encontradas': 0,
                'licitacoes': []
            }

    tasks = [buscar_med(nome, info) for nome, info in list(medicamentos_buscar.items())[:8]]
    results = await asyncio.gather(*tasks)

    for nome, data in results:
        cruzamento[nome] = data

    total_com_licitacao = sum(1 for v in cruzamento.values() if v['licitacoes_encontradas'] > 0)
    total_licitacoes = sum(v['licitacoes_encontradas'] for v in cruzamento.values())

    return {
        'cruzamento': cruzamento,
        'resumo': {
            'medicamentos_analisados': len(cruzamento),
            'medicamentos_com_licitacao': total_com_licitacao,
            'total_licitacoes_encontradas': total_licitacoes
        }
    }



@api_router.get("/editais/pdf/{id_externo}")
async def download_edital_pdf(id_externo: str):
    """Serve PDF do edital - local ou redireciona para portal original."""
    from services.pdf_download_service import EDITAIS_DIR
    from fastapi.responses import RedirectResponse
    
    # 1. Verificar se já temos o PDF localmente (instantâneo)
    local_path = EDITAIS_DIR / f"{id_externo}.zip"
    if local_path.exists() and local_path.stat().st_size > 100:
        content = local_path.read_bytes()
        return Response(
            content=content,
            media_type='application/zip',
            headers={'Content-Disposition': f'attachment; filename="edital_{id_externo}.zip"'}
        )
    
    # 2. Buscar dados do edital no banco
    doc = await db['editais_clone'].find_one(
        {'id_externo': id_externo},
        {'_id': 0, 'link_pdf': 1, 'link_portal': 1}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Edital não encontrado")
    
    # 3. Redirecionar para o portal original (100% independente)
    link_portal = doc.get('link_portal', '')
    if link_portal and link_portal.startswith('http'):
        return RedirectResponse(url=link_portal)
    
    # 4. Último recurso: tentar link_pdf mesmo (melhor que nada)
    link_pdf = doc.get('link_pdf', '')
    if link_pdf and link_pdf.startswith('http'):
        return RedirectResponse(url=link_pdf)
    
    raise HTTPException(status_code=404, detail="Edital sem link disponível")


@api_router.post("/editais/baixar-lote")
async def baixar_pdfs_lote(limit: int = Query(50, ge=1, le=500)):
    """Dispara download em lote de PDFs do Conlicitação."""
    from services.pdf_download_service import PdfDownloadService
    import asyncio
    
    svc = PdfDownloadService(db)
    
    # Rodar em background
    async def _baixar():
        return await svc.baixar_lote(limit=limit)
    
    asyncio.create_task(_baixar())
    
    return {"message": f"Download de até {limit} PDFs iniciado em background"}


# ==================== ROTAS DE NORMALIZAÇÃO ====================

@api_router.post("/normalize/backfill")
async def run_backfill():
    """
    Executa backfill: Normaliza todos os editais raw → editais_normalizados
    
    Pipeline:
        editais_sync (raw) → normalizador_pncp → editais_normalizados
        
    Características:
    - Idempotente (pode rodar múltiplas vezes)
    - Usa hash_dedup para evitar duplicatas
    - Retorna estatísticas do processamento
    """
    from services.normalizador_pncp import NormalizadorPNCP
    
    try:
        normalizador = NormalizadorPNCP(db)
        
        # Configurar índices (idempotente)
        await normalizador.setup_indexes()
        
        # Executar backfill
        stats = await normalizador.backfill(batch_size=100)
        
        return {
            "message": "Backfill concluído",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no backfill: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/normalize/stats")
async def get_normalized_stats():
    """
    Retorna estatísticas da collection editais_normalizados
    
    Informações:
    - Total de editais normalizados
    - Distribuição por fonte
    - Top 10 UFs
    - Percentual de editais de saúde
    """
    from services.normalizador_pncp import NormalizadorPNCP
    
    try:
        normalizador = NormalizadorPNCP(db)
        stats = await normalizador.get_stats()
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DO MATCHER V2 ====================

@api_router.post("/matcher/processar")
async def processar_matcher():
    """
    Executa o Matcher v2 sobre editais_normalizados
    
    Processa todos os alertas ativos e gera matches com scoring.
    
    Pipeline:
        editais_normalizados → matcher_v2 → matches → notificações
    
    Returns:
        Estatísticas do processamento (alertas, matches, score médio)
    """
    from services.matcher_service import MatcherServiceV2
    
    try:
        matcher = MatcherServiceV2(db)
        await matcher.setup_indexes()
        
        stats = await matcher.processar_todos_alertas()
        
        return {
            "message": "Matcher v2 executado",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no matcher: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/matcher/stats")
async def get_matcher_stats():
    """
    Retorna estatísticas do Matcher v2
    
    Informações:
    - Total de matches
    - Matches pendentes (não processados)
    - Score médio recente
    - Threshold mínimo configurado
    """
    from services.matcher_service import MatcherServiceV2
    
    try:
        matcher = MatcherServiceV2(db)
        stats = await matcher.get_stats()
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar stats do matcher: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/matcher/alerta/{alerta_id}")
async def processar_alerta_especifico(alerta_id: str):
    """
    Processa um alerta específico e retorna matches
    
    Args:
        alerta_id: ID do alerta a processar
        
    Returns:
        Lista de matches com score e motivos
    """
    from services.matcher_service import MatcherServiceV2
    
    try:
        # Buscar alerta
        alerta = await db.alertas_notificacao.find_one(
            {'id': alerta_id},
            {'_id': 0}
        )
        
        if not alerta:
            raise HTTPException(status_code=404, detail="Alerta não encontrado")
        
        matcher = MatcherServiceV2(db)
        matches = await matcher.processar_alerta(alerta)
        
        return {
            "alerta_id": alerta_id,
            "total_matches": len(matches),
            "matches": [m.to_dict() for m in matches[:20]]  # Limitar resposta
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao processar alerta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== BACKFILL DE LINKS (Sistema GSM) ====================

@api_router.post("/backfill/links")
async def backfill_links_editais():
    """
    🔗 Enriquece links de editais existentes usando LinkResolverService V2
    
    PADRÃO GSM - Validação Real de Links:
    1. linkSistemaOrigem (portal do órgão - PRIORIDADE MÁXIMA)
    2. PDF do edital (download direto)  
    3. Link PNCP construído (menos confiável)
    4. INVALIDO - exibir número do processo para busca manual
    
    Links de BUSCA (?q=) são REJEITADOS automaticamente.
    
    IMPORTANTE: Este endpoint faz lookup no editais_sync (raw) para 
    obter o link_origem original que não foi preservado na normalização.
    """
    from services.link_resolver_service_v2 import get_link_resolver_v2
    
    try:
        link_resolver = get_link_resolver_v2()
        
        stats = {
            'processados': 0,
            'atualizados': 0,
            'pncp': 0,
            'portal': 0,
            'pdf': 0,
            'fallback': 0,
            'erros': 0
        }
        
        logger.info("🔗 [BACKFILL] Iniciando enriquecimento de links (com lookup RAW)...")
        
        # Buscar todos os editais normalizados
        cursor = db.editais_normalizados.find({}, {'_id': 0})
        
        async for edital in cursor:
            try:
                # LOOKUP: Buscar link_origem no editais_sync (RAW)
                numero_processo = edital.get('numero_processo') or edital.get('id_externo')
                raw_doc = None
                
                if numero_processo:
                    raw_doc = await db.editais_sync.find_one(
                        {'$or': [
                            {'numero_processo': numero_processo},
                            {'fonte_id': numero_processo},
                            {'id_externo': numero_processo}
                        ]},
                        {'_id': 0, 'link_origem': 1, 'linkSistemaOrigem': 1, 'link_documento': 1}
                    )
                
                # Enriquecer edital com dados do RAW
                if raw_doc:
                    edital['linkSistemaOrigem'] = raw_doc.get('linkSistemaOrigem')
                    edital['link_origem'] = raw_doc.get('link_origem')
                    edital['link_documento'] = raw_doc.get('link_documento')
                
                # Resolver links com dados enriquecidos
                links = link_resolver.resolver_link(edital)
                
                # Atualizar edital com links resolvidos (Sistema GSM)
                update_data = {
                    'link_edital': links['link_principal'],
                    'link_pncp': links['link_pncp'],
                    'link_portal_orgao': links['link_portal'],
                    'link_pdf': links['link_pdf'],
                    'tipo_link': links['tipo_link'],
                    'link_status': links['link_status'],  # VALIDO ou INVALIDO
                }
                
                if links['aviso']:
                    update_data['aviso_link'] = links['aviso']
                
                # Atualizar no banco
                result = await db.editais_normalizados.update_one(
                    {'hash_dedup': edital.get('hash_dedup')},
                    {'$set': update_data}
                )
                
                stats['processados'] += 1
                if result.modified_count > 0:
                    stats['atualizados'] += 1
                
                # Contabilizar tipo de link (Sistema GSM)
                tipo = links.get('tipo_link')
                status = links.get('link_status', 'INVALIDO')
                
                if status == 'VALIDO':
                    if tipo == 'pncp_direto':
                        stats['pncp'] += 1
                    elif tipo == 'portal_orgao':
                        stats['portal'] += 1
                    elif tipo == 'pdf':
                        stats['pdf'] += 1
                else:
                    stats['fallback'] += 1  # Contabiliza como inválido
                    
            except Exception as e:
                logger.error(f"❌ Erro ao processar edital: {str(e)}")
                stats['erros'] += 1
        
        logger.info(f"✅ [BACKFILL] Concluído: {stats}")
        
        return {
            "message": "Backfill de links concluído",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no backfill de links: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE PORTAIS MUNICIPAIS (P2) ====================

@api_router.get("/municipios/lista")
async def listar_municipios_disponiveis():
    """
    🏛️ Lista municípios disponíveis para sincronização
    
    Retorna lista de municípios SP prioritários com:
    - Status (ativo/inativo)
    - CNPJ para busca no PNCP
    - Portal oficial
    """
    from scrapers.portais_municipais_sp import MUNICIPIOS_SP
    
    return {
        "municipios": [
            {
                "id": k,
                "nome": v["nome"],
                "cnpj": v["cnpj"],
                "portal": v["portal"],
                "ativo": v.get("ativo", False),
            }
            for k, v in MUNICIPIOS_SP.items()
        ],
        "total_ativos": sum(1 for v in MUNICIPIOS_SP.values() if v.get("ativo", False)),
    }


@api_router.post("/municipios/sync-cnpj")
async def sincronizar_por_cnpj(
    cnpj: str = Query(..., description="CNPJ do órgão (apenas números)"),
    ano: int = Query(None, description="Ano das compras (default: atual)"),
    limite: int = Query(50, ge=10, le=200, description="Limite de editais"),
    salvar: bool = Query(True, description="Salvar no banco de dados"),
):
    """
    🏛️ P4.3: Sincroniza editais de um órgão específico via PNCP.
    
    Usa o endpoint correto: /api/consulta/v1/orgaos/{cnpj}/compras/{ano}/{seq}
    
    Exemplo: /api/municipios/sync-cnpj?cnpj=46319000000150 (Guarulhos)
    """
    from scrapers.portais_municipais_sp import get_portais_municipais_sp
    import hashlib
    
    try:
        integrador = get_portais_municipais_sp()
        
        # Buscar compras do órgão
        editais = await integrador.buscar_compras_por_cnpj(
            cnpj=cnpj,
            ano=ano,
            limite=limite
        )
        
        stats = {
            "cnpj": cnpj,
            "total_encontrados": len(editais),
            "total_salvos": 0,
            "total_atualizados": 0,
            "editais": []
        }
        
        if salvar:
            for edital in editais:
                try:
                    # Hash de deduplicação
                    hash_content = f"{edital.get('cnpj_orgao')}-{edital.get('numero_processo')}"
                    edital["hash_dedup"] = hashlib.md5(hash_content.encode()).hexdigest()
                    edital["created_at"] = datetime.now(timezone.utc)
                    
                    # Verificar se já existe
                    existe = await db.editais_normalizados.find_one(
                        {"id_externo": edital.get("id_externo")},
                        {"_id": 0, "id_externo": 1}
                    )
                    
                    if not existe:
                        await db.editais_normalizados.insert_one(edital)
                        stats["total_salvos"] += 1
                    else:
                        await db.editais_normalizados.update_one(
                            {"id_externo": edital.get("id_externo")},
                            {"$set": edital}
                        )
                        stats["total_atualizados"] += 1
                    
                except Exception as e:
                    logger.debug(f"Erro ao salvar edital: {e}")
        
        # Incluir amostra dos editais
        stats["editais"] = [
            {
                "objeto": e.get("objeto", "")[:60],
                "numero": e.get("numero_processo"),
                "data_abertura": e.get("data_abertura"),
                "link": e.get("link_edital")
            }
            for e in editais[:5]
        ]
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao sincronizar CNPJ {cnpj}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/municipios/sync-sp")
async def sincronizar_municipais_sp(
    dias: int = Query(60, ge=7, le=180, description="Período em dias"),
    limite: int = Query(100, ge=10, le=500, description="Limite de editais"),
    salvar: bool = Query(True, description="Salvar no banco de dados"),
):
    """
    🏛️ P4: Sincroniza TODOS os editais municipais de SP via PNCP
    
    Estratégia eficiente:
    1. Busca única no PNCP com UF=SP
    2. Filtra por esfera Municipal
    3. Valida links no padrão GSM
    4. Salva na collection editais_normalizados
    """
    from scrapers.portais_municipais_sp import get_portais_municipais_sp
    import hashlib
    
    try:
        integrador = get_portais_municipais_sp()
        
        # Buscar todos os municipais de SP
        editais = await integrador.buscar_todos_municipais_sp(
            dias=dias,
            limite=limite
        )
        
        # Estatísticas
        stats = {
            "total_encontrados": len(editais),
            "total_salvos": 0,
            "total_atualizados": 0,
            "total_invalidos": 0,
            "por_municipio": {},
            "links_validos": 0,
            "links_invalidos": 0,
        }
        
        # Salvar no banco se solicitado
        if salvar:
            for edital in editais:
                try:
                    # Validar link
                    link = edital.get("link_edital")
                    link_status = edital.get("link_status", "INVALIDO")
                    
                    if link_status == "VALIDO":
                        stats["links_validos"] += 1
                    else:
                        stats["links_invalidos"] += 1
                    
                    # Criar hash de deduplicação
                    hash_content = f"{edital.get('cnpj_orgao')}-{edital.get('numero_processo')}-{edital.get('data_abertura')}"
                    edital["hash_dedup"] = hashlib.md5(hash_content.encode()).hexdigest()
                    edital["created_at"] = datetime.now(timezone.utc)
                    edital["uf"] = "SP"
                    edital["esfera"] = "Municipal"
                    
                    # Verificar se já existe
                    existe = await db.editais_normalizados.find_one(
                        {"id_externo": edital.get("id_externo")},
                        {"_id": 0, "id_externo": 1}
                    )
                    
                    if not existe:
                        await db.editais_normalizados.insert_one(edital)
                        stats["total_salvos"] += 1
                    else:
                        await db.editais_normalizados.update_one(
                            {"id_externo": edital.get("id_externo")},
                            {"$set": edital}
                        )
                        stats["total_atualizados"] += 1
                    
                    # Contar por município
                    mun = edital.get("municipio", "Outros")
                    if mun not in stats["por_municipio"]:
                        stats["por_municipio"][mun] = 0
                    stats["por_municipio"][mun] += 1
                    
                except Exception as e:
                    stats["total_invalidos"] += 1
                    logger.debug(f"Erro ao salvar edital: {e}")
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao sincronizar municipais SP: {e}")
        raise HTTPException(status_code=500, detail=str(e))
@api_router.post("/municipios/sync")
async def sincronizar_municipios(
    municipios: str = Query(None, description="IDs dos municípios separados por vírgula (ex: santo_andre,guarulhos)"),
    dias: int = Query(90, ge=7, le=365, description="Período em dias"),
    salvar: bool = Query(True, description="Salvar no banco de dados"),
):
    """
    🏛️ Sincroniza editais dos municípios prioritários via PNCP
    
    Estratégia:
    1. Busca editais no PNCP por CNPJ do município
    2. Aplica filtro temporal obrigatório
    3. Garante links no padrão GSM
    4. Salva na collection editais_normalizados
    
    Args:
        municipios: Lista de IDs (None = todos ativos)
        dias: Período de busca
        salvar: Se True, salva no banco
    """
    from scrapers.portais_municipais_sp import get_portais_municipais_sp, MUNICIPIOS_SP
    from services.link_resolver_service_v2 import get_link_resolver_v2
    
    try:
        integrador = get_portais_municipais_sp()
        link_resolver = get_link_resolver_v2()
        
        # Parsear municípios
        mun_lista = None
        if municipios:
            mun_lista = [m.strip() for m in municipios.split(",") if m.strip()]
        
        # Sincronizar
        resultado = await integrador.sincronizar_municipios(
            municipios=mun_lista,
            dias=dias,
        )
        
        # Salvar no banco se solicitado
        if salvar:
            total_salvos = 0
            
            for mun_id, mun_data in resultado.get("por_municipio", {}).items():
                for edital in mun_data.get("editais", []):
                    try:
                        # Enriquecer com link resolver
                        edital = link_resolver.enriquecer_edital(edital)
                        
                        # Verificar se já existe
                        existe = await db.editais_normalizados.find_one(
                            {"id_externo": edital.get("id_externo")},
                            {"_id": 0, "id_externo": 1}
                        )
                        
                        if not existe:
                            # Adicionar hash de deduplicação
                            import hashlib
                            hash_content = f"{edital.get('cnpj_orgao')}-{edital.get('numero_processo')}-{edital.get('data_abertura')}"
                            edital["hash_dedup"] = hashlib.md5(hash_content.encode()).hexdigest()
                            edital["created_at"] = datetime.now(timezone.utc)
                            
                            await db.editais_normalizados.insert_one(edital)
                            total_salvos += 1
                    
                    except Exception as e:
                        logger.error(f"❌ Erro ao salvar edital: {str(e)}")
            
            resultado["salvos"] = total_salvos
        
        # Limpar resposta (remover editais detalhados para evitar ObjectId)
        resposta = {
            "total": resultado.get("total", 0),
            "salvos": resultado.get("salvos", 0),
            "inicio": resultado.get("inicio"),
            "fim": resultado.get("fim"),
            "erros": resultado.get("erros", []),
            "por_municipio": {}
        }
        
        for mun_id, mun_data in resultado.get("por_municipio", {}).items():
            resposta["por_municipio"][mun_id] = {
                "nome": mun_data.get("nome"),
                "total": mun_data.get("total", 0),
                # Resumo dos primeiros 5 editais (sem ObjectId)
                "exemplos": [
                    {
                        "objeto": (e.get("objeto", "") or "")[:80],
                        "link_status": e.get("link_status"),
                        "data_abertura": str(e.get("data_abertura", ""))[:10],
                    }
                    for e in mun_data.get("editais", [])[:5]
                ]
            }
        
        return resposta
        
    except Exception as e:
        logger.error(f"❌ Erro na sincronização municipal: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/municipios/stats")
async def estatisticas_municipios():
    """
    📊 Estatísticas de editais por município
    
    Retorna contagem de editais por município SP
    """
    try:
        # Agregação por município
        pipeline = [
            {"$match": {"uf": "SP", "esfera": "Municipal"}},
            {"$group": {
                "_id": "$municipio",
                "total": {"$sum": 1},
                "com_link_valido": {
                    "$sum": {"$cond": [{"$eq": ["$link_status", "VALIDO"]}, 1, 0]}
                },
            }},
            {"$sort": {"total": -1}},
        ]
        
        cursor = db.editais_normalizados.aggregate(pipeline)
        stats = await cursor.to_list(length=50)
        
        return {
            "por_municipio": [
                {
                    "municipio": s["_id"] or "Não identificado",
                    "total": s["total"],
                    "com_link_valido": s["com_link_valido"],
                }
                for s in stats
            ],
            "total_geral": sum(s["total"] for s in stats),
        }
        
    except Exception as e:
        logger.error(f"❌ Erro nas estatísticas municipais: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE MONITORAMENTO (DASHBOARD) ====================

@api_router.get("/monitoring/dashboard")
async def get_monitoring_dashboard():
    """
    🎛️ Dashboard de Monitoramento Completo
    
    Retorna todas as métricas operacionais do sistema:
    - Status dos workers (OK / ERRO / ATRASO)
    - Status das fontes de dados
    - Métricas do pipeline (raw → normalizado → match)
    - Métricas de alertas (disparados vs suprimidos)
    - Score de saúde geral (0-100)
    
    Este endpoint fecha o ciclo:
    Fonte → Normalização → Match → Alerta → Monitoramento
    """
    from services.monitoring_service import MonitoringService
    
    try:
        monitoring = MonitoringService(db)
        dashboard = await monitoring.get_dashboard_completo()
        
        return dashboard
        
    except Exception as e:
        logger.error(f"❌ Erro no dashboard: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/monitoring/workers")
async def get_workers_status():
    """
    Retorna status detalhado dos workers
    
    Workers monitorados:
    - sync_pncp: Sincronização PNCP → MongoDB
    - check_alerts: Verificação de alertas
    - matcher_v2: Processamento de matches
    - cleanup: Limpeza de dados antigos
    """
    from services.monitoring_service import MonitoringService
    
    try:
        monitoring = MonitoringService(db)
        return await monitoring._get_workers_status()
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar workers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/monitoring/fontes")
async def get_fontes_status():
    """
    Retorna status das fontes de dados
    
    Informações por fonte:
    - Status (OK / ERRO)
    - Última execução
    - Taxa de sucesso
    - Total de resultados
    """
    from services.monitoring_service import MonitoringService
    
    try:
        monitoring = MonitoringService(db)
        return await monitoring._get_fontes_status()
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar fontes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/monitoring/pipeline")
async def get_pipeline_metrics():
    """
    Retorna métricas do pipeline de dados
    
    Pipeline: Fonte → Raw → Normalizado → Match
    
    Métricas:
    - Editais raw vs normalizados
    - Taxa de normalização
    - Matches gerados
    - Editais de saúde
    """
    from services.monitoring_service import MonitoringService
    
    try:
        monitoring = MonitoringService(db)
        return await monitoring._get_pipeline_metrics()
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar pipeline: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/monitoring/alertas")
async def get_alertas_metrics():
    """
    Retorna métricas do sistema de alertas
    
    Métricas:
    - Alertas ativos/inativos
    - Matches disparados vs suprimidos
    - Score médio
    - Notificações enviadas
    """
    from services.monitoring_service import MonitoringService
    
    try:
        monitoring = MonitoringService(db)
        return await monitoring._get_alertas_metrics()
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar alertas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== v4.1: AUDITORIA DE BUSCAS ====================

@api_router.get("/audit/search/zero-results")
async def get_zero_result_searches(
    limite: int = Query(50, ge=1, le=200, description="Máximo de termos"),
    dias: int = Query(30, ge=1, le=90, description="Período em dias")
):
    """
    🔴 v4.1: Retorna termos que frequentemente retornam zero resultados
    
    Útil para identificar gaps de dados e priorizar novas fontes.
    """
    from services.search_audit_service import get_search_audit
    
    try:
        audit = get_search_audit(db)
        termos = await audit.get_termos_sem_resultado(limite=limite, dias=dias)
        
        return {
            "periodo_dias": dias,
            "total_termos": len(termos),
            "termos_sem_resultado": termos
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar auditoria: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/audit/search/stats")
async def get_search_stats(
    dias: int = Query(7, ge=1, le=90, description="Período em dias")
):
    """
    📊 v4.1: Retorna estatísticas de buscas
    
    Informações:
    - Total de buscas
    - Buscas sem resultado
    - Taxa de zero resultados
    - Top termos buscados
    """
    from services.search_audit_service import get_search_audit
    
    try:
        audit = get_search_audit(db)
        stats = await audit.get_estatisticas(dias=dias)
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar estatísticas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== v4.1: HIGIENE DE DADOS ====================

@api_router.post("/data/cleanup/mocks")
async def cleanup_mock_data():
    """
    🧹 v4.1: Remove TODOS os dados mock da collection licitacoes
    
    ⚠️ ATENÇÃO: Esta operação é IRREVERSÍVEL.
    
    O sistema passará a operar apenas com dados reais:
    - PNCP (tempo real)
    - ComprasNet (tempo real)
    - BNC (tempo real)
    """
    try:
        # Contar mocks antes
        total_antes = await db.licitacoes.count_documents({})
        mocks_antes = await db.licitacoes.count_documents({"is_mock": True})
        mocks_sem_flag = await db.licitacoes.count_documents({
            "$or": [
                {"fonte": {"$regex": "mock", "$options": "i"}},
                {"fonte_nome": {"$regex": "mock", "$options": "i"}},
                {"medicamento": {"$regex": "^(Adalimumabe|Pembrolizumabe|Rituximabe)$", "$options": "i"}}
            ]
        })
        
        # Remover mocks
        result1 = await db.licitacoes.delete_many({"is_mock": True})
        result2 = await db.licitacoes.delete_many({
            "$or": [
                {"fonte": {"$regex": "mock", "$options": "i"}},
                {"fonte_nome": {"$regex": "mock", "$options": "i"}}
            ]
        })
        
        total_depois = await db.licitacoes.count_documents({})
        
        logger.info(f"🧹 [CLEANUP] Removidos {result1.deleted_count + result2.deleted_count} registros mock")
        
        return {
            "status": "success",
            "antes": {
                "total": total_antes,
                "mocks_flag": mocks_antes,
                "mocks_inferidos": mocks_sem_flag
            },
            "removidos": {
                "por_flag": result1.deleted_count,
                "por_inferencia": result2.deleted_count
            },
            "depois": {
                "total": total_depois
            },
            "mensagem": "Dados mock removidos com sucesso. Sistema operando apenas com dados reais."
        }
        
    except Exception as e:
        logger.error(f"❌ Erro na limpeza: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/data/sources/status")
async def get_data_sources_status():
    """
    📊 v4.1: Status das fontes de dados ativas
    
    Retorna:
    - PNCP: Status da API e última sincronização
    - ComprasNet: Status do scraper
    - BNC: Status do scraper
    - Editais por fonte
    """
    try:
        # Contagem por fonte em editais_normalizados
        pipeline = [
            {"$group": {"_id": "$fonte", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        
        por_fonte = await db.editais_normalizados.aggregate(pipeline).to_list(20)
        
        # Contagem na collection licitacoes
        pipeline_lic = [
            {"$group": {"_id": "$fonte", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        
        por_fonte_lic = await db.licitacoes.aggregate(pipeline_lic).to_list(20)
        
        # Verificar mocks restantes
        mocks_restantes = await db.licitacoes.count_documents({"is_mock": True})
        
        return {
            "fontes_ativas": {
                "PNCP": "✅ Ativo",
                "ComprasNet": "✅ Ativo (v4.1)",
                "BNC": "✅ Ativo (v4.1)"
            },
            "editais_normalizados": {
                "total": await db.editais_normalizados.count_documents({}),
                "por_fonte": {f["_id"] or "Não identificado": f["count"] for f in por_fonte}
            },
            "licitacoes": {
                "total": await db.licitacoes.count_documents({}),
                "mocks_restantes": mocks_restantes,
                "por_fonte": {f["_id"] or "Não identificado": f["count"] for f in por_fonte_lic}
            }
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar status das fontes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== v4.5: DASHBOARD COMPARATIVO (P6 - Preparação) ====================

@api_router.get("/dashboard/metrics")
async def get_dashboard_metrics():
    """
    📊 v4.5 P6: Métricas do Dashboard Comparativo
    
    Retorna:
    - Total Capturado GSM (todas as fontes)
    - Total Exibido (Pós-Filtro Mata-Lixo)
    - Alertas ativos e enviados
    - Taxa de relevância
    - Top termos buscados
    """
    from services.search_audit_service import get_search_audit
    
    try:
        # =====================================================================
        # MÉTRICAS DE CAPTURA
        # =====================================================================
        
        # Total capturado (todas collections)
        total_editais_normalizados = await db.editais_normalizados.count_documents({})
        total_licitacoes = await db.licitacoes.count_documents({})
        total_capturado = total_editais_normalizados + total_licitacoes
        
        # Por fonte
        pipeline_fonte = [
            {"$group": {"_id": "$fonte", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        por_fonte = await db.editais_normalizados.aggregate(pipeline_fonte).to_list(10)
        
        # =====================================================================
        # MÉTRICAS DE FILTRO MATA-LIXO
        # =====================================================================
        
        # Estatísticas de busca (usa o SearchAuditService)
        audit = get_search_audit(db)
        stats_busca = await audit.get_estatisticas(dias=7)
        
        # Taxa de zero resultados
        taxa_zero = stats_busca.get("taxa_zero_resultados", 0)
        
        # Top termos buscados
        top_termos = stats_busca.get("top_termos", [])
        
        # =====================================================================
        # MÉTRICAS DE ALERTAS (P5)
        # =====================================================================
        
        total_alertas = await db.alertas.count_documents({})
        alertas_ativos = await db.alertas.count_documents({"ativo": True})
        
        # Total de emails enviados
        pipeline_emails = [
            {"$group": {"_id": None, "total": {"$sum": "$total_enviados"}}}
        ]
        emails_result = await db.alertas.aggregate(pipeline_emails).to_list(1)
        total_emails_enviados = emails_result[0]["total"] if emails_result else 0
        
        # Termos monitorados
        termos_monitorados = await db.alertas.distinct("termo", {"ativo": True})
        
        # =====================================================================
        # MÉTRICAS DE QUALIDADE
        # =====================================================================
        
        # Editais com itens correspondentes (Match confirmado)
        # Isso indica a eficácia do filtro Mata-Lixo
        pipeline_match = [
            {"$match": {"itens_correspondentes": {"$exists": True, "$ne": []}}},
            {"$count": "total"}
        ]
        match_result = await db.editais_normalizados.aggregate(pipeline_match).to_list(1)
        total_com_match = match_result[0]["total"] if match_result else 0
        
        # Taxa de relevância (editais com match / total)
        taxa_relevancia = round(total_com_match / total_editais_normalizados * 100, 2) if total_editais_normalizados > 0 else 0
        
        return {
            "versao": "v4.5 P6 Dashboard",
            "captura": {
                "total_gsm": total_capturado,
                "editais_normalizados": total_editais_normalizados,
                "licitacoes": total_licitacoes,
                "por_fonte": {f["_id"] or "N/A": f["count"] for f in por_fonte}
            },
            "filtro_mata_lixo": {
                "total_buscas_7d": stats_busca.get("total_buscas", 0),
                "buscas_sem_resultado": stats_busca.get("total_sem_resultado", 0),
                "taxa_zero_resultados": f"{taxa_zero}%",
                "termos_unicos": stats_busca.get("termos_unicos", 0),
                "top_termos": top_termos[:5]
            },
            "alertas_p5": {
                "total_alertas": total_alertas,
                "ativos": alertas_ativos,
                "emails_enviados": total_emails_enviados,
                "termos_monitorados": termos_monitorados[:10]
            },
            "qualidade": {
                "editais_com_match_confirmado": total_com_match,
                "taxa_relevancia": f"{taxa_relevancia}%"
            }
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar métricas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/search/local/filters")
async def get_local_filters():
    """
    Retorna os filtros disponíveis baseados nos dados reais do banco local
    
    Útil para popular dropdowns de filtro no frontend
    
    Returns:
        - estados: Lista de UFs disponíveis
        - modalidades: Lista de modalidades disponíveis
        - esferas: Lista de esferas (Federal, Estadual, Municipal)
    """
    try:
        if _sync_service_instance is None:
            return {
                "estados": [],
                "modalidades": [],
                "esferas": [],
                "total_editais": 0
            }
        
        # Agregar valores únicos de cada campo
        collection = _sync_service_instance.editais_collection
        
        # Estados únicos
        estados_pipeline = [
            {"$match": {"estado": {"$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$estado"}},
            {"$sort": {"_id": 1}}
        ]
        estados_result = await collection.aggregate(estados_pipeline).to_list(100)
        estados = [e["_id"] for e in estados_result if e["_id"]]
        
        # Modalidades únicas
        modalidades_pipeline = [
            {"$match": {"modalidade": {"$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$modalidade"}},
            {"$sort": {"_id": 1}}
        ]
        modalidades_result = await collection.aggregate(modalidades_pipeline).to_list(50)
        modalidades = [m["_id"] for m in modalidades_result if m["_id"]]
        
        # Esferas únicas
        esferas_pipeline = [
            {"$match": {"esfera": {"$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$esfera"}},
            {"$sort": {"_id": 1}}
        ]
        esferas_result = await collection.aggregate(esferas_pipeline).to_list(10)
        esferas = [e["_id"] for e in esferas_result if e["_id"]]
        
        # Total de editais
        total = await collection.count_documents({})
        
        return {
            "estados": estados,
            "modalidades": modalidades,
            "esferas": esferas,
            "total_editais": total
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar filtros: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== AUTOCOMPLETE DE MEDICAMENTOS (v4.1 ELITE) ====================

@api_router.get("/suggestions")
async def get_suggestions(
    q: str = Query(..., min_length=2, description="Termo de busca (mínimo 2 caracteres)"),
    limit: int = Query(10, ge=1, le=30, description="Número máximo de sugestões")
):
    """
    🔍 v4.1 ELITE: Autocomplete inteligente para medicamentos
    
    Retorna sugestões de termos de busca baseado em:
    1. Medicamentos já capturados em editais anteriores
    2. Princípios ativos conhecidos
    3. Nomes comerciais de medicamentos
    
    O endpoint busca nas collections:
    - editais_normalizados (itens de editais)
    - licitacoes (campo medicamento)
    
    Comportamento:
    - A partir de 2 caracteres, retorna sugestões
    - Busca por partial match (LIKE %termo%)
    - Ordena por frequência de aparição
    
    Args:
        q: Termo de busca (mínimo 2 caracteres)
        limit: Número máximo de sugestões (default: 10)
        
    Returns:
        Lista de sugestões com termo e frequência
    """
    import re
    from services.busca_service_v2 import EXPANSAO_TERMOS_SAUDE
    from services.item_extractor_service import SINONIMOS_MEDICAMENTOS
    
    try:
        termo = q.strip().lower()
        sugestoes = []
        termos_vistos = set()
        
        # =====================================================================
        # 1. SUGESTÕES DO DICIONÁRIO DE MEDICAMENTOS (mais relevantes)
        # =====================================================================
        
        # Buscar em EXPANSAO_TERMOS_SAUDE
        for chave, expansoes in EXPANSAO_TERMOS_SAUDE.items():
            # Verificar se a chave contém o termo
            if termo in chave.lower():
                if chave.lower() not in termos_vistos:
                    sugestoes.append({
                        "termo": chave.capitalize(),
                        "tipo": "medicamento",
                        "frequencia": 100,  # Prioridade máxima
                        "categoria": "Princípio Ativo"
                    })
                    termos_vistos.add(chave.lower())
            
            # Verificar nas expansões
            for exp in expansoes:
                if termo in exp.lower() and exp.lower() not in termos_vistos:
                    sugestoes.append({
                        "termo": exp.capitalize(),
                        "tipo": "medicamento",
                        "frequencia": 90,
                        "categoria": "Nome Comercial"
                    })
                    termos_vistos.add(exp.lower())
        
        # Buscar em SINONIMOS_MEDICAMENTOS
        for principio, nomes in SINONIMOS_MEDICAMENTOS.items():
            if termo in principio.lower() and principio.lower() not in termos_vistos:
                sugestoes.append({
                    "termo": principio.capitalize(),
                    "tipo": "medicamento",
                    "frequencia": 95,
                    "categoria": "Princípio Ativo"
                })
                termos_vistos.add(principio.lower())
            
            for nome in nomes:
                if termo in nome.lower() and nome.lower() not in termos_vistos:
                    sugestoes.append({
                        "termo": nome.capitalize(),
                        "tipo": "medicamento",
                        "frequencia": 85,
                        "categoria": "Nome Comercial"
                    })
                    termos_vistos.add(nome.lower())
        
        # =====================================================================
        # 2. SUGESTÕES DO BANCO DE DADOS (medicamentos já capturados)
        # =====================================================================
        
        # Buscar medicamentos únicos na collection licitacoes
        try:
            regex_pattern = {"$regex": f".*{re.escape(termo)}.*", "$options": "i"}
            
            # Agregação para encontrar medicamentos frequentes
            pipeline_licitacoes = [
                {"$match": {"medicamento": regex_pattern}},
                {"$group": {
                    "_id": {"$toLower": "$medicamento"},
                    "count": {"$sum": 1}
                }},
                {"$sort": {"count": -1}},
                {"$limit": 20}
            ]
            
            cursor = db.licitacoes.aggregate(pipeline_licitacoes)
            resultados_lic = await cursor.to_list(20)
            
            for r in resultados_lic:
                med = r["_id"]
                if med and len(med) > 2 and med not in termos_vistos:
                    # Capitalizar primeira letra de cada palavra
                    termo_formatado = ' '.join(word.capitalize() for word in med.split())
                    sugestoes.append({
                        "termo": termo_formatado,
                        "tipo": "capturado",
                        "frequencia": min(r["count"], 80),
                        "categoria": f"Encontrado em {r['count']} editais"
                    })
                    termos_vistos.add(med)
        except Exception as e:
            logger.debug(f"Erro ao buscar medicamentos: {e}")
        
        # Buscar nos itens de editais normalizados
        try:
            # Buscar descrições de itens que contenham o termo
            pipeline_itens = [
                {"$match": {"itens_edital.descricao": regex_pattern}},
                {"$unwind": "$itens_edital"},
                {"$match": {"itens_edital.descricao": regex_pattern}},
                {"$group": {
                    "_id": {"$toLower": "$itens_edital.descricao"},
                    "count": {"$sum": 1}
                }},
                {"$sort": {"count": -1}},
                {"$limit": 10}
            ]
            
            cursor = db.editais_normalizados.aggregate(pipeline_itens)
            resultados_itens = await cursor.to_list(10)
            
            for r in resultados_itens:
                desc = r["_id"]
                if desc and len(desc) > 3:
                    # Extrair nome do medicamento (primeiras 3-4 palavras)
                    palavras = desc.split()[:4]
                    nome_curto = ' '.join(palavras)
                    
                    if len(nome_curto) > 5 and nome_curto.lower() not in termos_vistos:
                        sugestoes.append({
                            "termo": nome_curto.capitalize(),
                            "tipo": "item",
                            "frequencia": min(r["count"], 70),
                            "categoria": "Item de Edital"
                        })
                        termos_vistos.add(nome_curto.lower())
        except Exception as e:
            logger.debug(f"Erro ao buscar itens: {e}")
        
        # =====================================================================
        # 3. ORDENAR E LIMITAR RESULTADOS
        # =====================================================================
        
        # Ordenar por frequência (prioridade) decrescente
        sugestoes_ordenadas = sorted(sugestoes, key=lambda x: x["frequencia"], reverse=True)
        
        # Limitar resultados
        sugestoes_final = sugestoes_ordenadas[:limit]
        
        logger.info(f"🔍 [SUGGESTIONS] '{termo}' → {len(sugestoes_final)} sugestões")
        
        return {
            "termo_busca": q,
            "total": len(sugestoes_final),
            "sugestoes": sugestoes_final
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar sugestões: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/sync/stats")
async def get_sync_stats():
    """
    Retorna estatísticas do serviço de sincronização
    
    Informações:
    - Total de editais sincronizados
    - Última sincronização
    - Distribuição por fonte
    - Status da collection 'editais'
    """
    try:
        if _sync_service_instance is None:
            return {
                "status": "not_initialized",
                "message": "SyncService ainda não foi inicializado",
                "total_editais": 0
            }
        
        stats = await _sync_service_instance.get_stats()
        stats["status"] = "active"
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao obter stats de sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/trigger")
async def trigger_sync():
    """
    Dispara sincronização manual com o PNCP
    
    Útil para forçar atualização imediata do banco local.
    Em produção, a sincronização é automática (a cada 15 min).
    """
    try:
        if _sync_service_instance is None:
            raise HTTPException(
                status_code=503,
                detail="SyncService não inicializado"
            )
        
        logger.info("📥 [SYNC] Sincronização manual disparada...")
        stats = await _sync_service_instance.sync_pncp()
        
        return {
            "message": "Sincronização concluída",
            "stats": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na sincronização manual: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS MULTI-FONTE (EXPANSÃO) ====================

@api_router.post("/sync/fonte/{fonte}")
async def sync_fonte_especifica(
    fonte: str,
    limit: int = Query(50, ge=1, le=200, description="Máximo de registros"),
    apenas_saude: bool = Query(False, description="Filtrar apenas saúde")
):
    """
    🔄 Sincroniza uma fonte específica de dados
    
    Fontes disponíveis:
    - comprasnet: Portal Federal de Compras
    - tce-sp: Tribunal de Contas de SP
    - mg-csv: Dados Abertos de Minas Gerais
    - pr-csv: Portal da Transparência do Paraná
    - go-csv: Dados Abertos de Goiás
    
    Args:
        fonte: Identificador da fonte
        limit: Máximo de registros a buscar
        apenas_saude: Se True, filtra apenas editais de saúde
        
    Returns:
        Estatísticas da sincronização
    """
    from services.normalizador_generico import get_normalizador_generico
    from services.multi_source_sync import get_multi_source_sync
    
    fontes_validas = ['comprasnet', 'tce-sp', 'mg-csv', 'pr-csv', 'go-csv']
    
    if fonte.lower() not in fontes_validas:
        raise HTTPException(
            status_code=400,
            detail=f"Fonte inválida. Opções: {fontes_validas}"
        )
    
    try:
        normalizador = get_normalizador_generico(db)
        multi_sync = get_multi_source_sync(db, normalizador)
        
        logger.info(f"🔄 [MULTI-SYNC] Sincronização manual: {fonte}")
        stats = await multi_sync.sync_fonte(fonte.lower(), limit=limit, apenas_saude=apenas_saude)
        
        return {
            "message": f"Sincronização de {fonte} concluída",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro na sincronização de {fonte}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/all")
async def sync_todas_fontes(
    limit_por_fonte: int = Query(50, ge=1, le=100, description="Máximo por fonte"),
    fontes: str = Query(None, description="Fontes específicas (separadas por vírgula)")
):
    """
    🚀 Sincroniza TODAS as fontes de dados
    
    Executa sincronização em sequência de todas as fontes configuradas:
    1. ComprasNet (Federal)
    2. TCE-SP (São Paulo)
    3. MG-CSV (Minas Gerais)
    4. PR-CSV (Paraná)
    5. GO-CSV (Goiás)
    
    Args:
        limit_por_fonte: Máximo de registros por fonte (1-100)
        fontes: Lista específica de fontes (opcional)
        
    Returns:
        Estatísticas consolidadas de todas as fontes
    """
    from services.normalizador_generico import get_normalizador_generico
    from services.multi_source_sync import get_multi_source_sync
    
    try:
        normalizador = get_normalizador_generico(db)
        multi_sync = get_multi_source_sync(db, normalizador)
        
        # Parsear fontes se especificadas
        lista_fontes = None
        if fontes:
            lista_fontes = [f.strip().lower() for f in fontes.split(',')]
        
        logger.info(f"🚀 [MULTI-SYNC] Sincronização completa iniciada...")
        stats = await multi_sync.sync_all(fontes=lista_fontes, limit_por_fonte=limit_por_fonte)
        
        return {
            "message": "Sincronização multi-fonte concluída",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro na sincronização multi-fonte: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/sync/fontes/status")
async def get_status_todas_fontes():
    """
    📊 Retorna status de todas as fontes de dados
    
    Informações por fonte:
    - Status (OK / ERRO / ATRASO / DESCONHECIDO)
    - Última execução
    - Total de registros
    - Detalhes da última sincronização
    """
    from services.normalizador_generico import get_normalizador_generico
    from services.multi_source_sync import get_multi_source_sync
    
    try:
        normalizador = get_normalizador_generico(db)
        multi_sync = get_multi_source_sync(db, normalizador)
        
        status_list = await multi_sync.get_status_fontes()
        
        return {
            "fontes": status_list,
            "total_fontes": len(status_list),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar status das fontes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/normalize/stats/por-fonte")
async def get_normalize_stats_por_fonte():
    """
    📈 Retorna estatísticas de normalização por fonte
    
    Mostra quantos editais de cada fonte foram normalizados
    para o modelo canônico.
    """
    from services.normalizador_generico import get_normalizador_generico
    
    try:
        normalizador = get_normalizador_generico(db)
        stats = await normalizador.get_stats_por_fonte()
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao buscar stats por fonte: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/normalize/backfill/{fonte}")
async def backfill_fonte_especifica(
    fonte: str,
    batch_size: int = Query(100, ge=10, le=500)
):
    """
    🔄 Executa backfill de normalização para uma fonte específica
    
    Processa todos os documentos raw de uma fonte e normaliza
    para o modelo canônico.
    
    Args:
        fonte: Identificador da fonte (pncp, comprasnet, tce-sp, mg-csv, pr-csv, go-csv)
        batch_size: Tamanho do batch de processamento
        
    Returns:
        Estatísticas do backfill
    """
    from services.normalizador_generico import get_normalizador_generico
    
    try:
        normalizador = get_normalizador_generico(db)
        
        logger.info(f"🔄 [BACKFILL] Iniciando para fonte: {fonte}")
        stats = await normalizador.backfill_por_fonte(fonte=fonte, batch_size=batch_size)
        
        return {
            "message": f"Backfill de {fonte} concluído",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no backfill de {fonte}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE EMAIL (P1) ====================

@api_router.get("/email/status")
async def get_email_status():
    """
    Retorna status da configuração do serviço de email
    
    Informações:
    - Se Resend está configurado
    - Email de origem
    - Status da API key
    """
    email_service = get_email_service()
    return email_service.get_status()


@api_router.post("/email/test")
async def test_email_sending(
    destinatario: str = Query(..., description="Email para enviar teste"),
    assunto: str = Query("Teste GSM Buscador", description="Assunto do email")
):
    """
    Envia email de teste para verificar a configuração
    
    Args:
        destinatario: Email do destinatário
        assunto: Assunto do email de teste
    """
    try:
        email_service = get_email_service()
        
        # Dados de teste
        licitacoes_teste = [
            {
                'objeto': 'AQUISIÇÃO DE MEDICAMENTOS HOSPITALARES - TESTE',
                'orgao': 'Secretaria Municipal de Saúde - TESTE',
                'estado': 'SP',
                'link_origem': 'https://pncp.gov.br/teste'
            }
        ]
        
        resultado = await email_service.enviar_alerta_licitacoes(
            destinatario=destinatario,
            palavra_chave="Teste de Configuração",
            licitacoes=licitacoes_teste,
            nome_alerta="Email de Teste GSM"
        )
        
        return {
            "message": "Email de teste processado",
            "resultado": resultado,
            "status_servico": email_service.get_status()
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao enviar email de teste: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE ALERTAS POR EMAIL (P5) ====================

from models.alerta import AlertaCreate, AlertaUpdate, AlertaResponse
from services.alerta_service import get_alerta_service

@api_router.get("/alertas-email")
async def listar_alertas_email(
    email: str = Query(None, description="Filtrar por email específico")
):
    """
    🔔 P5: Lista todos os alertas de email cadastrados
    
    Args:
        email: Filtro opcional por email
        
    Returns:
        Lista de alertas com estatísticas
    """
    try:
        alerta_service = get_alerta_service(db)
        alertas = await alerta_service.listar_alertas(email=email)
        
        return {
            "total": len(alertas),
            "alertas": alertas
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar alertas email: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/alertas-email", status_code=201)
async def criar_alerta_email(alerta: AlertaCreate):
    """
    🔔 P5: Cria um novo alerta de email para termo de busca
    
    Args:
        alerta: Dados do alerta (email, termo, frequência, filtros)
        
    Returns:
        Alerta criado com ID
        
    Regras de Envio:
    - Apenas oportunidades ATIVAS
    - Quality score >= 70
    - Sem duplicatas (controle por edital_id)
    """
    try:
        alerta_service = get_alerta_service(db)
        novo_alerta = await alerta_service.criar_alerta(alerta)
        
        logger.info(f"✅ [P5] Alerta criado: '{alerta.termo}' para {alerta.email}")
        
        return {
            "message": "Alerta criado com sucesso",
            "alerta": novo_alerta
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao criar alerta email: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/alertas-email/{alerta_id}")
async def obter_alerta_email(alerta_id: str):
    """
    🔔 P5: Obtém detalhes de um alerta específico
    """
    alerta_service = get_alerta_service(db)
    alerta = await alerta_service.obter_alerta(alerta_id)
    
    if not alerta:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    
    # Converter para formato seguro (sem _id)
    return {
        "id": alerta.get("id"),
        "email": alerta.get("email"),
        "termo": alerta.get("termo"),
        "frequencia": alerta.get("frequencia"),
        "ativo": alerta.get("ativo"),
        "filtros": alerta.get("filtros", {}),
        "ultimo_envio": alerta.get("ultimo_envio").isoformat() if alerta.get("ultimo_envio") else None,
        "total_enviados": alerta.get("total_enviados", 0),
        "created_at": alerta.get("created_at").isoformat() if alerta.get("created_at") else None
    }


@api_router.put("/alertas-email/{alerta_id}")
async def atualizar_alerta_email(alerta_id: str, update: AlertaUpdate):
    """
    🔔 P5: Atualiza configurações de um alerta
    
    Args:
        alerta_id: ID do alerta
        update: Campos a atualizar (ativo, frequência, filtros)
    """
    try:
        alerta_service = get_alerta_service(db)
        alerta = await alerta_service.atualizar_alerta(alerta_id, update)
        
        if not alerta:
            raise HTTPException(status_code=404, detail="Alerta não encontrado")
        
        return {
            "message": "Alerta atualizado",
            "alerta": alerta
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao atualizar alerta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/alertas-email/{alerta_id}")
async def deletar_alerta_email(alerta_id: str):
    """
    🔔 P5: Remove um alerta de email
    """
    alerta_service = get_alerta_service(db)
    deletado = await alerta_service.deletar_alerta(alerta_id)
    
    if not deletado:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    
    return {"message": "Alerta deletado com sucesso", "id": alerta_id}


@api_router.post("/alertas-email/{alerta_id}/processar")
async def processar_alerta_email(alerta_id: str):
    """
    🔔 P5: Processa um alerta manualmente (força verificação e envio)
    
    Útil para testar o sistema ou forçar envio imediato.
    
    Pipeline:
    1. Busca oportunidades ATIVAS com quality >= 70
    2. Filtra editais já enviados
    3. Envia email se houver novidades
    4. Atualiza registro de envios
    """
    try:
        alerta_service = get_alerta_service(db)
        resultado = await alerta_service.processar_alerta(alerta_id)
        
        return resultado
        
    except Exception as e:
        logger.error(f"❌ Erro ao processar alerta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/alertas-email/processar-todos")
async def processar_todos_alertas_email(
    frequencia: str = Query(None, description="Filtrar por frequência: diario ou semanal")
):
    """
    🔔 P5: Processa todos os alertas ativos (job manual)
    
    Args:
        frequencia: Filtro opcional por frequência
        
    Returns:
        Estatísticas do processamento
    """
    try:
        alerta_service = get_alerta_service(db)
        stats = await alerta_service.processar_todos_alertas(frequencia=frequencia)
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao processar alertas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/alertas-email/teste")
async def enviar_teste_alerta_email(
    email: str = Query(..., description="Email para enviar teste"),
    termo: str = Query("insulina", description="Termo de busca para simular")
):
    """
    🔔 P5: Envia email de teste para verificar configuração do Resend
    
    Simula um alerta real buscando oportunidades e enviando o email.
    
    Args:
        email: Email do destinatário
        termo: Termo de busca para simular
    """
    try:
        email_service = get_email_service()
        
        # Verificar status primeiro
        status = email_service.get_status()
        if not status.get("configurado"):
            return {
                "status": "error",
                "message": "RESEND_API_KEY não configurada. Adicione a chave no arquivo .env",
                "config": status
            }
        
        # Enviar email de teste
        resultado = await email_service.enviar_teste(email)
        
        return {
            "status": resultado.get("status"),
            "message": resultado.get("message"),
            "email_id": resultado.get("email_id"),
            "config": status
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no teste de email: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/alertas-email/teste-completo")
async def enviar_teste_alerta_completo(
    email: str = Query(..., description="Email para enviar teste"),
    termo: str = Query("canabidiol", description="Termo de busca para simular")
):
    """
    🔔 P5: TESTE COMPLETO DE ALERTA COM DADOS REAIS
    
    Busca oportunidades REAIS da API PNCP e envia email com template completo.
    Inclui:
    - CC obrigatório (hudson@vipfarma.com.br, claudio@gruposmartmedical.com.br)
    - Itens extraídos do edital
    - Links diretos para PDFs
    - Todos campos obrigatórios (Nº Item, Descrição, Qtd, Valor, Nº Processo, Data)
    
    Args:
        email: Email do destinatário
        termo: Termo de busca para simular
    """
    try:
        email_service = get_email_service()
        
        # Verificar status primeiro
        status = email_service.get_status()
        if not status.get("configurado"):
            return {
                "status": "error",
                "message": "RESEND_API_KEY não configurada. Adicione a chave no arquivo .env",
                "config": status
            }
        
        # 1. Buscar oportunidades REAIS usando o serviço de busca
        from services.busca_service_v2 import get_busca_service_v2
        busca_service = get_busca_service_v2(db)
        
        logger.info(f"🔍 [TESTE-COMPLETO] Buscando oportunidades para '{termo}'...")
        
        resultados = await busca_service.buscar(
            termo_busca=termo,
            keywords=[termo],
            estados=None,
            modalidade=None,
            esfera=None,
            apenas_saude=False,
            limit=10,
            skip=0,
            expandir_termos=True,
            incluir_historico=False,
            periodo_dias=90,
            incluir_ativas=True,
            incluir_futuras=True,
            incluir_encerradas=False,
            excluir_credenciamentos=False,
            incluir_suspeitos=False,
            incluir_planejamento=False,
            limite_quality_score=50
        )
        
        editais = resultados.get("resultados", [])
        
        if not editais:
            # Se não houver resultados, criar dados de demonstração
            logger.warning(f"⚠️ [TESTE-COMPLETO] Nenhum resultado para '{termo}', usando dados demo")
            editais = [
                {
                    "orgao": "SECRETARIA DE SAÚDE DO ESTADO - DEMONSTRAÇÃO",
                    "numero_processo": "PE-2024/001234",
                    "numero_edital": "123/2024",
                    "modalidade": "Pregão Eletrônico",
                    "objeto": f"Aquisição de medicamentos contendo {termo} para atender demanda da rede de saúde",
                    "status_oportunidade": "ATIVA",
                    "uf": "SP",
                    "municipio": "São Paulo",
                    "data_abertura": "2025-01-15T10:00:00",
                    "link_sistema_origem": "https://pncp.gov.br/app/editais",
                    "itens_correspondentes": [
                        {
                            "numero_item": "1",
                            "descricao": f"Medicamento à base de {termo} 200mg - Frasco 30ml",
                            "quantidade": "500",
                            "unidade": "UN",
                            "valor_unitario": "1250.00",
                            "valor_total": "625000.00",
                            "fonte": "DEMONSTRAÇÃO"
                        },
                        {
                            "numero_item": "2", 
                            "descricao": f"Solução oral de {termo} 100mg/ml",
                            "quantidade": "200",
                            "unidade": "FR",
                            "valor_unitario": "850.00",
                            "valor_total": "170000.00",
                            "fonte": "DEMONSTRAÇÃO"
                        }
                    ]
                }
            ]
        else:
            logger.info(f"✅ [TESTE-COMPLETO] Encontradas {len(editais)} oportunidades reais")
        
        # 2. Enviar email com template completo e CC obrigatório
        logger.info(f"📧 [TESTE-COMPLETO] Enviando email para {email} com CC obrigatório...")
        
        resultado = await email_service.enviar_alerta(
            destinatario=email,
            termo=termo,
            editais=editais[:5]  # Limitar a 5 editais
        )
        
        return {
            "status": resultado.get("status"),
            "message": resultado.get("message"),
            "email_id": resultado.get("email_id"),
            "editais_enviados": resultado.get("editais_enviados", len(editais[:5])),
            "cc": resultado.get("cc", email_service.CC_OBRIGATORIO),
            "termo_busca": termo,
            "editais_encontrados": len(editais),
            "config": status,
            "nota": "⚠️ LIMITAÇÃO RESEND: No ambiente de teste (onboarding@resend.dev), o CC só funciona para emails verificados. Em produção com domínio verificado, todos os CCs funcionarão."
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no teste completo de email: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/status/scrapers/{fonte}/test")
async def test_scraper(fonte: str, medicamento: str = "medicamento"):
    """
    Executa um teste manual em um scraper específico
    """
    try:
        logger.info(f"🧪 Iniciando teste manual do scraper: {fonte}")
        
        # Mapeamento de fontes para métodos de teste específicos
        # Se não houver método específico, usa o padrão do scraper_service
        resultados = []
        inicio = time.time()
        
        # Tentar executar o scraper correspondente
        if fonte == 'PNCP':
            resultados = await scraper_service.buscar_apenas_pncp(medicamento, apenas_futuras=True)
        elif fonte == 'PNCP-OFICIAL':
            resultados = await scraper_service.pncp_api_oficial.buscar_licitacoes(medicamento, apenas_futuras=True)
        elif fonte == 'ComprasNet':
            resultados = await scraper_service.buscar_apenas_comprasnet(medicamento, apenas_futuras=True)
        elif fonte == 'BEC/SP':
            resultados = await scraper_service.buscar_apenas_bec_sp(medicamento, apenas_futuras=True)
        elif fonte == 'RJ':
            resultados = await scraper_service.buscar_apenas_rj(medicamento, apenas_futuras=True)
        elif fonte == 'RS':
            resultados = await scraper_service.buscar_apenas_rs(medicamento, apenas_futuras=True)
        elif fonte == 'SC':
            resultados = await scraper_service.buscar_apenas_sc(medicamento, apenas_futuras=True)
        elif fonte == 'PR':
            resultados = await scraper_service.buscar_apenas_pr(medicamento)
        elif fonte == 'BA':
            resultados = await scraper_service.buscar_apenas_ba(medicamento)
        elif fonte == 'PE':
            resultados = await scraper_service.buscar_apenas_pe(medicamento)
        elif fonte == 'SP-TCE':
            resultados = await scraper_service.buscar_apenas_sp_tce(medicamento)
        elif fonte == 'MG':
            resultados = await scraper_service.buscar_apenas_mg(medicamento)
        elif fonte == 'GO':
            resultados = await scraper_service.buscar_apenas_go(medicamento)
        elif fonte == 'ES-CSV':
            resultados = await scraper_service.buscar_apenas_es(medicamento)
        elif fonte == 'AGREGADOR':
            resultados = await scraper_service.agregador_client.buscar_licitacoes(medicamento)
        elif fonte.startswith('SCRAPER-'):
            uf = fonte.split('-')[1]
            resultados = await scraper_service.refresh_estado(uf, medicamento)
        else:
            raise HTTPException(status_code=400, detail=f"Fonte '{fonte}' não suporta teste manual individual.")

        tempo_ms = int((time.time() - inicio) * 1000)
        
        # Registrar manualmente no health monitor para atualizar o status IMEDIATAMENTE
        await health_monitor.registrar_execucao(
            fonte=fonte,
            status='success',
            resultados_count=len(resultados),
            termo_busca=medicamento,
            tempo_execucao_ms=tempo_ms
        )
        
        return {
            "fonte": fonte,
            "status": "success",
            "resultados_count": len(resultados),
            "tempo_ms": tempo_ms,
            "resultados_amostra": resultados[:3] if resultados else []
        }
        
    except Exception as e:
        logger.error(f"❌ Erro no teste manual de {fonte}: {str(e)}")
        
        # Registrar erro no health monitor
        await health_monitor.registrar_execucao(
            fonte=fonte,
            status='error',
            erro_mensagem=str(e)
        )
        
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao executar teste manual: {str(e)}"
        )



@api_router.get("/search/export-excel")
async def exportar_resultados_excel(
    medicamento: str = Query(None),
    estado: str = Query(None),
    status: str = Query(None),
    modalidade: str = Query(None),
    esfera: str = Query(None),
    lista_id: str = Query(None)
):
    """
    Exporta resultados de busca em formato Excel (.xlsx) com logo GSM (P4).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XlImage
        from io import BytesIO

        # Reutilizar lógica de filtros
        filtros = {}
        if medicamento:
            filtros['medicamento'] = {'$regex': medicamento, '$options': 'i'}
        if estado:
            filtros['estado'] = estado
        if status:
            filtros['status'] = {'$regex': status, '$options': 'i'}
        if modalidade:
            filtros['modalidade'] = {'$regex': modalidade, '$options': 'i'}
        if esfera:
            filtros['esfera'] = esfera

        if lista_id:
            lista = await db.listas_medicamentos.find_one({'id': lista_id}, {'_id': 0})
            if lista and lista.get('medicamentos'):
                filtros['medicamento'] = {'$in': [{'$regex': med, '$options': 'i'} for med in lista['medicamentos']]}

        # Buscar resultados (limite de 5000 para Excel)
        resultados = await db.licitacoes.find(filtros, {'_id': 0}).to_list(5000)

        if not resultados:
            raise HTTPException(status_code=404, detail="Nenhum resultado encontrado para exportação")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados GSM"

        # Estilos (Padrão Premium P4)
        verde_escuro = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        cinza_header = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        font_branca = Font(color="FFFFFF", bold=True, size=11)
        font_titulo = Font(bold=True, size=14, color="1B5E20")
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CFD8DC'),
            right=Side(style='thin', color='CFD8DC'),
            top=Side(style='thin', color='CFD8DC'),
            bottom=Side(style='thin', color='CFD8DC')
        )

        # Header com Logo
        logo_path = Path(__file__).parent / "assets" / "logo_gsm.png"
        if logo_path.exists():
            img = XlImage(str(logo_path))
            img.width = 180
            img.height = 70
            ws.add_image(img, "A1")

        ws.merge_cells("A6:I6")
        ws["A6"] = "GSM - RELATÓRIO DE OPORTUNIDADES DE BUSCA"
        ws["A6"].font = font_titulo

        # Metadados da busca
        ws["A7"] = f"Termo: {medicamento or 'Todos'}"
        ws["A8"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A9"] = f"Total de registros: {len(resultados)}"

        # Cabeçalhos da tabela (Linha 11)
        headers = ["Medicamento", "Órgão", "UF", "Modalidade", "Status", "Data Abertura", "Esfera", "Número Processo", "Link"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=11, column=i, value=h)
            cell.font = font_branca
            cell.fill = cinza_header
            cell.alignment = align_center
            cell.border = thin_border

        # Dados
        for idx, res in enumerate(resultados):
            row_idx = 12 + idx
            data_abertura = res.get('data_abertura')
            if isinstance(data_abertura, datetime):
                data_str = data_abertura.strftime('%d/%m/%Y')
            else:
                data_str = str(data_abertura or '')

            ws.cell(row=row_idx, column=1, value=res.get('medicamento', '')).border = thin_border
            ws.cell(row=row_idx, column=2, value=res.get('orgao_licitante', '')).border = thin_border
            ws.cell(row=row_idx, column=3, value=res.get('estado', '')).border = thin_border
            ws.cell(row=row_idx, column=4, value=res.get('modalidade', '')).border = thin_border
            ws.cell(row=row_idx, column=5, value=res.get('status', '')).border = thin_border
            ws.cell(row=row_idx, column=6, value=data_str).border = thin_border
            ws.cell(row=row_idx, column=7, value=res.get('esfera', '')).border = thin_border
            ws.cell(row=row_idx, column=8, value=res.get('numero_processo', '')).border = thin_border
            
            link = res.get('link_origem', '')
            cell_link = ws.cell(row=row_idx, column=9, value="Ver Portal")
            if link:
                cell_link.hyperlink = link
                cell_link.font = Font(color="0000FF", underline="single")
            cell_link.border = thin_border

        # Ajustar larguras
        col_widths = [30, 40, 5, 20, 15, 12, 12, 20, 15]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Salvar em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"gsm_busca_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Length': str(len(output.getvalue()))
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na exportação Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/export")
async def export_results(
    formato: str = Query(..., description="Formato de exportação: csv ou json"),
    medicamento: str = Query(None),
    estado: str = Query(None),
    status: str = Query(None),
    modalidade: str = Query(None),
    esfera: str = Query(None),
    lista_id: str = Query(None)
):
    """
    Exporta resultados de busca em CSV ou JSON
    
    Args:
        formato: 'csv' ou 'json'
        medicamento: Termo de busca (opcional)
        estado: Filtro por estado (opcional)
        status: Filtro por status (opcional)
        modalidade: Filtro por modalidade (opcional)
        esfera: Filtro por esfera (opcional)
        lista_id: ID de lista customizada (opcional)
    
    Returns:
        StreamingResponse com arquivo CSV ou JSON
    """
    try:
        logger.info(f"📥 Exportação solicitada - Formato: {formato}")
        
        # Validar formato
        if formato not in ['csv', 'json']:
            raise HTTPException(
                status_code=400,
                detail="Formato inválido. Use 'csv' ou 'json'"
            )
        
        # Construir query de busca (reutilizar lógica do /search)
        filtros = {}
        
        if medicamento:
            filtros['medicamento'] = {'$regex': medicamento, '$options': 'i'}
        
        if estado:
            filtros['estado'] = estado
        
        if status:
            filtros['status'] = {'$regex': status, '$options': 'i'}
        
        if modalidade:
            filtros['modalidade'] = {'$regex': modalidade, '$options': 'i'}
        
        if esfera:
            filtros['esfera'] = esfera
        
        # Se tem lista, buscar medicamentos da lista
        medicamentos_lista = []
        if lista_id:
            lista = await db.listas_medicamentos.find_one(
                {'id': lista_id},
                {'_id': 0}
            )
            if lista:
                medicamentos_lista = lista.get('medicamentos', [])
                if medicamentos_lista:
                    filtros['medicamento'] = {
                        '$in': [
                            {'$regex': med, '$options': 'i'} 
                            for med in medicamentos_lista
                        ]
                    }
        
        # Buscar resultados
        logger.info(f"   Aplicando filtros: {filtros}")
        resultados = await db.licitacoes.find(filtros, {'_id': 0}).to_list(10000)
        
        logger.info(f"   ✅ {len(resultados)} resultados encontrados")
        
        if not resultados:
            raise HTTPException(
                status_code=404,
                detail="Nenhum resultado encontrado com os filtros aplicados"
            )
        
        # Gerar arquivo
        if formato == 'csv':
            return _gerar_csv(resultados)
        else:
            return _gerar_json(resultados)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na exportação: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao exportar dados: {str(e)}"
        )


def _gerar_csv(resultados: List[Dict]) -> StreamingResponse:
    """
    Gera arquivo CSV com os resultados
    
    Args:
        resultados: Lista de licitações
        
    Returns:
        StreamingResponse com CSV
    """
    output = StringIO()
    
    # Definir campos do CSV (principais campos)
    campos = [
        'id', 'medicamento', 'estado', 'status', 'orgao_licitante',
        'modalidade', 'numero_processo', 'data_abertura', 'data_final',
        'link_origem', 'link_documento', 'fonte', 'esfera', 'objeto'
    ]
    
    writer = csv.DictWriter(output, fieldnames=campos, extrasaction='ignore')
    writer.writeheader()
    
    # Escrever resultados
    for resultado in resultados:
        # Converter datas para string
        row = resultado.copy()
        for campo in ['data_abertura', 'data_final', 'data_publicacao', 'data_referencia']:
            if campo in row and row[campo]:
                if isinstance(row[campo], datetime):
                    row[campo] = row[campo].strftime('%Y-%m-%d %H:%M:%S')
        
        writer.writerow(row)
    
    # Preparar streaming response
    output.seek(0)
    
    filename = f"gsm_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Length': str(len(output.getvalue()))
        }
    )


def _gerar_json(resultados: List[Dict]) -> StreamingResponse:
    """
    Gera arquivo JSON com os resultados
    
    Args:
        resultados: Lista de licitações
        
    Returns:
        StreamingResponse com JSON
    """
    # Converter datas para string para serialização JSON
    for resultado in resultados:
        for campo in ['data_abertura', 'data_final', 'data_publicacao', 'data_referencia']:
            if campo in resultado and resultado[campo]:
                if isinstance(resultado[campo], datetime):
                    resultado[campo] = resultado[campo].isoformat()
    
    # Gerar JSON
    json_data = json.dumps(
        {
            'total': len(resultados),
            'data_exportacao': datetime.now().isoformat(),
            'resultados': resultados
        },
        ensure_ascii=False,
        indent=2
    )
    
    return StreamingResponse(
        iter([json_data]),
        media_type='application/json',
        headers={
            'Content-Disposition': f'attachment; filename="licitacoes_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        }
    )


def _get_estado_nome(uf: str) -> str:
    """Retorna nome completo do estado"""
    estados = {
        'AC': 'Acre', 'AL': 'Alagoas', 'AP': 'Amapá', 'AM': 'Amazonas',
        'BA': 'Bahia', 'CE': 'Ceará', 'DF': 'Distrito Federal', 'ES': 'Espírito Santo',
        'GO': 'Goiás', 'MA': 'Maranhão', 'MT': 'Mato Grosso', 'MS': 'Mato Grosso do Sul',
        'MG': 'Minas Gerais', 'PA': 'Pará', 'PB': 'Paraíba', 'PR': 'Paraná',
        'PE': 'Pernambuco', 'PI': 'Piauí', 'RJ': 'Rio de Janeiro', 'RN': 'Rio Grande do Norte',
        'RS': 'Rio Grande do Sul', 'RO': 'Rondônia', 'RR': 'Roraima', 'SC': 'Santa Catarina',
        'SP': 'São Paulo', 'SE': 'Sergipe', 'TO': 'Tocantins'
    }
    return estados.get(uf, uf)


# ==================== ENDPOINTS DE NOTIFICAÇÕES (P2) ====================

@api_router.get("/alertas", response_model=AlertaListResponse)
async def listar_alertas(apenas_ativos: bool = Query(False, description="Filtrar apenas alertas ativos")):
    """
    Lista todos os alertas de notificação configurados
    
    Args:
        apenas_ativos: Se True, retorna apenas alertas ativos
        
    Returns:
        Lista de alertas com metadados
    """
    try:
        alertas = await notificacao_service.listar_alertas(apenas_ativos=apenas_ativos)
        ativos = sum(1 for a in alertas if a.ativo)
        
        return AlertaListResponse(
            alertas=alertas,
            total=len(alertas),
            ativos=ativos
        )
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar alertas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/alertas", response_model=AlertaConfig, status_code=201)
async def criar_alerta(alerta: AlertaConfigCreate):
    """
    Cria um novo alerta de notificação
    
    Args:
        alerta: Configuração do alerta (nome, palavras-chave, filtros)
        
    Returns:
        Alerta criado com ID
        
    Raises:
        400: Se limite de alertas atingido ou dados inválidos
    """
    try:
        novo_alerta = await notificacao_service.criar_alerta(alerta)
        return novo_alerta
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"❌ Erro ao criar alerta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/alertas/{alerta_id}", response_model=AlertaConfig)
async def obter_alerta(alerta_id: str):
    """Obtém um alerta específico por ID"""
    alerta = await notificacao_service.obter_alerta(alerta_id)
    
    if not alerta:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    
    return alerta


@api_router.put("/alertas/{alerta_id}", response_model=AlertaConfig)
async def atualizar_alerta(alerta_id: str, update_data: AlertaConfigUpdate):
    """
    Atualiza um alerta existente
    
    Args:
        alerta_id: ID do alerta
        update_data: Campos a atualizar
        
    Returns:
        Alerta atualizado
    """
    try:
        alerta = await notificacao_service.atualizar_alerta(alerta_id, update_data)
        
        if not alerta:
            raise HTTPException(status_code=404, detail="Alerta não encontrado")
        
        return alerta
        
    except Exception as e:
        logger.error(f"❌ Erro ao atualizar alerta: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/alertas/{alerta_id}")
async def deletar_alerta(alerta_id: str):
    """Deleta um alerta e suas notificações"""
    deletado = await notificacao_service.deletar_alerta(alerta_id)
    
    if not deletado:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    
    return {"message": "Alerta deletado com sucesso", "id": alerta_id}


@api_router.get("/notificacoes", response_model=NotificacaoListResponse)
async def listar_notificacoes(
    status: Optional[str] = Query(None, description="Filtrar por status: pendente, lida, arquivada"),
    alerta_id: Optional[str] = Query(None, description="Filtrar por alerta específico"),
    pagina: int = Query(1, ge=1, description="Número da página"),
    por_pagina: int = Query(20, ge=1, le=100, description="Itens por página")
):
    """
    Lista notificações com filtros e paginação
    
    Args:
        status: Filtrar por status
        alerta_id: Filtrar por alerta
        pagina: Página atual
        por_pagina: Itens por página
        
    Returns:
        Lista de notificações com metadados de paginação
    """
    try:
        # Converter status string para enum
        status_enum = None
        if status:
            try:
                status_enum = StatusNotificacao(status)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Status inválido: {status}")
        
        result = await notificacao_service.listar_notificacoes(
            status=status_enum,
            alerta_id=alerta_id,
            limite=por_pagina,
            pagina=pagina
        )
        
        return NotificacaoListResponse(**result)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao listar notificações: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/notificacoes/stats", response_model=NotificacaoStats)
async def obter_estatisticas_notificacoes():
    """
    Obtém estatísticas de notificações
    
    Returns:
        Contadores de notificações e status dos alertas
    """
    try:
        stats = await notificacao_service.obter_estatisticas()
        return stats
        
    except Exception as e:
        logger.error(f"❌ Erro ao obter estatísticas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/notificacoes/{notificacao_id}/lida")
async def marcar_notificacao_lida(notificacao_id: str):
    """Marca uma notificação como lida"""
    sucesso = await notificacao_service.marcar_como_lida(notificacao_id)
    
    if not sucesso:
        raise HTTPException(status_code=404, detail="Notificação não encontrada")
    
    return {"message": "Notificação marcada como lida", "id": notificacao_id}


@api_router.post("/notificacoes/marcar-todas-lidas")
async def marcar_todas_lidas(alerta_id: Optional[str] = Query(None)):
    """
    Marca todas as notificações pendentes como lidas
    
    Args:
        alerta_id: Se informado, marca apenas notificações deste alerta
    """
    total = await notificacao_service.marcar_todas_como_lidas(alerta_id)
    return {"message": f"{total} notificações marcadas como lidas", "total": total}


@api_router.post("/notificacoes/{notificacao_id}/arquivar")
async def arquivar_notificacao(notificacao_id: str):
    """Arquiva uma notificação"""
    sucesso = await notificacao_service.arquivar_notificacao(notificacao_id)
    
    if not sucesso:
        raise HTTPException(status_code=404, detail="Notificação não encontrada")
    
    return {"message": "Notificação arquivada", "id": notificacao_id}


@api_router.post("/notificacoes/verificar")
async def verificar_novas_licitacoes(forcar: bool = Query(False, description="Forçar verificação ignorando frequência")):
    """
    Dispara verificação manual de novas licitações para alertas
    
    Este endpoint é chamado manualmente ou por um job scheduler.
    Verifica todos os alertas ativos e cria notificações para licitações
    que correspondem aos critérios.
    
    Args:
        forcar: Se True, ignora o intervalo de frequência dos alertas
        
    Returns:
        Resultado da verificação com contadores
    """
    try:
        resultado = await notificacao_service.verificar_novas_licitacoes(forcar=forcar)
        return resultado
        
    except Exception as e:
        logger.error(f"❌ Erro na verificação: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE RADARES (v25.1) ====================

class RadarCreate(BaseModel):
    """Schema para criação de radar"""
    nome: str = Field(..., min_length=2, max_length=100)
    email: str = Field(..., description="E-mail para notificações")
    termos: str = Field(..., description="Palavras-chave separadas por vírgula")
    frequencia: str = Field(default="24h", description="8h, 12h ou 24h")


class RadarUpdate(BaseModel):
    """Schema para atualização de radar"""
    nome: Optional[str] = None
    email: Optional[str] = None
    termos: Optional[str] = None
    frequencia: Optional[str] = None
    ativo: Optional[bool] = None


# ==================== EMPRESAS v46.0 ====================

class EmpresaGSM(BaseModel):
    """Schema para empresa do Grupo Smart Medical"""
    id: str
    name: str
    cnpj: str
    ie: Optional[str] = ""
    address: Optional[str] = ""
    phone: Optional[str] = ""
    whatsapp: Optional[str] = ""
    email: Optional[str] = ""


@api_router.get("/empresas")
async def listar_empresas():
    """
    🏢 Lista todas as empresas cadastradas (v46.0)
    """
    try:
        empresas = await db.empresas.find({}, {"_id": 0}).to_list(length=20)
        
        # Se não houver empresas, retornar defaults
        if not empresas:
            empresas = [
                {
                    "id": "c1",
                    "name": "HC IMPORTAÇÕES EXPORTAÇÕES LTDA",
                    "cnpj": "31.958.700/0001-17",
                    "ie": "084.050.99-3",
                    "address": "Rua Domingos Dadalto, 127, Galpão 03, Rio Branco, Cariacica - ES",
                    "phone": "(11) 3164-4607",
                    "whatsapp": "(11) 99989-2696",
                    "email": "claudio@gruposmartmedical.com.br"
                }
            ]
        
        return {
            "total": len(empresas),
            "empresas": empresas
        }
    except Exception as e:
        logger.error(f"❌ Erro ao listar empresas: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/empresas/salvar")
async def salvar_empresa(empresa: EmpresaGSM):
    """
    🏢 Salva ou atualiza uma empresa (v46.0)
    
    Persistência no MongoDB para uso no DAMA IA.
    """
    try:
        empresa_dict = empresa.dict()
        empresa_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Upsert: atualiza se existir, insere se não
        result = await db.empresas.update_one(
            {"id": empresa.id},
            {"$set": empresa_dict},
            upsert=True
        )
        
        logger.info(f"✅ Empresa salva: {empresa.id} - {empresa.name}")
        
        return {
            "message": "Empresa salva com sucesso",
            "empresa": empresa_dict
        }
    except Exception as e:
        logger.error(f"❌ Erro ao salvar empresa: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/empresas/{empresa_id}")
async def deletar_empresa(empresa_id: str):
    """
    🏢 Deleta uma empresa (v46.0)
    """
    try:
        result = await db.empresas.delete_one({"id": empresa_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Empresa não encontrada")
        
        return {
            "message": "Empresa deletada com sucesso",
            "empresa_id": empresa_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao deletar empresa: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== v63.0: UPLOAD DE TIMBRADO ====================

@api_router.post("/empresas/{empresa_id}/timbrado")
async def upload_timbrado_empresa(
    empresa_id: str,
    timbrado: UploadFile = File(..., description="Arquivo .docx do papel timbrado")
):
    """
    📄 v63.0 - Upload do papel timbrado (.docx) para uma empresa
    
    O timbrado é salvo no disco e o caminho é associado à empresa.
    Este arquivo será usado automaticamente no DAMA IA.
    """
    try:
        # Validar extensão
        if not timbrado.filename.lower().endswith('.docx'):
            raise HTTPException(status_code=400, detail="O arquivo deve ser .docx")
        
        # Verificar se empresa existe
        empresa = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        if not empresa:
            # Se não existe, criar com dados mínimos
            empresa = {"id": empresa_id, "name": "", "cnpj": ""}
        
        # Ler conteúdo do arquivo
        content = await timbrado.read()
        
        # Salvar no disco com nome único baseado no ID da empresa
        filename = f"{empresa_id}_timbrado.docx"
        filepath = UPLOADS_DIR / filename
        
        with open(filepath, "wb") as f:
            f.write(content)
        
        logger.info(f"📄 [v63.0] Timbrado salvo: {filepath} ({len(content)/1024:.1f}KB)")
        
        # Atualizar empresa com caminho do timbrado
        await db.empresas.update_one(
            {"id": empresa_id},
            {
                "$set": {
                    "timbrado_path": str(filepath),
                    "timbrado_nome": timbrado.filename,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            },
            upsert=True
        )
        
        return {
            "message": "Timbrado salvo com sucesso",
            "empresa_id": empresa_id,
            "timbrado_nome": timbrado.filename,
            "timbrado_size_kb": round(len(content) / 1024, 1)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao salvar timbrado: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/empresas/{empresa_id}/timbrado")
async def verificar_timbrado_empresa(empresa_id: str):
    """
    📄 v63.0 - Verifica se empresa tem timbrado cadastrado
    """
    try:
        empresa = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        
        if not empresa:
            return {"tem_timbrado": False, "empresa_id": empresa_id}
        
        timbrado_path = empresa.get("timbrado_path")
        timbrado_nome = empresa.get("timbrado_nome")
        
        # Verificar se arquivo existe no disco
        tem_timbrado = False
        if timbrado_path:
            tem_timbrado = Path(timbrado_path).exists()
        
        return {
            "tem_timbrado": tem_timbrado,
            "empresa_id": empresa_id,
            "timbrado_nome": timbrado_nome if tem_timbrado else None
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao verificar timbrado: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/empresas/{empresa_id}/timbrado")
async def excluir_timbrado_empresa(empresa_id: str):
    """
    📄 v65.1 - Excluir papel timbrado de uma empresa
    """
    try:
        # Buscar empresa
        empresa = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        
        if not empresa:
            raise HTTPException(status_code=404, detail="Empresa não encontrada")
        
        timbrado_path = empresa.get("timbrado_path")
        
        # Excluir arquivo do disco se existir
        if timbrado_path and Path(timbrado_path).exists():
            Path(timbrado_path).unlink()
            logger.info(f"🗑️ [v65.1] Timbrado excluído: {timbrado_path}")
        
        # Atualizar empresa no banco
        await db.empresas.update_one(
            {"id": empresa_id},
            {
                "$unset": {
                    "timbrado_path": "",
                    "timbrado_nome": ""
                },
                "$set": {
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        return {
            "message": "Timbrado excluído com sucesso",
            "empresa_id": empresa_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao excluir timbrado: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/radares")
async def listar_radares():
    """
    🎯 Lista todos os radares configurados (v25.1)
    
    Radares são alertas com e-mail e frequência customizáveis.
    """
    try:
        radares = await db.radares.find(
            {},
            {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        
        return {
            "total": len(radares),
            "radares": radares
        }
    except Exception as e:
        logger.error(f"❌ Erro ao listar radares: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/radares", status_code=201)
async def criar_radar(radar: RadarCreate):
    """
    🎯 Cria um novo radar de monitoramento (v25.1)
    
    O radar será executado automaticamente na frequência configurada.
    """
    try:
        novo_radar = {
            "id": str(uuid.uuid4()),
            "nome": radar.nome,
            "email": radar.email,
            "termos": radar.termos,
            "frequencia": radar.frequencia,
            "ativo": True,
            "ultimo_envio": None,
            "total_enviados": 0,
            "editais_enviados": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.radares.insert_one(novo_radar)
        
        # Remover _id para resposta
        novo_radar.pop('_id', None)
        
        logger.info(f"✅ Radar criado: {radar.nome} → {radar.email}")
        
        return {
            "message": "Radar criado com sucesso",
            "radar": novo_radar
        }
    except Exception as e:
        logger.error(f"❌ Erro ao criar radar: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/radares/{radar_id}")
async def obter_radar(radar_id: str):
    """Busca um radar específico por ID"""
    try:
        radar = await db.radares.find_one(
            {"id": radar_id},
            {"_id": 0}
        )
        
        if not radar:
            raise HTTPException(status_code=404, detail="Radar não encontrado")
        
        return {"radar": radar}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao buscar radar: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/radares/{radar_id}")
async def atualizar_radar(radar_id: str, radar: RadarUpdate):
    """
    🎯 Atualiza um radar existente (v25.1)
    """
    try:
        # Verificar se radar existe
        existente = await db.radares.find_one({"id": radar_id})
        if not existente:
            raise HTTPException(status_code=404, detail="Radar não encontrado")
        
        # Preparar dados para atualização
        update_data = {
            k: v for k, v in radar.dict(exclude_unset=True).items() 
            if v is not None
        }
        
        if not update_data:
            raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
        
        update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        await db.radares.update_one(
            {"id": radar_id},
            {"$set": update_data}
        )
        
        # Buscar radar atualizado
        radar_atualizado = await db.radares.find_one(
            {"id": radar_id},
            {"_id": 0}
        )
        
        logger.info(f"✅ Radar atualizado: {radar_id}")
        
        return {
            "message": "Radar atualizado com sucesso",
            "radar": radar_atualizado
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao atualizar radar: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/radares/{radar_id}")
async def deletar_radar(radar_id: str):
    """
    🎯 Deleta um radar (v25.1)
    """
    try:
        result = await db.radares.delete_one({"id": radar_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Radar não encontrado")
        
        logger.info(f"✅ Radar deletado: {radar_id}")
        
        return {
            "message": "Radar deletado com sucesso",
            "radar_id": radar_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao deletar radar: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/radares/{radar_id}/executar")
async def executar_radar(radar_id: str):
    """
    🎯 Executa um radar manualmente (v25.1)
    
    Busca editais que correspondem aos termos e envia por e-mail.
    """
    try:
        radar = await db.radares.find_one({"id": radar_id}, {"_id": 0})
        
        if not radar:
            raise HTTPException(status_code=404, detail="Radar não encontrado")
        
        # Importar serviço de busca
        from services.busca_service_v2 import get_busca_service_v2
        from services.email_service import get_email_service
        
        busca_service = get_busca_service_v2(db)
        email_service = get_email_service()
        
        # Buscar editais
        termos = [t.strip() for t in radar['termos'].split(',') if t.strip()]
        
        resultado = await busca_service.buscar(
            keywords=termos,
            limit=20,
            incluir_ativas=True
        )
        
        editais = resultado.get('resultados', [])
        
        # Filtrar editais já enviados
        enviados_ids = set(radar.get('editais_enviados', []))
        novos_editais = [e for e in editais if e.get('id') not in enviados_ids]
        
        if novos_editais:
            # Enviar e-mail
            await email_service.enviar_alerta_radar(
                email=radar['email'],
                nome_radar=radar['nome'],
                editais=novos_editais[:10]
            )
            
            # Atualizar radar
            novos_ids = [e.get('id') for e in novos_editais if e.get('id')]
            await db.radares.update_one(
                {"id": radar_id},
                {
                    "$set": {
                        "ultimo_envio": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    },
                    "$inc": {"total_enviados": len(novos_editais)},
                    "$addToSet": {"editais_enviados": {"$each": novos_ids}}
                }
            )
        
        return {
            "message": "Radar executado com sucesso",
            "radar_id": radar_id,
            "editais_encontrados": len(editais),
            "novos_editais": len(novos_editais),
            "email_enviado": len(novos_editais) > 0
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao executar radar: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DE PREÇOS (v25.1) ====================

@api_router.get("/precos/search")
async def buscar_precos(
    q: str = Query(..., description="Termo de busca para preços"),
    uf: str = Query(None, description="Filtrar por UF"),
    limite: int = Query(100, ge=1, le=500, description="Limite de resultados"),
    use_cache: bool = Query(True, description="Usar cache (15 dias)"),
    meses: int = Query(24, ge=3, le=24, description="Período em meses (3, 6, 9, 12, 18, 24)")
):
    """
    🎯 Central de Preços v43.0 - Inteligência de Mercado
    
    Consulta preços históricos dos últimos 24 meses:
    - Fonte primária: API PNCP (itens homologados)
    - Fonte secundária: Base local MongoDB
    - Agregações: Mínimo, Máximo, Média, Mediana
    - Cache: 15 dias para economia de processamento
    
    Retorna:
    - Big Numbers (min, max, média)
    - Lista de itens com órgão, quantidade, valor, data
    """
    try:
        from services.precos_service import get_precos_service
        
        precos_service = get_precos_service(db)
        
        logger.info(f"[PRECOS] Buscando: {q} | UF: {uf or 'BR'} | Limite: {limite} | Meses: {meses}")
        
        # Buscar preços com agregações e agrupamento por apresentação
        resumo, apresentacoes = await precos_service.buscar_precos(
            termo=q,
            uf=uf,
            limit=limite,
            use_cache=use_cache,
            meses=meses
        )
        
        # Formatar resposta para o frontend
        return {
            "termo": resumo.termo_pesquisado,
            "total": resumo.total_registros,
            "periodo": f"{meses} meses",
            
            # Big Numbers para o dashboard
            "agregacoes": {
                "minimo": resumo.preco_minimo,
                "maximo": resumo.preco_maximo,
                "medio": resumo.preco_medio,
                "mediana": resumo.preco_mediana,
                "desvio_padrao": resumo.desvio_padrao
            },
            
            # Período dos dados
            "periodo_dados": {
                "data_mais_antiga": resumo.data_mais_antiga,
                "data_mais_recente": resumo.data_mais_recente
            },
            
            # Agrupamento por apresentação (NOVO v44)
            "apresentacoes": [
                {
                    "nome": ap.apresentacao,
                    "total": ap.total_registros,
                    "preco_minimo": ap.preco_minimo,
                    "preco_maximo": ap.preco_maximo,
                    "preco_medio": ap.preco_medio,
                    "preco_mediana": ap.preco_mediana,
                    "itens": [
                        {
                            "orgao": it.orgao,
                            "uf": it.uf,
                            "municipio": it.municipio,
                            "descricao": it.descricao,
                            "quantidade": it.quantidade,
                            "unidade": it.unidade,
                            "valor_unitario": it.valor_unitario,
                            "valor_total": it.valor_total,
                            "data_homologacao": it.data_homologacao,
                            "modalidade": it.modalidade,
                            "numero_processo": it.numero_processo,
                            "fonte": it.fonte
                        }
                        for it in ap.itens
                    ],
                    "tendencia": ap.tendencia
                }
                for ap in apresentacoes
            ],
            
            # Lista flat de itens para compatibilidade
            "resultados": [
                {
                    "orgao": it.orgao,
                    "uf": it.uf,
                    "municipio": it.municipio,
                    "descricao": it.descricao,
                    "quantidade": it.quantidade,
                    "unidade": it.unidade,
                    "valor_unitario": it.valor_unitario,
                    "valor_total": it.valor_total,
                    "data_homologacao": it.data_homologacao,
                    "modalidade": it.modalidade,
                    "numero_processo": it.numero_processo,
                    "fonte": it.fonte
                }
                for it in resumo.itens[:limite]
            ],
            
            "fontes": ["PNCP", "Base Local"],
            "cache_usado": use_cache
        }
        
    except Exception as e:
        logger.error(f"[PRECOS] Erro na busca: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/precos/export-excel")
async def exportar_precos_excel(
    q: str = Query(..., description="Termo de busca"),
    uf: str = Query(None, description="Filtrar por UF"),
    limite: int = Query(200, ge=1, le=500),
    meses: int = Query(24, ge=3, le=24, description="Período em meses"),
):
    """Exporta resultados de preços em formato Excel (.xlsx) com logo GSM"""
    try:
        from services.precos_service import get_precos_service
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XlImage
        
        precos_service = get_precos_service(db)
        resumo, apresentacoes = await precos_service.buscar_precos(
            termo=q, uf=uf, limit=limite, use_cache=False, meses=meses
        )
        
        wb = Workbook()
        
        # Estilos
        verde_escuro = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        verde_claro = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        azul_claro = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        cinza_header = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        font_branca = Font(color="FFFFFF", bold=True, size=11)
        font_titulo = Font(bold=True, size=14, color="1B5E20")
        font_subtitulo = Font(bold=True, size=11, color="546E7A")
        font_valor = Font(bold=True, size=11, color="2E7D32")
        font_normal = Font(size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CFD8DC'),
            right=Side(style='thin', color='CFD8DC'),
            top=Side(style='thin', color='CFD8DC'),
            bottom=Side(style='thin', color='CFD8DC')
        )
        
        # === ABA RESUMO ===
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        ws_resumo.sheet_properties.tabColor = "1B5E20"
        
        # Logo
        logo_path = Path(__file__).parent / "assets" / "logo_gsm.png"
        if logo_path.exists():
            img = XlImage(str(logo_path))
            img.width = 200
            img.height = 80
            ws_resumo.add_image(img, "A1")
        
        # Header
        ws_resumo.merge_cells("A6:G6")
        ws_resumo["A6"] = "RELATÓRIO DE PREÇOS - CENTRAL DE INTELIGÊNCIA GSM"
        ws_resumo["A6"].font = font_titulo
        
        ws_resumo.merge_cells("A7:G7")
        ws_resumo["A7"] = f"Termo pesquisado: {q.upper()}"
        ws_resumo["A7"].font = font_subtitulo
        
        ws_resumo.merge_cells("A8:G8")
        from datetime import datetime
        ws_resumo["A8"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Fonte: PNCP + Base Local | Período: {meses} meses"
        ws_resumo["A8"].font = Font(size=9, color="90A4AE")
        
        # Big Numbers
        row = 10
        labels = ["Menor Preço", "Preço Médio", "Mediana", "Maior Preço", "Total Registros"]
        values = [
            f"R$ {resumo.preco_minimo:,.2f}",
            f"R$ {resumo.preco_medio:,.2f}",
            f"R$ {resumo.preco_mediana:,.2f}",
            f"R$ {resumo.preco_maximo:,.2f}",
            str(resumo.total_registros)
        ]
        for i, (label, val) in enumerate(zip(labels, values)):
            col = 2 + i
            ws_resumo.cell(row=row, column=col, value=label).font = Font(bold=True, size=9, color="78909C")
            ws_resumo.cell(row=row, column=col).alignment = align_center
            ws_resumo.cell(row=row+1, column=col, value=val).font = Font(bold=True, size=13, color="1B5E20")
            ws_resumo.cell(row=row+1, column=col).alignment = align_center
            ws_resumo.cell(row=row, column=col).fill = verde_claro
            ws_resumo.cell(row=row+1, column=col).fill = verde_claro
        
        # Tabela de apresentações
        row = 14
        ws_resumo.cell(row=row, column=1, value="APRESENTAÇÕES POR DOSAGEM").font = Font(bold=True, size=12, color="1B5E20")
        row += 1
        headers_ap = ["Apresentação", "Registros", "Menor Preço", "Preço Médio", "Mediana", "Maior Preço"]
        for i, h in enumerate(headers_ap, 1):
            cell = ws_resumo.cell(row=row, column=i, value=h)
            cell.font = font_branca
            cell.fill = cinza_header
            cell.alignment = align_center
            cell.border = thin_border
        
        for ap in apresentacoes:
            row += 1
            vals = [
                ap.apresentacao, ap.total_registros,
                f"R$ {ap.preco_minimo:,.2f}", f"R$ {ap.preco_medio:,.2f}",
                f"R$ {ap.preco_mediana:,.2f}", f"R$ {ap.preco_maximo:,.2f}"
            ]
            for i, v in enumerate(vals, 1):
                cell = ws_resumo.cell(row=row, column=i, value=v)
                cell.font = font_normal
                cell.border = thin_border
                if i >= 3:
                    cell.alignment = align_right
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        
        # Ajustar larguras
        for col_idx in range(1, 8):
            ws_resumo.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # === ABAS POR APRESENTAÇÃO ===
        for ap in apresentacoes:
            nome_aba = ap.apresentacao[:31].replace("/", "-")  # Excel limita 31 chars
            ws = wb.create_sheet(title=nome_aba)
            
            # Header da aba
            ws.merge_cells("A1:G1")
            ws["A1"] = f"{ap.apresentacao} - {ap.total_registros} registros"
            ws["A1"].font = Font(bold=True, size=12, color="1B5E20")
            
            ws.merge_cells("A2:G2")
            ws["A2"] = f"Min: R$ {ap.preco_minimo:,.2f} | Médio: R$ {ap.preco_medio:,.2f} | Mediana: R$ {ap.preco_mediana:,.2f} | Max: R$ {ap.preco_maximo:,.2f}"
            ws["A2"].font = Font(size=10, color="546E7A")
            
            # Headers da tabela
            headers = ["Órgão Comprador", "UF", "Descrição", "Quantidade", "Valor Unitário", "Data", "Fonte"]
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=i, value=h)
                cell.font = font_branca
                cell.fill = cinza_header
                cell.alignment = align_center
                cell.border = thin_border
            
            # Dados
            for idx, item in enumerate(ap.itens):
                r = 5 + idx
                desc_limpa = (item.descricao or '').replace('<br>', ' ').replace('&nbsp;', ' ')
                import re as re_mod
                desc_limpa = re_mod.sub(r'<[^>]+>', '', desc_limpa)
                
                row_data = [
                    item.orgao,
                    item.uf or '-',
                    desc_limpa[:100],
                    f"{item.quantidade} {item.unidade}",
                    item.valor_unitario,
                    (item.data_homologacao or '-')[:10],
                    item.fonte
                ]
                for i, v in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=i, value=v)
                    cell.font = font_normal
                    cell.border = thin_border
                    if r % 2 == 0:
                        cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
                
                # Formatar valor como moeda
                ws.cell(row=r, column=5).number_format = 'R$ #,##0.00'
                ws.cell(row=r, column=5).font = font_valor
                ws.cell(row=r, column=5).alignment = align_right
            
            # Ajustar larguras
            widths = [35, 6, 45, 15, 18, 12, 8]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        
        # Gerar arquivo em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"precos_{q.replace(' ', '_').lower()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        logger.error(f"[PRECOS] Erro ao exportar Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== SINCRONIZAÇÃO PNCP v52.1 ====================
from services.pncp_sync_service import get_pncp_sync_service

@api_router.post("/sync/pncp")
async def sincronizar_pncp(
    termo: str = Query(..., description="Termo de busca (ex: canabidiol, insulina)"),
    dias: int = Query(default=30, description="Últimos N dias"),
    limite: int = Query(default=100, description="Máximo de registros")
):
    """
    🔄 v52.1: Sincronização DIRETA com API PNCP
    
    Busca dados em: https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao
    Salva no MongoDB local na collection: editais_gsm
    
    100% INDEPENDENTE - Sem intermediários de terceiros.
    """
    try:
        sync_service = get_pncp_sync_service(db)
        
        logger.info(f"🔄 [SYNC-PNCP] Iniciando: termo='{termo}', dias={dias}, limite={limite}")
        
        resultado = await sync_service.sincronizar_termo(
            termo=termo,
            dias=dias,
            limite=limite
        )
        
        return {
            "status": "sucesso",
            "message": f"Sincronização concluída: {resultado['novos_inseridos']} novos, {resultado['atualizados']} atualizados",
            "detalhes": resultado
        }
        
    except Exception as e:
        logger.error(f"❌ [SYNC-PNCP] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/sync/stats")
async def stats_sincronizacao():
    """
    📊 v52.1: Estatísticas do banco local GSM
    
    Mostra quantos editais estão clonados localmente.
    """
    try:
        sync_service = get_pncp_sync_service(db)
        stats = await sync_service.get_stats()
        
        return {
            "status": "sucesso",
            "banco_local": stats,
            "independencia": "100% - Dados próprios no MongoDB"
        }
        
    except Exception as e:
        logger.error(f"❌ [SYNC-STATS] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/popular-base")
async def popular_base_local(
    termos: str = Query(default="canabidiol,insulina,prolia,denosumabe,medicamento", description="Termos separados por vírgula"),
    dias: int = Query(default=60, description="Últimos N dias")
):
    """
    🚀 v52.1: Popular banco local com múltiplos termos
    
    Sincroniza dados do PNCP para vários termos de uma vez.
    Ideal para setup inicial do sistema independente.
    """
    try:
        sync_service = get_pncp_sync_service(db)
        
        lista_termos = [t.strip() for t in termos.split(",") if t.strip()]
        
        resultados = []
        total_novos = 0
        total_atualizados = 0
        
        for termo in lista_termos:
            logger.info(f"🔄 [POPULAR] Sincronizando: {termo}")
            resultado = await sync_service.sincronizar_termo(
                termo=termo,
                dias=dias,
                limite=100
            )
            resultados.append({
                "termo": termo,
                "novos": resultado["novos_inseridos"],
                "atualizados": resultado["atualizados"]
            })
            total_novos += resultado["novos_inseridos"]
            total_atualizados += resultado["atualizados"]
        
        return {
            "status": "sucesso",
            "message": f"Base populada: {total_novos} novos editais, {total_atualizados} atualizados",
            "por_termo": resultados,
            "totais": {
                "novos": total_novos,
                "atualizados": total_atualizados
            }
        }
        
    except Exception as e:
        logger.error(f"❌ [POPULAR] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== INGESTÃO MASSIVA v57.0 ====================

from services.clone_massivo_service import get_clone_massivo_service
from services.ingestao_massiva_service import get_ingestao_massiva_service
from services.clone_agregador_service import get_clone_agregador_service


@api_router.post("/clone/agregador")
async def clonar_agregador(
    max_paginas: int = Query(default=100, description="Máximo de páginas a processar")
):
    """
    🚀 v59.0: CLONE AGREGADOR - Espelhamento Total
    
    Clona TODOS os dados da API Agregador para o MongoDB local.
    OBJETIVO: Igualar o volume de 950+ resultados do parceiro.
    """
    try:
        clone_service = get_clone_agregador_service(db)
        
        logger.info(f"🚀 [CLONE-AGREGADOR] Iniciando clonagem massiva...")
        
        stats = await clone_service.clonar_tudo(max_paginas=max_paginas)
        
        return {
            "status": "sucesso",
            "message": f"Clonagem concluída: {stats.get('total_clonados', 0)} novos, {stats.get('total_editais_gsm', 0)} total no banco",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [CLONE-AGREGADOR] Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/motor/sincronizar")
async def forcar_sincronizacao():
    """
    🔄 v59.0: Força sincronização imediata com Agregador
    
    Executa o motor de sincronização manualmente (normalmente roda a cada 15 min).
    """
    try:
        from services.motor_sincronizacao_gsm import get_motor_sincronizacao
        
        motor = get_motor_sincronizacao(db)
        stats = await motor.sincronizar()
        
        return {
            "status": "sucesso",
            "message": f"Sincronização concluída: {stats.get('novos', 0)} novos editais",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [MOTOR] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/motor/status")
async def status_motor():
    """
    📊 v59.0: Status do motor de sincronização
    """
    try:
        from services.motor_sincronizacao_gsm import get_motor_sincronizacao
        
        motor = get_motor_sincronizacao(db)
        status = await motor.get_status()
        
        return {
            "status": "ok",
            "motor": status,
            "mensagem": f"Motor GSM ativo. {status.get('total_editais', 0)} editais no banco."
        }
        
    except Exception as e:
        logger.error(f"❌ [MOTOR] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/fontes-diretas/sincronizar")
async def sincronizar_fontes_diretas():
    """
    🌐 v59.0: Sincronização direta com portais (sem Agregador)
    
    Acessa diretamente:
    - PNCP (pncp.gov.br)
    - ComprasNet (compras.dados.gov.br)
    - BNC, BLL, Licitar Digital
    """
    try:
        from services.fontes_diretas_gsm import get_fontes_diretas
        
        fontes = get_fontes_diretas(db)
        stats = await fontes.sincronizar_todas_fontes()
        
        return {
            "status": "sucesso",
            "message": f"Sincronização direta: {stats.get('total_novos', 0)} novos editais",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [FONTES-DIRETAS] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/ingestao-massiva")
async def executar_ingestao_massiva(
    termos: str = Query(default=None, description="Termos específicos (opcional). Se vazio, usa termos padrão de saúde."),
    dias: int = Query(default=365, description="Período em dias para buscar"),
    limite: int = Query(default=500, description="Limite de registros por fonte")
):
    """
    🚀 v57.0: INGESTÃO MASSIVA - Data Mirroring Multi-Fonte
    
    Ingere dados de múltiplas fontes:
    - PNCP API Oficial
    - ComprasNet API
    - Dados de demonstração (para teste de volume)
    
    OBJETIVO: Popular editais_gsm com 900+ documentos para termos como "canabidiol".
    
    ⚠️ NOTA: Execute este endpoint para popular o banco antes de usar a busca.
    """
    try:
        ingestao_service = get_ingestao_massiva_service(db)
        
        lista_termos = None
        if termos:
            lista_termos = [t.strip() for t in termos.split(",") if t.strip()]
        
        logger.info(f"🚀 [INGESTAO-MASSIVA] Iniciando ingestão: termos={lista_termos}")
        
        stats = await ingestao_service.executar_ingestao_completa(
            termos=lista_termos,
            dias=dias,
            limite_por_fonte=limite
        )
        
        return {
            "status": "sucesso",
            "message": f"Ingestão concluída: {stats['total_inseridos']} inseridos",
            "total_editais_gsm": stats.get("total_editais_gsm", 0),
            "detalhes": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [INGESTAO-MASSIVA] Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/sync/stats-gsm")
async def stats_gsm():
    """
    📊 v57.0: Estatísticas da collection editais_gsm
    """
    try:
        ingestao_service = get_ingestao_massiva_service(db)
        stats = await ingestao_service.get_stats()
        
        return {
            "status": "sucesso",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [STATS-GSM] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/popular-termo")
async def popular_termo_especifico(
    termo: str = Query(..., description="Termo a ser populado"),
    quantidade: int = Query(default=500, description="Quantidade de editais a gerar")
):
    """
    🚀 v57.0: Popular banco com editais específicos para um termo.
    
    Gera dados de demonstração para um termo específico, 
    permitindo atingir o volume de 900+ resultados.
    """
    try:
        ingestao_service = get_ingestao_massiva_service(db)
        
        logger.info(f"🚀 [POPULAR-TERMO] Gerando {quantidade} editais para '{termo}'")
        
        stats = await ingestao_service._gerar_dados_demonstracao(
            termos=[termo],
            limite=quantidade
        )
        
        return {
            "status": "sucesso",
            "message": f"Populado: {stats['inseridos']} editais para '{termo}'",
            "total_editais_gsm": await db.editais_gsm.count_documents({}),
            "detalhes": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [POPULAR-TERMO] Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sync/clone-massivo")
async def executar_clone_massivo(
    termos: str = Query(default=None, description="Termos para filtrar (opcional). Se vazio, clona TUDO.")
):
    """
    🔄 v55.0: CLONE MASSIVO - Espelhamento de Dados
    
    Clona TODOS os dados de:
    - editais_normalizados (PNCP)
    - licitacoes (histórico)
    
    Para a collection própria editais_gsm.
    
    OBJETIVO: Igualar volume do parceiro (centenas de resultados).
    """
    try:
        clone_service = get_clone_massivo_service(db)
        
        lista_termos = None
        if termos:
            lista_termos = [t.strip() for t in termos.split(",") if t.strip()]
        
        logger.info(f"🔄 [CLONE-MASSIVO] Iniciando clone: termos={lista_termos}")
        
        stats = await clone_service.executar_clone_massivo(termos=lista_termos)
        
        return {
            "status": "sucesso",
            "message": f"Clone concluído: {stats['novos_inseridos']} novos, {stats['atualizados']} atualizados",
            "total_editais_gsm": stats.get("total_editais_gsm", 0),
            "detalhes": stats
        }
        
    except Exception as e:
        logger.error(f"❌ [CLONE-MASSIVO] Erro: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/sync/volume")
async def verificar_volume():
    """
    Verificar volume de dados no sistema
    """
    try:
        clone_service = get_clone_massivo_service(db)
        stats = await clone_service.get_stats()
        
        return {
            "status": "sucesso",
            "volume": stats,
            "recomendacao": "Execute /api/sync/clone-massivo para igualar volumes" if stats["editais_gsm"] < stats["editais_normalizados"] else "Volume OK"
        }
        
    except Exception as e:
        logger.error(f"Erro volume: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync/batch-import")
async def batch_import_avisos(data: dict):
    """
    v74.0: Import direto de dados da API do Agregador.
    Aceita array de avisos no formato original da API.
    """
    try:
        avisos = data.get("avisos", [])
        if not avisos:
            raise HTTPException(status_code=400, detail="Campo 'avisos' vazio")
        
        collection = db.editais_gsm
        imported = 0
        updated = 0
        errors = 0
        
        for aviso in avisos:
            try:
                aviso_id = str(aviso.get('id', ''))
                if not aviso_id:
                    errors += 1
                    continue
                
                import hashlib
                id_gsm = hashlib.md5(f"AGREGADOR-{aviso_id}".encode()).hexdigest()
                
                # Extrair dados
                portal_nome = aviso.get('portalNome', '') or ''
                uasg_nome = aviso.get('uasgNome', '') or ''
                uasg = str(aviso.get('uasg', ''))
                uf_completo = aviso.get('uf', '') or ''
                municipio = aviso.get('municipio') or aviso.get('orgaoDetalhe', {}).get('municipio', '') if isinstance(aviso.get('orgaoDetalhe'), dict) else ''
                objeto = (aviso.get('objeto', '') or '').upper()
                pregao = aviso.get('pregao', '') or ''
                url = aviso.get('url', '') or ''
                
                # Extrair UF sigla
                UF_MAP = {
                    "Acre": "AC", "Alagoas": "AL", "Amapá": "AP", "Amazonas": "AM",
                    "Bahia": "BA", "Ceará": "CE", "Distrito Federal": "DF",
                    "Espírito Santo": "ES", "Goiás": "GO", "Maranhão": "MA",
                    "Mato Grosso": "MT", "Mato Grosso do Sul": "MS", "Minas Gerais": "MG",
                    "Pará": "PA", "Paraíba": "PB", "Paraná": "PR", "Pernambuco": "PE",
                    "Piauí": "PI", "Rio de Janeiro": "RJ", "Rio Grande do Norte": "RN",
                    "Rio Grande do Sul": "RS", "Rondônia": "RO", "Roraima": "RR",
                    "Santa Catarina": "SC", "São Paulo": "SP", "Sergipe": "SE", "Tocantins": "TO"
                }
                uf_sigla = UF_MAP.get(uf_completo, uf_completo[:2] if uf_completo else '')
                
                # Extrair link PDF dos anexos
                link_pdf = None
                anexos_raw = aviso.get('anexo', []) or []
                for a in anexos_raw:
                    nome = (a.get('nome', '') or '').lower()
                    a_url = a.get('url', '')
                    if a_url and 'edital' in nome:
                        link_pdf = a_url
                        break
                if not link_pdf:
                    for a in anexos_raw:
                        nome = (a.get('nome', '') or '').lower()
                        a_url = a.get('url', '')
                        if a_url and ('.pdf' in nome or '.zip' in nome):
                            link_pdf = a_url
                            break
                if not link_pdf:
                    for a in anexos_raw:
                        a_url = a.get('url', '')
                        if a_url:
                            link_pdf = a_url
                            break
                
                # Formatar itens com todos os campos
                itens_formatados = []
                for item in (aviso.get('item', []) or []):
                    itens_formatados.append({
                        "grupo": str(item.get('grupo', '')),
                        "numero": str(item.get('numero', '')) if item.get('numero') else '',
                        "descricao": item.get('descricao', ''),
                        "exclusivo_me_epp": item.get('exclusivoMeEpp', -1),
                        "quantidade": str(item.get('quantidade', '')),
                        "unidade": item.get('unidade', ''),
                        "decreto_7174": item.get('decreto7174', 0),
                        "valor_total": item.get('valorTotal'),
                        "valor_unitario": item.get('valorUnitario')
                    })
                
                # Formatar anexos
                anexos_formatados = [
                    {"nome": a.get('nome', ''), "url": a.get('url', ''), "codigo": a.get('codigo')}
                    for a in anexos_raw if a.get('url')
                ]
                
                # Detectar saúde
                termos_saude = ['medicamento', 'farmac', 'hospital', 'saude', 'saúde', 'insulina', 'canabidiol', 'seringas', 'luvas', 'medicamentos', 'injetável', 'comprimido', 'droga', 'remédio', 'cirurg', 'ortopéd', 'prótese', 'laborat', 'diagnóst', 'vacina']
                is_saude = any(t in objeto.lower() for t in termos_saude)
                
                doc = {
                    "id_gsm": id_gsm,
                    "id_externo": aviso_id,
                    "numero_controle_pncp": "",
                    "fonte_origem": "AGREGADOR",
                    "fonte": portal_nome,
                    "portal_captura": portal_nome,
                    "dados_orgao": {
                        "uasg": uasg,
                        "cnpj": "",
                        "nome": uasg_nome,
                        "uf": uf_sigla,
                        "municipio": municipio or ''
                    },
                    "objeto": objeto,
                    "orgao": uasg_nome,
                    "estado": uf_sigla,
                    "uf": uf_sigla,
                    "municipio": municipio or '',
                    "uasg": uasg,
                    "modalidade": aviso.get('tipo', 'Pregão Eletrônico'),
                    "status": "ABERTA",
                    "data_publicacao": aviso.get('dataPublicacao'),
                    "data_abertura": aviso.get('dataFinal'),
                    "data_inicial": aviso.get('dataInicial'),
                    "data_final": aviso.get('dataFinal'),
                    "link_documento": link_pdf or url,
                    "link_pdf": link_pdf,
                    "link_origem": url,
                    "link_portal": url,
                    "anexos": anexos_formatados,
                    "numero_processo": pregao,
                    "numero_licitacao": pregao,
                    "itens_clonados": itens_formatados,
                    "valor_total_estimado": aviso.get('valorTotalEstimado'),
                    "is_srp": aviso.get('isSrp', False),
                    "exclusivo_me_epp": aviso.get('exclusivoMeEpp', -1),
                    "palavras_chave": aviso.get('palavrasChave', []),
                    "sincronizado_em": datetime.now(timezone.utc),
                    "atualizado_em": datetime.now(timezone.utc),
                    "is_saude": is_saude,
                    "is_clone_agregador": True
                }
                
                result = await collection.update_one(
                    {"id_gsm": id_gsm},
                    {"$set": doc},
                    upsert=True
                )
                
                if result.upserted_id:
                    imported += 1
                elif result.modified_count > 0:
                    updated += 1
                    
            except Exception as item_err:
                errors += 1
                logger.error(f"Erro importando aviso {aviso.get('id', '?')}: {item_err}")
        
        total = await collection.count_documents({})
        
        return {
            "status": "sucesso",
            "importados": imported,
            "atualizados": updated,
            "erros": errors,
            "total_avisos_recebidos": len(avisos),
            "total_editais_gsm": total
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro batch-import: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROTAS DAMA VIGÊNCIA (P0) ====================

@api_router.post("/dama/vigencia/sync")
async def sync_vigencia():
    """Força re-scraping das resoluções CMED para atualizar status de vigência."""
    from services.vigencia_service import get_vigencia_service
    svc = get_vigencia_service(db)
    resolucoes = await svc.scrape_resolucoes()
    stats = await svc.get_stats()
    return {"total": len(resolucoes), "stats": stats}


@api_router.get("/dama/vigencia/check")
async def check_vigencia(referencia: str = Query(..., description="Ex: Resolução 07/2022")):
    """Verifica vigência de uma resolução CMED específica."""
    from services.vigencia_service import get_vigencia_service
    svc = get_vigencia_service(db)
    return await svc.verificar_vigencia(referencia)


@api_router.get("/dama/vigencia/resolucoes")
async def listar_resolucoes_vigencia():
    """Lista todas as resoluções CMED com status de vigência."""
    from services.vigencia_service import get_vigencia_service
    svc = get_vigencia_service(db)
    stats = await svc.get_stats()
    if stats['total'] == 0:
        await svc.scrape_resolucoes()
    cursor = svc.collection.find({}, {'_id': 0}).sort([('ano', -1), ('numero', -1)])
    resolucoes = await cursor.to_list(length=200)
    return {"resolucoes": resolucoes, "total": len(resolucoes), "stats": stats}


@api_router.get("/dama/vigencia/stats")
async def stats_vigencia():
    """Estatísticas de vigência das resoluções CMED."""
    from services.vigencia_service import get_vigencia_service
    svc = get_vigencia_service(db)
    return await svc.get_stats()


@api_router.post("/dama/vigencia/validar-esclarecimento")
async def validar_vigencia_para_esclarecimento():
    """
    Valida vigência das normas-chave usadas no Esclarecimento Técnico.
    Retorna status de cada norma e se a geração é segura.
    """
    from services.vigencia_service import get_vigencia_service
    svc = get_vigencia_service(db)
    
    stats = await svc.get_stats()
    if stats['total'] == 0:
        await svc.scrape_resolucoes()
    
    normas_chave = [
        "Resolução 07/2022",
        "Resolução 13/2022",
        "Resolução 02/2004",
        "Resolução 01/2003",
    ]
    
    resultados = []
    bloqueios = []
    for norma in normas_chave:
        resultado = await svc.verificar_vigencia(norma)
        resultados.append(resultado)
        if resultado.get('encontrada') and not resultado.get('pode_usar'):
            bloqueios.append(resultado)
    
    vigentes = await svc.get_resolucoes_vigentes()
    
    return {
        "validacao": resultados,
        "bloqueios": bloqueios,
        "tem_bloqueio": len(bloqueios) > 0,
        "normas_vigentes": vigentes[:10],
        "stats": stats,
    }


# ==================== DAMA CHECKLIST AUTOMATIZADO (P2) ====================

@api_router.post("/dama/checklist")
async def executar_dama_checklist(data: dict):
    """
    Executa checklist DAMA completo para um medicamento.
    Verifica: vigencia normativa, janela aberta, publicacao oficial.
    """
    from services.dama_checklist_service import get_checklist_service

    medicamento = data.get('medicamento', '').strip()
    if not medicamento:
        raise HTTPException(status_code=400, detail="Campo 'medicamento' obrigatorio")

    normas = data.get('normas', [])
    svc = get_checklist_service(db)
    resultado = await svc.executar_checklist(medicamento, normas)
    return resultado




# ==================== RADAR FARMACEUTICO - INTELIGENCIA DESABASTECIMENTO ====================

@api_router.get("/radar-farmaceutico/lista-interesse")
async def listar_interesse_estrategico():
    """Lista medicamentos da Lista de Interesse Estrategica."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    items = await svc.listar_interesse()
    return {"items": items, "total": len(items)}


@api_router.post("/radar-farmaceutico/lista-interesse")
async def adicionar_interesse(data: dict):
    """Adiciona medicamento a Lista de Interesse."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    if not data.get('medicamento'):
        raise HTTPException(status_code=400, detail="Campo 'medicamento' obrigatorio")
    svc = get_radar_farmaceutico_service(db)
    item = await svc.adicionar_interesse(data)
    return item


@api_router.put("/radar-farmaceutico/lista-interesse/{item_id}")
async def atualizar_interesse(item_id: str, data: dict):
    """Atualiza medicamento na Lista de Interesse."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    ok = await svc.atualizar_interesse(item_id, data)
    if not ok:
        raise HTTPException(status_code=404, detail="Item nao encontrado ou sem alteracoes")
    return {"message": "Atualizado", "id": item_id}


@api_router.delete("/radar-farmaceutico/lista-interesse/{item_id}")
async def remover_interesse(item_id: str):
    """Remove medicamento da Lista de Interesse."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    ok = await svc.remover_interesse(item_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Item nao encontrado")
    return {"message": "Removido", "id": item_id}


@api_router.post("/radar-farmaceutico/seed")
async def seed_lista_interesse():
    """Insere seeds iniciais na Lista de Interesse (se vazia)."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    result = await svc.seed_lista_interesse()
    return result


@api_router.post("/radar-farmaceutico/scan")
async def executar_scan_desabastecimento(background_tasks: BackgroundTasks):
    """
    Executa scan de desabastecimento em background.
    Cruza DOU/ANVISA com lista de interesse e detecta oportunidades criticas.
    """
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service

    async def _run_scan():
        try:
            svc = get_radar_farmaceutico_service(db)
            resultado = await svc.executar_scan()
            logger.info(f"Radar Farmaceutico scan concluido: {resultado}")
        except Exception as e:
            logger.error(f"Radar Farmaceutico scan erro: {e}")

    import asyncio
    asyncio.create_task(_run_scan())
    return {"status": "processando", "mensagem": "Scan de desabastecimento iniciado em background. Atualize em ~30s."}


@api_router.get("/radar-farmaceutico/desabastecimento")
async def listar_desabastecimento(limite: int = Query(50, ge=1, le=200)):
    """Lista registros de desabastecimento detectados (cruzados com lista de interesse)."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    items = await svc.listar_desabastecimento(limite=limite)
    stats = await svc.estatisticas()
    return {"items": items, "total": len(items), "estatisticas": stats}


@api_router.get("/radar-farmaceutico/stats")
async def stats_radar_farmaceutico():
    """Estatisticas do Radar Farmaceutico."""
    from services.radar_farmaceutico_service import get_radar_farmaceutico_service
    svc = get_radar_farmaceutico_service(db)
    return await svc.estatisticas()


# ==================== LMR - RADAR DE IMPORTACAO (IN 428/2026) ====================

@api_router.get("/dama/lmr-analysis")
async def listar_oportunidades_lmr(limite: int = Query(20, ge=1, le=100)):
    """
    Lista oportunidades de importacao rankeadas por score.
    Baseado na IN 428/2026 (Lista de Medicamentos de Referencia).
    Smart Cache: TTL 24h.
    """
    cached = smart_cache.get('lmr_analysis', limite=limite)
    if cached is not None:
        logger.info("SmartCache HIT: lmr_analysis")
        return cached

    from services.lmr_service import get_lmr_service
    svc = get_lmr_service(db)
    resultado = await svc.listar_oportunidades(limite=limite)

    smart_cache.set(resultado, namespace='lmr_analysis', limite=limite)
    return resultado


@api_router.post("/dama/lmr-analise-medicamento")
async def analisar_medicamento_lmr(data: dict):
    """
    Analisa um medicamento especifico sob as regras da IN 428/2026.
    Retorna estrategia tributaria, margens e recomendacao.
    Smart Cache: TTL 24h por medicamento+tipo+preco.
    """
    from services.lmr_service import get_lmr_service

    medicamento = data.get('medicamento', '').strip()
    if not medicamento:
        raise HTTPException(status_code=400, detail="Campo 'medicamento' obrigatorio")

    preco_referencia = data.get('preco_referencia', 0)
    tipo_produto = data.get('tipo_produto', 'sintetico')

    cache_key_params = dict(medicamento=medicamento.lower(), preco=preco_referencia, tipo=tipo_produto)
    cached = smart_cache.get('lmr_medicamento', **cache_key_params)
    if cached is not None:
        logger.info(f"SmartCache HIT: lmr_medicamento '{medicamento}'")
        return cached

    svc = get_lmr_service(db)
    resultado = await svc.analisar_medicamento(medicamento, preco_referencia, tipo_produto)

    smart_cache.set(resultado, namespace='lmr_medicamento', **cache_key_params)
    return resultado



# ==================== PROVA DOCUMENTAL PDF (P2) ====================

@api_router.post("/dama/prova-documental")
async def gerar_prova_documental(data: dict):
    """
    Gera PDF de Prova Documental formal para processos de licitacao.
    Inclui analise tributaria LMR se disponivel.
    """
    from services.prova_documental_service import gerar_prova_documental_pdf

    medicamento = data.get('medicamento', '')
    fonte = data.get('fonte', 'ANVISA/DOU')
    titulo = data.get('titulo', '')
    descricao = data.get('descricao', '')
    data_pub = data.get('data_publicacao', '')
    link = data.get('link', '')
    tipo_alerta = data.get('tipo_alerta', '')
    risco = data.get('risco', '')
    classificacao_dama = data.get('classificacao_dama', '')
    empresa_id = data.get('empresa_id', '')

    empresa_data = None
    if empresa_id:
        emp = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        if emp:
            empresa_data = emp

    # Buscar analise LMR para o medicamento (automatico)
    analise_lmr = None
    if medicamento:
        try:
            from services.lmr_service import get_lmr_service
            lmr_svc = get_lmr_service(db)
            analise_lmr = await lmr_svc.analisar_medicamento(medicamento)
        except Exception as e:
            logger.warning(f"Analise LMR para PDF falhou: {e}")

    pdf_bytes = gerar_prova_documental_pdf(
        medicamento=medicamento,
        fonte=fonte,
        titulo_documento=titulo,
        descricao=descricao,
        data_publicacao=data_pub,
        link=link,
        tipo_alerta=tipo_alerta,
        risco=risco,
        empresa=empresa_data,
        classificacao_dama=classificacao_dama,
        analise_lmr=analise_lmr,
    )

    nome_arquivo = f"prova_documental_{medicamento.replace(' ', '_')[:30]}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
        }
    )



@api_router.get("/dama/prova-documental-lmr/{alerta_id}")
async def gerar_prova_documental_lmr(alerta_id: str):
    """
    Gera PDF de Prova Documental LMR a partir de um alerta de oportunidade.
    Chamado diretamente pelo link do e-mail de alerta.
    """
    from services.prova_documental_service import gerar_prova_documental_pdf
    from services.lmr_service import get_lmr_service

    # Buscar alerta no MongoDB
    alerta = await db.oportunidades_alertas.find_one({"id": alerta_id}, {"_id": 0})
    if not alerta:
        raise HTTPException(status_code=404, detail="Alerta nao encontrado")

    medicamento = alerta.get('medicamento', '')

    # Gerar analise LMR completa
    lmr_svc = get_lmr_service(db)
    analise_lmr = await lmr_svc.analisar_medicamento(medicamento)

    pdf_bytes = gerar_prova_documental_pdf(
        medicamento=medicamento,
        fonte='DAMA Intelligence - Radar LMR (IN 428/2026)',
        titulo_documento=f'Analise de Oportunidade LMR - Score {alerta.get("oportunidade_score", 0)}%',
        descricao=alerta.get('recomendacao', ''),
        data_publicacao=alerta.get('criado_em', ''),
        link='',
        tipo_alerta=f'LMR {alerta.get("categoria_lmr", "N/A").upper()}',
        risco='ALTO' if alerta.get('oportunidade_score', 0) >= 80 else 'MEDIO',
        classificacao_dama='oportunidade_importacao',
        analise_lmr=analise_lmr,
    )

    nome_arquivo = f"prova_documental_LMR_{medicamento.replace(' ', '_')[:25]}_{datetime.now().strftime('%Y%m%d')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{nome_arquivo}"',
        }
    )



# ==================== ROTAS DAMA IA (v63.0) ====================

from fastapi import File, UploadFile, Form
from fastapi.responses import Response
import tempfile
import zipfile
@api_router.post("/dama/process")
async def processar_dama(
    edital: List[UploadFile] = File(..., description="PDF(s) do edital (máx 15MB cada)"),
    timbrado: Optional[UploadFile] = File(None, description="DOCX do timbrado (opcional se empresa já tem cadastrado)"),
    empresa_id: str = Form(default="c1", description="ID da empresa GSM"),
    custo_unitario: float = Form(default=0.0, description="Custo unitário para cálculo de margem"),
    moeda: str = Form(default="BRL", description="Moeda: BRL ou USD"),
    itens_config: Optional[str] = Form(None, description="JSON com itens configurados")
):
    """
    🎯 DAMA IA v73.1 - Motor de Inteligência para Licitações
    
    NOVA LÓGICA v73.1:
    - Suporta múltiplos arquivos PDF
    - Se empresa tem timbrado cadastrado, usa ele automaticamente
    - Se não tem, usa o arquivo enviado no request
    - Suporta configuração de itens (participação e preços)
    - SEM cálculo tributário (removido para performance)
    
    Pipeline:
    1. OCR do(s) PDF(s) usando Gemini 2.5 Flash
    2. Extração de dados estruturados (órgão, itens, valores)
    3. Geração de proposta via IA
    4. Injeção no template Word (tag {{TEXTO_DAMA}})
    5. Geração de declarações obrigatórias
    6. Retorno do ZIP com kit completo
    
    Limite de arquivo: 15MB por PDF
    """
    try:
        # v66.0: Processar lista de arquivos PDF
        all_edital_content = b''
        filenames = []
        
        for ed in edital:
            if not ed.filename.lower().endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"O arquivo {ed.filename} deve ser PDF")
            
            content = await ed.read()
            
            if len(content) > MAX_UPLOAD_SIZE:
                raise HTTPException(
                    status_code=413, 
                    detail=f"Arquivo {ed.filename} muito grande. Máximo: 15MB. Atual: {len(content) / (1024*1024):.1f}MB"
                )
            
            all_edital_content += content
            filenames.append(ed.filename)
        
        logger.info(f"🚀 [DAMA v66.0] Processando {len(filenames)} PDF(s): {', '.join(filenames)} ({len(all_edital_content)/1024:.1f}KB total) | Empresa: {empresa_id}")
        
        # v63.0: Buscar empresa e verificar timbrado cadastrado
        empresa_doc = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        empresa_data = None
        timbrado_content = None
        
        if empresa_doc:
            empresa_data = {
                "nome": empresa_doc.get("name", ""),
                "cnpj": empresa_doc.get("cnpj", ""),
                "ie": empresa_doc.get("ie", ""),
                "endereco": empresa_doc.get("address", ""),
                "telefone": empresa_doc.get("phone", ""),
                "whatsapp": empresa_doc.get("whatsapp", ""),
                "email": empresa_doc.get("email", "")
            }
            logger.info(f"🏢 [DAMA] Usando empresa do banco: {empresa_data['nome']}")
            
            # v63.0: Tentar usar timbrado cadastrado
            timbrado_path = empresa_doc.get("timbrado_path")
            if timbrado_path and Path(timbrado_path).exists():
                with open(timbrado_path, "rb") as f:
                    timbrado_content = f.read()
                logger.info(f"📄 [DAMA v63.0] Usando timbrado cadastrado: {timbrado_path}")
        
        # v63.0: Se não tem timbrado cadastrado, usar o enviado no request
        if timbrado_content is None:
            if timbrado is None:
                raise HTTPException(
                    status_code=400, 
                    detail="Empresa não tem timbrado cadastrado. Envie o arquivo .docx ou cadastre na aba Configurações."
                )
            
            if not timbrado.filename.lower().endswith('.docx'):
                raise HTTPException(status_code=400, detail="O arquivo do timbrado deve ser DOCX")
            
            timbrado_content = await timbrado.read()
            logger.info(f"📄 [DAMA v63.0] Usando timbrado do request: {timbrado.filename}")
        
        # v63.0: Processar itens configurados
        itens_configurados = None
        if itens_config:
            try:
                itens_configurados = json.loads(itens_config)
                logger.info(f"📋 [DAMA v63.0] Itens configurados: {len(itens_configurados)}")
            except json.JSONDecodeError:
                logger.warning("⚠️ [DAMA] Erro ao parsear itens_config JSON")
        
        # Importar serviço DAMA
        from services.dama_service import get_dama_service
        
        dama_service = get_dama_service()
        
        # Processar documentos com novo serviço
        zip_content, stats = await dama_service.processar_edital(
            pdf_content=all_edital_content,
            docx_template_content=timbrado_content,
            empresa_id=empresa_id,
            custo_unitario=custo_unitario,
            empresa_data=empresa_data
        )
        
        logger.info(f"✅ [DAMA v66.0] Kit gerado: {stats}")
        
        # v65.0: Salvar ZIP no disco para download confiável
        import uuid
        zip_id = str(uuid.uuid4())[:8]
        zip_filename = f"kit_licitacao_{zip_id}.zip"
        kits_dir = Path(__file__).parent / "uploads" / "kits"
        kits_dir.mkdir(parents=True, exist_ok=True)
        zip_path = kits_dir / zip_filename
        
        with open(zip_path, "wb") as f:
            f.write(zip_content)
        
        logger.info(f"📦 [DAMA v65.0] ZIP salvo: {zip_path} ({len(zip_content)/1024:.1f}KB)")
        
        # Retornar ZIP diretamente com headers corretos
        return Response(
            content=zip_content,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={zip_filename}",
                "Content-Length": str(len(zip_content)),
                "X-Zip-Id": zip_id,
                "X-Zip-Filename": zip_filename,
                "Access-Control-Expose-Headers": "Content-Disposition, X-Zip-Id, X-Zip-Filename"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro no DAMA: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar documentos: {str(e)}")


# v65.0: Endpoint para download de kit salvo
@api_router.get("/dama/download/{zip_id}")
async def download_kit_dama(zip_id: str):
    """
    📥 v65.0 - Download de kit DAMA salvo
    """
    try:
        kits_dir = Path(__file__).parent / "uploads" / "kits"
        
        # Procurar arquivo pelo ID
        for f in kits_dir.glob(f"kit_licitacao_{zip_id}*.zip"):
            with open(f, "rb") as file:
                content = file.read()
            
            return Response(
                content=content,
                media_type="application/zip",
                headers={
                    "Content-Disposition": f"attachment; filename={f.name}",
                    "Content-Length": str(len(content))
                }
            )
        
        raise HTTPException(status_code=404, detail="Kit não encontrado")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro no download: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dama/extrair-itens")
async def extrair_itens_edital(
    edital: UploadFile = File(..., description="PDF do edital")
):
    """
    🔍 DAMA v61.0 - Extrair itens do edital para configuração
    
    Usa Gemini para:
    1. OCR do PDF
    2. Identificar todos os itens licitados
    3. Retornar lista para o usuário configurar (participar ou não, preço)
    """
    try:
        if not edital.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail="Arquivo deve ser PDF")
        
        edital_content = await edital.read()
        
        if len(edital_content) > MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=413, detail="Arquivo muito grande (máx 15MB)")
        
        logger.info(f"🔍 [DAMA] Extraindo itens de: {edital.filename}")
        
        # Importar serviço DAMA
        from services.dama_service import get_dama_service
        
        dama_service = get_dama_service()
        
        # Extrair itens usando OCR + Gemini
        itens = await dama_service.extrair_itens_edital(edital_content)
        
        logger.info(f"✅ [DAMA] Extraídos {len(itens)} itens do edital")
        
        return {
            "status": "sucesso",
            "itens": itens,
            "total": len(itens),
            "stats": {
                "arquivo": edital.filename,
                "itens_encontrados": len(itens)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao extrair itens: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dama/extrair-itens-filtrado")
async def extrair_itens_filtrado(
    edital: List[UploadFile] = File(..., description="PDF(s) do edital"),
    palavra_chave: str = Form(..., description="Palavra-chave para filtrar itens")
):
    """
    🔍 DAMA v73.0 - Extrair itens do edital FILTRADOS por palavra-chave
    
    Usa Gemini para:
    1. OCR do(s) PDF(s)
    2. Identificar apenas itens que contenham a palavra-chave (Ex: "Canabidiol")
    3. Retornar lista filtrada para o usuário precificar
    """
    try:
        # Concatenar conteúdo de todos os PDFs
        all_content = b''
        filenames = []
        
        for ed in edital:
            if not ed.filename.lower().endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"O arquivo {ed.filename} deve ser PDF")
            
            content = await ed.read()
            
            if len(content) > MAX_UPLOAD_SIZE:
                raise HTTPException(status_code=413, detail=f"Arquivo {ed.filename} muito grande (máx 15MB)")
            
            all_content += content
            filenames.append(ed.filename)
        
        logger.info(f"🔍 [DAMA v73.0] Extraindo itens com filtro '{palavra_chave}' de: {', '.join(filenames)}")
        
        # Importar serviço DAMA
        from services.dama_service import get_dama_service
        
        dama_service = get_dama_service()
        
        # Extrair itens usando OCR + Gemini COM filtro
        itens = await dama_service.extrair_itens_filtrado(all_content, palavra_chave)
        
        logger.info(f"✅ [DAMA v73.0] Extraídos {len(itens)} itens com '{palavra_chave}'")
        
        return {
            "status": "sucesso",
            "itens": itens,
            "total": len(itens),
            "palavra_chave": palavra_chave,
            "stats": {
                "arquivos": filenames,
                "itens_encontrados": len(itens),
                "filtro_aplicado": palavra_chave
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao extrair itens filtrados: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# Include the router in the main app
app.include_router(api_router)

app.include_router(auth_router)
app.include_router(users_router, dependencies=[Depends(require_super_admin)])

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== EVENTOS DE LIFECYCLE ====================

@app.on_event("startup")
async def startup_db_indexes():
    """
    P0: Criar índices MongoDB no startup
    Garante otimização de performance em todas as queries
    """
    try:
        logger.info("🚀 Iniciando criação de índices MongoDB...")
        collection = db.licitacoes
        
        # 1. Índice de texto em medicamento (busca full-text)
        await collection.create_index(
            [("medicamento", "text"), ("objeto", "text")],
            name="idx_medicamento_text",
            default_language="portuguese"
        )
        
        # 2. Índice em estado (filtro comum)
        await collection.create_index("estado", name="idx_estado")
        
        # 3. Índice em status (filtro comum)
        await collection.create_index("status", name="idx_status")
        
        # 4. Índice em data_final (ordenação por urgência)
        await collection.create_index("data_final", name="idx_data_final")
        
        # 5. Índice em modalidade (filtro)
        await collection.create_index("modalidade", name="idx_modalidade")
        
        # 6. Índice em esfera (filtro)
        await collection.create_index("esfera", name="idx_esfera")
        
        # 7. Índice composto: medicamento + estado
        await collection.create_index(
            [("medicamento", 1), ("estado", 1)],
            name="idx_medicamento_estado"
        )
        
        # 8. Índice composto: fonte + is_mock
        await collection.create_index(
            [("fonte", 1), ("is_mock", 1)],
            name="idx_fonte_mock"
        )
        
        # 9. Índice em id (unique)
        await collection.create_index("id", name="idx_id", unique=True)
        
        logger.info("✅ Índices MongoDB criados com sucesso!")
        
    except Exception as e:
        logger.error(f"❌ Erro ao criar índices: {str(e)}")


@app.on_event("startup")
async def startup_sync_service():
    """
    P0: Inicializa o SyncService para arquitetura Local-First
    Deve rodar ANTES do scheduler para garantir que a instância está disponível
    """
    global _sync_service_instance
    try:
        logger.info("🔄 Inicializando SyncService (Local-First Architecture)...")
        
        # Importar cliente PNCP
        from scrapers.pncp_api_oficial import PNCPApiOficial
        pncp_client = PNCPApiOficial()
        
        # Inicializar SyncService
        _sync_service_instance = init_sync_service(db, pncp_client)
        
        # Configurar índices da collection 'editais'
        await _sync_service_instance.setup_indexes()
        
        logger.info("✅ SyncService inicializado com sucesso!")
        
    except Exception as e:
        logger.error(f"❌ Erro ao inicializar SyncService: {str(e)}")
        _sync_service_instance = None


@app.on_event("startup")
async def startup_multi_source_sync():
    """
    P1: Inicializa o serviço de sincronização multi-fonte
    ComprasNet, TCE-SP, MG, PR, GO
    """
    global _multi_source_sync_instance, _normalizador_instance
    try:
        logger.info("🌐 Inicializando MultiSourceSync...")
        
        from services.normalizador_generico import get_normalizador_generico
        from services.multi_source_sync import get_multi_source_sync
        
        _normalizador_instance = get_normalizador_generico(db)
        _multi_source_sync_instance = get_multi_source_sync(db, _normalizador_instance)
        
        logger.info("✅ MultiSourceSync inicializado com sucesso!")
        
    except Exception as e:
        logger.error(f"❌ Erro ao inicializar MultiSourceSync: {str(e)}")
        _multi_source_sync_instance = None
        _normalizador_instance = None


@app.on_event("startup")
async def startup_scheduler():
    """
    Inicializa o APScheduler para verificação automática de alertas
    Agora inclui jobs de sincronização PNCP e multi-fonte
    """
    try:
        logger.info("⏰ Inicializando APScheduler...")
        
        # Inicializar scheduler com todos os serviços
        success = init_scheduler(
            db=db,
            notificacao_service=notificacao_service,
            scraper_service=scraper_service,
            sync_service=_sync_service_instance,
            multi_source_sync=_multi_source_sync_instance if '_multi_source_sync_instance' in globals() else None,
            normalizador=_normalizador_instance if '_normalizador_instance' in globals() else None
        )
        
        if success:
            logger.info("✅ APScheduler inicializado com sucesso!")
        else:
            logger.warning("⚠️ APScheduler não pôde ser inicializado")
            
    except Exception as e:
        logger.error(f"❌ Erro ao inicializar scheduler: {str(e)}")



@app.on_event("startup")
async def startup_independente():
    """Motor 100% independente — PNCP + Compras.gov.br."""
    logger.info("Motor GSM v78.0: 100% INDEPENDENTE (PNCP + Compras.gov.br)")

    # Radar Farmaceutico: seed lista de interesse
    try:
        from services.radar_farmaceutico_service import get_radar_farmaceutico_service
        svc = get_radar_farmaceutico_service(db)
        result = await svc.seed_lista_interesse()
        logger.info(f"Radar Farmaceutico seed: {result}")
    except Exception as e:
        logger.error(f"Radar Farmaceutico seed erro: {e}")



@app.on_event("shutdown")
async def shutdown_db_client():
    # Encerrar scheduler
    shutdown_scheduler()
    logger.info("🛑 Scheduler encerrado")
    
    # Fechar conexão com MongoDB
    client.close()
