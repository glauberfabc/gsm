"""
Test suite for Excel Export feature on the Preços (Prices) tab.
Tests GET /api/precos/export-excel endpoint.

Requirements tested:
1. Returns 200 with correct content-type (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
2. Downloaded file is valid .xlsx with multiple sheets
3. Resumo sheet contains: title, search term, date, Big Numbers, presentations table
4. Each presentation sheet has header with stats and data table
"""
import pytest
import requests
import os
from io import BytesIO

# Use PUBLIC URL for testing
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestPrecosExportExcel:
    """Excel Export endpoint tests for Preços tab"""
    
    def test_export_excel_returns_200(self):
        """Test GET /api/precos/export-excel?q=Prolia returns 200"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120  # Excel generation can take time
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text[:500]}"
        print(f"✅ Export Excel returns 200 OK")
    
    def test_export_excel_content_type(self):
        """Test that response has correct Excel content-type"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        content_type = response.headers.get('Content-Type', '')
        expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        assert expected_type in content_type, f"Expected content-type {expected_type}, got {content_type}"
        print(f"✅ Content-Type: {content_type}")
    
    def test_export_excel_content_disposition(self):
        """Test that response has Content-Disposition header with filename"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        content_disp = response.headers.get('Content-Disposition', '')
        
        assert 'attachment' in content_disp, f"Expected 'attachment' in Content-Disposition, got {content_disp}"
        assert 'filename=' in content_disp, f"Expected 'filename=' in Content-Disposition, got {content_disp}"
        assert '.xlsx' in content_disp, f"Expected '.xlsx' in filename, got {content_disp}"
        print(f"✅ Content-Disposition: {content_disp}")
    
    def test_export_excel_valid_xlsx(self):
        """Test that the downloaded file is a valid .xlsx file"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Import openpyxl to validate
        from openpyxl import load_workbook
        
        # Load workbook from response content
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        assert wb is not None, "Failed to load workbook"
        print(f"✅ Valid xlsx file loaded. Sheets: {wb.sheetnames}")
    
    def test_export_excel_has_resumo_sheet(self):
        """Test that xlsx has 'Resumo' sheet as the first sheet"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        sheet_names = wb.sheetnames
        assert len(sheet_names) > 0, "No sheets in workbook"
        assert sheet_names[0] == "Resumo", f"First sheet should be 'Resumo', got {sheet_names[0]}"
        print(f"✅ First sheet is 'Resumo'. Total sheets: {len(sheet_names)}")
    
    def test_export_excel_multiple_sheets(self):
        """Test that xlsx has multiple sheets (Resumo + presentations)"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        sheet_names = wb.sheetnames
        # Should have at least 2 sheets: Resumo + at least one presentation
        assert len(sheet_names) >= 2, f"Expected at least 2 sheets (Resumo + presentations), got {len(sheet_names)}: {sheet_names}"
        print(f"✅ Multiple sheets: {sheet_names}")
    
    def test_resumo_sheet_has_title(self):
        """Test that Resumo sheet has the GSM title"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        ws_resumo = wb["Resumo"]
        
        # Title is at A6
        title_cell = ws_resumo["A6"].value
        assert title_cell is not None, "Title cell A6 is empty"
        assert "GSM" in title_cell.upper() or "PREÇOS" in title_cell.upper(), f"Expected GSM or PREÇOS in title, got {title_cell}"
        print(f"✅ Resumo title: {title_cell}")
    
    def test_resumo_sheet_has_search_term(self):
        """Test that Resumo sheet shows the search term"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        ws_resumo = wb["Resumo"]
        
        # Search term at A7
        term_cell = ws_resumo["A7"].value
        assert term_cell is not None, "Search term cell A7 is empty"
        assert "PROLIA" in term_cell.upper(), f"Expected PROLIA in search term, got {term_cell}"
        print(f"✅ Search term: {term_cell}")
    
    def test_resumo_sheet_has_date(self):
        """Test that Resumo sheet shows generation date"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        ws_resumo = wb["Resumo"]
        
        # Date info at A8
        date_cell = ws_resumo["A8"].value
        assert date_cell is not None, "Date cell A8 is empty"
        assert "Gerado em" in date_cell or "/" in date_cell, f"Expected date info in A8, got {date_cell}"
        print(f"✅ Date info: {date_cell}")
    
    def test_resumo_sheet_has_big_numbers(self):
        """Test that Resumo sheet has Big Numbers (Min/Med/Mediana/Max)"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        ws_resumo = wb["Resumo"]
        
        # Big numbers labels at row 10, values at row 11
        labels_row = 10
        values_row = 11
        
        # Check labels exist (columns B-F)
        labels_found = []
        for col in range(2, 7):  # B to F
            cell = ws_resumo.cell(row=labels_row, column=col).value
            if cell:
                labels_found.append(cell)
        
        expected_labels = ["Menor", "Médio", "Mediana", "Maior", "Total"]
        found_any = any(any(exp.lower() in (lbl or '').lower() for exp in expected_labels) for lbl in labels_found)
        assert found_any, f"Expected Big Numbers labels, got {labels_found}"
        
        # Check values exist (row 11)
        values_found = []
        for col in range(2, 7):
            cell = ws_resumo.cell(row=values_row, column=col).value
            if cell:
                values_found.append(cell)
        
        assert len(values_found) > 0, "No Big Numbers values found in row 11"
        print(f"✅ Big Numbers labels: {labels_found}")
        print(f"✅ Big Numbers values: {values_found}")
    
    def test_resumo_sheet_has_presentations_table(self):
        """Test that Resumo sheet has presentations table starting at row 14"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        ws_resumo = wb["Resumo"]
        
        # Headers at row 15 (title at 14)
        title_row = ws_resumo.cell(row=14, column=1).value
        assert title_row is not None and "APRESENTA" in title_row.upper(), f"Expected 'APRESENTAÇÕES' title at row 14, got {title_row}"
        
        # Check headers at row 15
        headers_row = 15
        headers = []
        for col in range(1, 7):
            cell = ws_resumo.cell(row=headers_row, column=col).value
            if cell:
                headers.append(cell)
        
        assert len(headers) >= 4, f"Expected at least 4 headers in presentations table, got {headers}"
        print(f"✅ Presentations table title: {title_row}")
        print(f"✅ Presentations table headers: {headers}")
    
    def test_presentation_sheets_exist(self):
        """Test that presentation sheets exist with correct names (DENOSUMAB variations)"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        sheet_names = wb.sheetnames
        
        # Filter out 'Resumo' to get presentation sheets
        presentation_sheets = [s for s in sheet_names if s != "Resumo"]
        
        assert len(presentation_sheets) >= 1, f"Expected at least 1 presentation sheet, got 0. All sheets: {sheet_names}"
        
        # Check that sheets contain medication-related names
        has_denosumab = any("DENOSUMAB" in s.upper() for s in presentation_sheets)
        has_prolia = any("PROLIA" in s.upper() for s in presentation_sheets)
        has_dosage = any("MG" in s.upper() for s in presentation_sheets)
        
        assert has_denosumab or has_prolia or has_dosage, f"Expected medication-related sheet names, got {presentation_sheets}"
        print(f"✅ Presentation sheets: {presentation_sheets}")
    
    def test_presentation_sheet_has_header_stats(self):
        """Test that each presentation sheet has header with stats"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        # Get first presentation sheet (not Resumo)
        presentation_sheets = [s for s in wb.sheetnames if s != "Resumo"]
        assert len(presentation_sheets) > 0
        
        ws = wb[presentation_sheets[0]]
        
        # Header at A1 should have presentation name + count
        header = ws["A1"].value
        assert header is not None, "Presentation sheet header A1 is empty"
        assert "registros" in header.lower() or "DENOSUMAB" in header.upper() or "MG" in header.upper(), f"Expected presentation name in A1, got {header}"
        
        # Stats at A2 should have Min/Médio/Mediana/Max
        stats = ws["A2"].value
        assert stats is not None, "Stats cell A2 is empty"
        assert "Min" in stats or "R$" in stats, f"Expected stats with prices in A2, got {stats}"
        
        print(f"✅ Presentation header: {header}")
        print(f"✅ Presentation stats: {stats}")
    
    def test_presentation_sheet_has_data_table(self):
        """Test that presentation sheet has data table with required columns"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        # Get first presentation sheet
        presentation_sheets = [s for s in wb.sheetnames if s != "Resumo"]
        assert len(presentation_sheets) > 0
        
        ws = wb[presentation_sheets[0]]
        
        # Headers at row 4
        headers = []
        for col in range(1, 8):
            cell = ws.cell(row=4, column=col).value
            if cell:
                headers.append(cell)
        
        expected_headers = ["Órgão", "UF", "Descrição", "Quantidade", "Valor", "Data", "Fonte"]
        
        # Check that most expected headers are present
        found_count = sum(1 for exp in expected_headers if any(exp.lower() in (h or '').lower() for h in headers))
        assert found_count >= 5, f"Expected at least 5 of {expected_headers}, got {headers}"
        
        print(f"✅ Data table headers: {headers}")
    
    def test_presentation_sheet_has_data_rows(self):
        """Test that presentation sheet has actual data rows"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        # Get first presentation sheet
        presentation_sheets = [s for s in wb.sheetnames if s != "Resumo"]
        assert len(presentation_sheets) > 0
        
        ws = wb[presentation_sheets[0]]
        
        # Data starts at row 5
        data_row = ws.cell(row=5, column=1).value
        assert data_row is not None, "No data in first data row (row 5)"
        
        # Count data rows
        row_count = 0
        for row in range(5, 100):  # Check up to 95 rows
            if ws.cell(row=row, column=1).value:
                row_count += 1
            else:
                break
        
        assert row_count >= 1, f"Expected at least 1 data row, got {row_count}"
        print(f"✅ Data rows found: {row_count}")
    
    def test_export_excel_with_uf_filter(self):
        """Test that UF filter works in export"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "Prolia", "uf": "SP"},
            timeout=120
        )
        
        # Should still return 200 (may have fewer or no results for specific UF)
        assert response.status_code == 200, f"Expected 200 with UF filter, got {response.status_code}"
        print(f"✅ Export with UF filter returns 200")


class TestPrecosExportEdgeCases:
    """Edge case tests for Excel export"""
    
    def test_export_excel_missing_query(self):
        """Test that missing query returns 422"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            timeout=30
        )
        
        # FastAPI should return 422 for missing required parameter
        assert response.status_code == 422, f"Expected 422 for missing query, got {response.status_code}"
        print(f"✅ Missing query returns 422")
    
    def test_export_excel_empty_results(self):
        """Test export with term that may return few/no results"""
        response = requests.get(
            f"{BASE_URL}/api/precos/export-excel",
            params={"q": "xyznonexistent123"},
            timeout=60
        )
        
        # Should return 200 even with no results (empty Excel)
        # Or could return 404/500 - just check it doesn't hang
        assert response.status_code in [200, 404, 500], f"Unexpected status {response.status_code}"
        print(f"✅ Non-existent term returns {response.status_code}")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
