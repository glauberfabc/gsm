"""
Test suite for Search Export feature to Excel.
Tests GET /api/search/export-excel endpoint.
"""
import pytest
import requests
import os
from io import BytesIO
from openpyxl import load_workbook

# Use PUBLIC URL for testing or localhost default
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:8000').rstrip('/')

class TestSearchExportExcel:
    """Excel Export endpoint tests for Search results"""
    
    def test_export_excel_returns_200(self):
        """Test GET /api/search/export-excel?medicamento=Insulina returns 200"""
        response = requests.get(
            f"{BASE_URL}/api/search/export-excel",
            params={"medicamento": "Insulina"},
            timeout=120
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text[:500]}"
        print(f"✅ Export Excel returns 200 OK")
    
    def test_export_excel_content_type(self):
        """Test that response has correct Excel content-type"""
        response = requests.get(
            f"{BASE_URL}/api/search/export-excel",
            params={"medicamento": "Insulina"},
            timeout=120
        )
        
        assert response.status_code == 200
        content_type = response.headers.get('Content-Type', '')
        expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        assert expected_type in content_type, f"Expected content-type {expected_type}, got {content_type}"
        print(f"✅ Content-Type correct: {content_type}")
    
    def test_export_excel_valid_xlsx(self):
        """Test that the downloaded file is a valid .xlsx file and has the correct title"""
        response = requests.get(
            f"{BASE_URL}/api/search/export-excel",
            params={"medicamento": "Insulina"},
            timeout=120
        )
        
        assert response.status_code == 200
        
        # Load workbook from response content
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        assert wb is not None, "Failed to load workbook"
        ws = wb.active
        assert ws.title == "Resultados GSM", f"Expected sheet title 'Resultados GSM', got {ws.title}"
        
        # Check header title in A6
        header_title = ws["A6"].value
        assert "GSM" in header_title.upper(), f"Expected GSM in header, got {header_title}"
        
        # Check if table headers exist at row 11
        first_header = ws["A11"].value
        assert "MEDICAMENTO" in first_header.upper(), f"Expected 'Medicamento' header at A11, got {first_header}"
        
        print(f"✅ Valid xlsx file loaded with correct structure")

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
